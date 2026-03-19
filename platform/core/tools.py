"""
Janovum Platform — Unified Tools Registry
The orchestrator's hands. Every tool the AI can call lives here.

This file:
  1. Collects tool definitions from all modules (browser, search, files)
  2. Adds built-in tools (email, image gen, code exec, shell, voice, scheduling)
  3. Provides get_all_tools() for tool definitions and execute_tool() for execution
  4. One function the engine's agent_loop() calls — routes to the right module

Usage:
  from core.tools import get_all_tools, execute_tool
  result = agent_loop(messages, system_prompt, get_all_tools(), execute_tool)
"""

import json
import os
import sys
import time
import subprocess
import smtplib
import imaplib
import email as email_lib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

PLATFORM_DIR = Path(__file__).parent.parent

# ── Import module tools ──
sys.path.insert(0, str(PLATFORM_DIR))
from modules.web_search import TOOLS as WEB_TOOLS, execute_tool as web_execute
from modules.browser_agent import TOOLS as BROWSER_TOOLS, execute_tool as browser_execute
from modules.file_manager import TOOLS as FILE_TOOLS, execute_tool as file_execute


# ══════════════════════════════════════════════════════════════════════
# BUILT-IN TOOLS — capabilities that don't have their own module yet
# ══════════════════════════════════════════════════════════════════════

BUILTIN_TOOLS = [
    # ── Email ──
    {
        "name": "email_send",
        "description": "Send an email via SMTP",
        "input_schema": {
            "type": "object",
            "properties": {
                "to": {"type": "string", "description": "Recipient email address"},
                "subject": {"type": "string", "description": "Email subject line"},
                "body": {"type": "string", "description": "Email body text"},
                "smtp_server": {"type": "string", "description": "SMTP server (default: smtp.gmail.com)", "default": "smtp.gmail.com"},
                "smtp_port": {"type": "integer", "description": "SMTP port (default: 587)", "default": 587},
            },
            "required": ["to", "subject", "body"]
        }
    },
    {
        "name": "email_read",
        "description": "Read recent emails from an inbox via IMAP",
        "input_schema": {
            "type": "object",
            "properties": {
                "folder": {"type": "string", "description": "Mailbox folder (default: INBOX)", "default": "INBOX"},
                "count": {"type": "integer", "description": "Number of recent emails to read (default: 5)", "default": 5},
                "unread_only": {"type": "boolean", "description": "Only return unread emails", "default": True},
                "imap_server": {"type": "string", "description": "IMAP server (default: imap.gmail.com)", "default": "imap.gmail.com"},
            },
            "required": []
        }
    },
    # ── Image Generation ──
    {
        "name": "image_generate",
        "description": "Generate an image from a text prompt using AI (Pollinations.ai, free, no API key)",
        "input_schema": {
            "type": "object",
            "properties": {
                "prompt": {"type": "string", "description": "Text description of the image to generate"},
                "width": {"type": "integer", "description": "Image width in pixels (default: 1024)", "default": 1024},
                "height": {"type": "integer", "description": "Image height in pixels (default: 1024)", "default": 1024},
                "save_as": {"type": "string", "description": "Filename to save as (default: auto-generated)"}
            },
            "required": ["prompt"]
        }
    },
    # ── Code Execution ──
    {
        "name": "code_execute",
        "description": "Execute Python code in a secure sandbox and return the output",
        "input_schema": {
            "type": "object",
            "properties": {
                "code": {"type": "string", "description": "Python code to execute"},
                "timeout": {"type": "integer", "description": "Max execution time in seconds (default: 30)", "default": 30}
            },
            "required": ["code"]
        }
    },
    {
        "name": "shell_execute",
        "description": "Execute a shell command and return the output",
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string", "description": "Shell command to execute"},
                "timeout": {"type": "integer", "description": "Max execution time in seconds (default: 30)", "default": 30}
            },
            "required": ["command"]
        }
    },
    # ── Text-to-Speech ──
    {
        "name": "voice_tts",
        "description": "Convert text to speech audio file using edge-tts (free, 400+ voices)",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to convert to speech"},
                "voice": {"type": "string", "description": "Voice ID (default: en-US-AriaNeural). Options: en-US-AriaNeural, en-US-GuyNeural, en-US-JennyNeural, en-US-DavisNeural, en-GB-SoniaNeural, en-GB-RyanNeural"},
                "save_as": {"type": "string", "description": "Output filename (default: auto-generated .mp3)"}
            },
            "required": ["text"]
        }
    },
    # ── Speech-to-Text ──
    {
        "name": "voice_stt",
        "description": "Transcribe audio file to text using Whisper (local, free)",
        "input_schema": {
            "type": "object",
            "properties": {
                "audio_path": {"type": "string", "description": "Path to the audio file to transcribe"}
            },
            "required": ["audio_path"]
        }
    },
    # ── HTTP Requests ──
    {
        "name": "http_request",
        "description": "Make an HTTP request (GET, POST, PUT, DELETE) to any API or URL",
        "input_schema": {
            "type": "object",
            "properties": {
                "method": {"type": "string", "description": "HTTP method: GET, POST, PUT, DELETE", "default": "GET"},
                "url": {"type": "string", "description": "The URL to request"},
                "headers": {"type": "object", "description": "Optional headers as key-value pairs"},
                "body": {"type": "string", "description": "Optional request body (for POST/PUT)"},
                "json_body": {"type": "object", "description": "Optional JSON body (for POST/PUT)"}
            },
            "required": ["url"]
        }
    },
    # ── Scheduling ──
    {
        "name": "schedule_task",
        "description": "Schedule a one-time or recurring task (saved to disk, survives restarts)",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Task name/identifier"},
                "action": {"type": "string", "description": "What to do: 'run_code', 'http_request', 'send_email'"},
                "action_data": {"type": "object", "description": "Data for the action (code, url, email details, etc.)"},
                "run_at": {"type": "string", "description": "ISO datetime for one-time execution (e.g. 2026-03-20T14:00:00)"},
                "interval_minutes": {"type": "integer", "description": "For recurring: run every N minutes"}
            },
            "required": ["name", "action", "action_data"]
        }
    },
    {
        "name": "schedule_list",
        "description": "List all scheduled tasks",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "schedule_cancel",
        "description": "Cancel a scheduled task by name",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Name of the task to cancel"}
            },
            "required": ["name"]
        }
    },
    # ── Memory (agent persistent context) ──
    {
        "name": "memory_read",
        "description": "Read the agent's persistent memory (context, history, notes, contacts)",
        "input_schema": {
            "type": "object",
            "properties": {
                "memory_type": {"type": "string", "description": "Type: context, history, notes, contacts", "default": "context"}
            }
        }
    },
    {
        "name": "memory_write",
        "description": "Save information to the agent's persistent memory",
        "input_schema": {
            "type": "object",
            "properties": {
                "memory_type": {"type": "string", "description": "Type: context, history, notes, contacts", "default": "notes"},
                "content": {"type": "string", "description": "Content to save"},
                "append": {"type": "boolean", "description": "Append to existing (true) or overwrite (false)", "default": True}
            },
            "required": ["content"]
        }
    },
    # ── Document Parsing ──
    {
        "name": "document_read_pdf",
        "description": "Extract text from a PDF file",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the PDF file"},
                "pages": {"type": "string", "description": "Page range: 'all', '1-5', '3' (default: all)", "default": "all"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "document_read_docx",
        "description": "Extract text from a Word document (.docx)",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the .docx file"}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "document_read_csv",
        "description": "Read and parse a CSV file, return as structured data",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to the CSV file"},
                "max_rows": {"type": "integer", "description": "Max rows to return (default: 100)", "default": 100}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "document_ocr",
        "description": "Extract text from an image using OCR (Tesseract)",
        "input_schema": {
            "type": "object",
            "properties": {
                "image_path": {"type": "string", "description": "Path to the image file"}
            },
            "required": ["image_path"]
        }
    },
    # ── Messaging (Telegram, Discord, SMS) ──
    {
        "name": "telegram_send",
        "description": "Send a message via Telegram bot",
        "input_schema": {
            "type": "object",
            "properties": {
                "chat_id": {"type": "string", "description": "Telegram chat ID to send to"},
                "text": {"type": "string", "description": "Message text"},
                "parse_mode": {"type": "string", "description": "Message format: HTML or Markdown", "default": "HTML"}
            },
            "required": ["chat_id", "text"]
        }
    },
    {
        "name": "discord_send",
        "description": "Send a message via Discord webhook",
        "input_schema": {
            "type": "object",
            "properties": {
                "webhook_url": {"type": "string", "description": "Discord webhook URL"},
                "content": {"type": "string", "description": "Message text"},
                "username": {"type": "string", "description": "Bot display name (optional)"}
            },
            "required": ["webhook_url", "content"]
        }
    },
    {
        "name": "sms_send",
        "description": "Send an SMS via Twilio",
        "input_schema": {
            "type": "object",
            "properties": {
                "to": {"type": "string", "description": "Phone number to send to (e.g. +1234567890)"},
                "body": {"type": "string", "description": "SMS message body"}
            },
            "required": ["to", "body"]
        }
    },
    # ── VPS Management ──
    {
        "name": "vps_execute",
        "description": "Execute a command on the VPS via SSH",
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string", "description": "Shell command to run on the VPS"},
                "timeout": {"type": "integer", "description": "Max seconds to wait (default: 30)", "default": 30}
            },
            "required": ["command"]
        }
    },
    {
        "name": "vps_upload",
        "description": "Upload a local file to the VPS via SCP",
        "input_schema": {
            "type": "object",
            "properties": {
                "local_path": {"type": "string", "description": "Local file path to upload"},
                "remote_path": {"type": "string", "description": "Destination path on VPS"}
            },
            "required": ["local_path", "remote_path"]
        }
    },
    {
        "name": "vps_download",
        "description": "Download a file from the VPS to local machine via SCP",
        "input_schema": {
            "type": "object",
            "properties": {
                "remote_path": {"type": "string", "description": "File path on VPS"},
                "local_path": {"type": "string", "description": "Local destination path"}
            },
            "required": ["remote_path", "local_path"]
        }
    },
    # ── Git / GitHub ──
    {
        "name": "git_execute",
        "description": "Run a git command in a specified directory",
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string", "description": "Git command (e.g. 'status', 'add -A', 'commit -m msg', 'push')"},
                "working_dir": {"type": "string", "description": "Directory to run git in (default: platform dir)"}
            },
            "required": ["command"]
        }
    },
    # ── Skills / Plugins ──
    {
        "name": "skill_list",
        "description": "List all available skills/plugins loaded in the toolkit",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "skill_load",
        "description": "Load and read a skill file (.md or .py) to learn how to perform a specific task",
        "input_schema": {
            "type": "object",
            "properties": {
                "skill_name": {"type": "string", "description": "Name of the skill to load (without extension)"}
            },
            "required": ["skill_name"]
        }
    },
    {
        "name": "skill_create",
        "description": "Create a new skill file (.md) that teaches the AI how to do a new task",
        "input_schema": {
            "type": "object",
            "properties": {
                "skill_name": {"type": "string", "description": "Name for the new skill"},
                "description": {"type": "string", "description": "One-line description of what the skill does"},
                "instructions": {"type": "string", "description": "Full instructions/steps for the skill in markdown"}
            },
            "required": ["skill_name", "description", "instructions"]
        }
    },
    # ── Client Management ──
    {
        "name": "client_list",
        "description": "List all clients configured in the toolkit",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "client_add",
        "description": "Add a new client to the toolkit with their configuration",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Unique client identifier (lowercase, no spaces)"},
                "business_name": {"type": "string", "description": "Client's business name"},
                "phone": {"type": "string", "description": "Client's phone number"},
                "email": {"type": "string", "description": "Client's email"},
                "services": {"type": "array", "items": {"type": "string"}, "description": "List of services the client offers"},
                "business_hours": {"type": "string", "description": "Business hours (e.g. 'Mon-Fri 9am-5pm')"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "client_start",
        "description": "Start a client's receptionist/agent service",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID to start"}
            },
            "required": ["client_id"]
        }
    },
    {
        "name": "client_stop",
        "description": "Stop a client's receptionist/agent service",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID to stop"}
            },
            "required": ["client_id"]
        }
    },
    # ── Knowledge / Learning ──
    {
        "name": "knowledge_save",
        "description": "Save a piece of knowledge to the persistent knowledge base (facts, how-tos, discoveries, API info, anything useful)",
        "input_schema": {
            "type": "object",
            "properties": {
                "topic": {"type": "string", "description": "Knowledge topic/category (e.g. 'api', 'how-to', 'fact', 'contact', 'business')"},
                "title": {"type": "string", "description": "Short title for this knowledge entry"},
                "content": {"type": "string", "description": "The knowledge content — be detailed"},
                "tags": {"type": "array", "items": {"type": "string"}, "description": "Tags for searchability"}
            },
            "required": ["topic", "title", "content"]
        }
    },
    {
        "name": "knowledge_search",
        "description": "Search the knowledge base for stored information",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Search query"},
                "topic": {"type": "string", "description": "Optional: filter by topic"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "knowledge_list",
        "description": "List all knowledge entries, optionally filtered by topic",
        "input_schema": {
            "type": "object",
            "properties": {
                "topic": {"type": "string", "description": "Optional topic filter"}
            }
        }
    },
    # ── Account / Signup Automation ──
    {
        "name": "browser_signup",
        "description": "Automate account signup on a website using Selenium (fill form, handle CAPTCHA, verify email)",
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "Signup page URL"},
                "fields": {"type": "object", "description": "Form fields to fill: {selector: value, ...}"},
                "submit_selector": {"type": "string", "description": "CSS selector for the submit button"},
                "email": {"type": "string", "description": "Email for signup (will auto-verify if IMAP configured)"},
                "headless": {"type": "boolean", "description": "Run headless (true) or visible (false)", "default": False}
            },
            "required": ["url", "fields"]
        }
    },
    {
        "name": "browser_login",
        "description": "Log into a website using Selenium and return the session cookies",
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "Login page URL"},
                "username_selector": {"type": "string", "description": "CSS selector for username field"},
                "password_selector": {"type": "string", "description": "CSS selector for password field"},
                "username": {"type": "string", "description": "Username/email"},
                "password": {"type": "string", "description": "Password"},
                "submit_selector": {"type": "string", "description": "CSS selector for login button"}
            },
            "required": ["url", "username", "password"]
        }
    },
    {
        "name": "browser_interactive",
        "description": "Run a multi-step browser session — navigate, click, type, screenshot, extract data in sequence",
        "input_schema": {
            "type": "object",
            "properties": {
                "steps": {
                    "type": "array",
                    "description": "List of browser actions to execute in order",
                    "items": {
                        "type": "object",
                        "properties": {
                            "action": {"type": "string", "description": "Action: navigate, click, type, screenshot, wait, extract_text, extract_links, scroll"},
                            "selector": {"type": "string", "description": "CSS selector (for click, type, extract_text)"},
                            "value": {"type": "string", "description": "Value (for type: text to type, for navigate: URL, for wait: seconds)"}
                        },
                        "required": ["action"]
                    }
                },
                "headless": {"type": "boolean", "description": "Run headless or visible", "default": True}
            },
            "required": ["steps"]
        }
    },
    # ── Data Intelligence / Scraping ──
    {
        "name": "scrape_google",
        "description": "Scrape Google search results for a query (uses Selenium to bypass blocks)",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Google search query"},
                "num_results": {"type": "integer", "description": "Number of results to return", "default": 10}
            },
            "required": ["query"]
        }
    },
    {
        "name": "scrape_social",
        "description": "Scrape public social media profiles/posts (Twitter/X, LinkedIn, Instagram, Reddit)",
        "input_schema": {
            "type": "object",
            "properties": {
                "platform": {"type": "string", "description": "Platform: twitter, linkedin, instagram, reddit"},
                "target": {"type": "string", "description": "Username, URL, or search query"},
                "action": {"type": "string", "description": "Action: profile, posts, search", "default": "profile"}
            },
            "required": ["platform", "target"]
        }
    },
    {
        "name": "scrape_business_info",
        "description": "Look up business information — Google Maps, Yelp, website, contacts",
        "input_schema": {
            "type": "object",
            "properties": {
                "business_name": {"type": "string", "description": "Business name to look up"},
                "location": {"type": "string", "description": "City/state/zip (optional)"}
            },
            "required": ["business_name"]
        }
    },
    # ── Lead Generation ──
    {
        "name": "lead_find",
        "description": "Find potential business leads — scrape directories, Google Maps, Yelp for businesses matching criteria",
        "input_schema": {
            "type": "object",
            "properties": {
                "industry": {"type": "string", "description": "Industry type (e.g. 'dentist', 'plumber', 'restaurant')"},
                "location": {"type": "string", "description": "City, state, or zip code"},
                "max_results": {"type": "integer", "description": "Max leads to find", "default": 20}
            },
            "required": ["industry", "location"]
        }
    },
    {
        "name": "lead_enrich",
        "description": "Enrich a lead with more data — find email, phone, website, social media, reviews",
        "input_schema": {
            "type": "object",
            "properties": {
                "business_name": {"type": "string", "description": "Business name"},
                "location": {"type": "string", "description": "Business location"},
                "website": {"type": "string", "description": "Business website (if known)"}
            },
            "required": ["business_name"]
        }
    },
    # ── Notifications / Alerts ──
    {
        "name": "notify",
        "description": "Send a notification via the best available channel (Telegram > Discord > Email > SMS)",
        "input_schema": {
            "type": "object",
            "properties": {
                "message": {"type": "string", "description": "Notification message"},
                "priority": {"type": "string", "description": "Priority: low, normal, high, urgent", "default": "normal"},
                "channel": {"type": "string", "description": "Force specific channel: telegram, discord, email, sms (optional)"}
            },
            "required": ["message"]
        }
    },
    # ── JSON/Data Processing ──
    {
        "name": "json_transform",
        "description": "Transform, filter, or restructure JSON data using a Python expression",
        "input_schema": {
            "type": "object",
            "properties": {
                "data": {"type": "object", "description": "The JSON data to transform"},
                "expression": {"type": "string", "description": "Python expression to transform the data (variable 'data' holds the input)"}
            },
            "required": ["data", "expression"]
        }
    },
    # ── QR Code ──
    {
        "name": "qr_generate",
        "description": "Generate a QR code image from text or URL",
        "input_schema": {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "Text or URL to encode in QR code"},
                "save_as": {"type": "string", "description": "Filename to save as (default: auto)"}
            },
            "required": ["content"]
        }
    },
    # ── Screenshot / Screen ──
    {
        "name": "screenshot_desktop",
        "description": "Take a screenshot of the local desktop screen",
        "input_schema": {
            "type": "object",
            "properties": {
                "save_as": {"type": "string", "description": "Filename to save as (default: auto)"}
            }
        }
    },
    # ── PDF Generation ──
    {
        "name": "pdf_create",
        "description": "Create a PDF document from text/HTML content",
        "input_schema": {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "Text or HTML content for the PDF"},
                "title": {"type": "string", "description": "PDF title"},
                "save_as": {"type": "string", "description": "Filename to save as"}
            },
            "required": ["content"]
        }
    },
    # ── Clipboard ──
    {
        "name": "clipboard_read",
        "description": "Read the current clipboard contents",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "clipboard_write",
        "description": "Write text to the clipboard",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to copy to clipboard"}
            },
            "required": ["text"]
        }
    },
    # ── Process Management ──
    {
        "name": "process_list",
        "description": "List running processes on this machine",
        "input_schema": {
            "type": "object",
            "properties": {
                "filter": {"type": "string", "description": "Optional: filter by process name"}
            }
        }
    },
    {
        "name": "process_kill",
        "description": "Kill a process by name or PID",
        "input_schema": {
            "type": "object",
            "properties": {
                "target": {"type": "string", "description": "Process name or PID to kill"}
            },
            "required": ["target"]
        }
    },
    # ── Environment / System Info ──
    {
        "name": "system_info",
        "description": "Get system information — OS, CPU, RAM, disk, Python version, installed packages",
        "input_schema": {
            "type": "object",
            "properties": {
                "detail": {"type": "string", "description": "What info: 'all', 'os', 'cpu', 'ram', 'disk', 'python', 'packages'", "default": "all"}
            }
        }
    },
    # ── Wait / Delay ──
    {
        "name": "wait",
        "description": "Wait/sleep for a specified number of seconds (useful in multi-step workflows)",
        "input_schema": {
            "type": "object",
            "properties": {
                "seconds": {"type": "number", "description": "Seconds to wait"}
            },
            "required": ["seconds"]
        }
    },
    # ── Video / Camera ──
    {
        "name": "camera_capture",
        "description": "Capture a photo from the webcam/camera",
        "input_schema": {
            "type": "object",
            "properties": {
                "save_as": {"type": "string", "description": "Filename to save as (default: auto)"},
                "camera_index": {"type": "integer", "description": "Camera index (default: 0)", "default": 0}
            }
        }
    },
    {
        "name": "video_record",
        "description": "Record video from webcam for a specified duration",
        "input_schema": {
            "type": "object",
            "properties": {
                "duration": {"type": "integer", "description": "Recording duration in seconds"},
                "save_as": {"type": "string", "description": "Filename to save as"},
                "camera_index": {"type": "integer", "description": "Camera index (default: 0)", "default": 0}
            },
            "required": ["duration"]
        }
    },
    {
        "name": "video_stream_start",
        "description": "Start a live video stream from webcam — frames saved to disk for real-time viewing",
        "input_schema": {
            "type": "object",
            "properties": {
                "fps": {"type": "integer", "description": "Frames per second to capture (default: 2)", "default": 2},
                "camera_index": {"type": "integer", "description": "Camera index (default: 0)", "default": 0},
                "stream_id": {"type": "string", "description": "Unique stream identifier"}
            },
            "required": ["stream_id"]
        }
    },
    {
        "name": "video_stream_stop",
        "description": "Stop a live video stream",
        "input_schema": {
            "type": "object",
            "properties": {
                "stream_id": {"type": "string", "description": "Stream ID to stop"}
            },
            "required": ["stream_id"]
        }
    },
    {
        "name": "screen_record",
        "description": "Record the desktop screen for a specified duration",
        "input_schema": {
            "type": "object",
            "properties": {
                "duration": {"type": "integer", "description": "Recording duration in seconds"},
                "save_as": {"type": "string", "description": "Filename to save as"},
                "fps": {"type": "integer", "description": "Frames per second (default: 10)", "default": 10}
            },
            "required": ["duration"]
        }
    },
    # ── Audio Recording ──
    {
        "name": "audio_record",
        "description": "Record audio from microphone for a specified duration",
        "input_schema": {
            "type": "object",
            "properties": {
                "duration": {"type": "integer", "description": "Recording duration in seconds"},
                "save_as": {"type": "string", "description": "Filename to save as (default: auto .wav)"}
            },
            "required": ["duration"]
        }
    },
    {
        "name": "audio_play",
        "description": "Play an audio file",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to audio file to play"}
            },
            "required": ["file_path"]
        }
    },
    # ── Compression ──
    {
        "name": "zip_create",
        "description": "Create a ZIP archive from files or a directory",
        "input_schema": {
            "type": "object",
            "properties": {
                "paths": {"type": "array", "items": {"type": "string"}, "description": "List of file/directory paths to compress"},
                "save_as": {"type": "string", "description": "Output ZIP filename"}
            },
            "required": ["paths", "save_as"]
        }
    },
    {
        "name": "zip_extract",
        "description": "Extract a ZIP archive",
        "input_schema": {
            "type": "object",
            "properties": {
                "zip_path": {"type": "string", "description": "Path to ZIP file"},
                "extract_to": {"type": "string", "description": "Directory to extract to"}
            },
            "required": ["zip_path", "extract_to"]
        }
    },
    # ── Translation ──
    {
        "name": "translate",
        "description": "Translate text between languages (free via MyMemory API or LibreTranslate)",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to translate"},
                "from_lang": {"type": "string", "description": "Source language code (e.g. 'en')", "default": "en"},
                "to_lang": {"type": "string", "description": "Target language code (e.g. 'es', 'fr', 'de', 'zh')"}
            },
            "required": ["text", "to_lang"]
        }
    },
    # ── Weather ──
    {
        "name": "weather_get",
        "description": "Get current weather for a location (Open-Meteo, free, no key)",
        "input_schema": {
            "type": "object",
            "properties": {
                "location": {"type": "string", "description": "City name or 'lat,lon'"}
            },
            "required": ["location"]
        }
    },
    # ── News ──
    {
        "name": "news_get",
        "description": "Get latest news headlines (Google News RSS, free)",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "News topic to search for (optional)"},
                "count": {"type": "integer", "description": "Number of articles (default: 10)", "default": 10}
            }
        }
    },
    # ── Finance ──
    {
        "name": "stock_price",
        "description": "Get stock/crypto price (via free APIs)",
        "input_schema": {
            "type": "object",
            "properties": {
                "symbol": {"type": "string", "description": "Stock ticker (e.g. AAPL, TSLA) or crypto (e.g. bitcoin)"}
            },
            "required": ["symbol"]
        }
    },
    # ── Network / DNS / Whois ──
    {
        "name": "whois_lookup",
        "description": "Look up WHOIS information for a domain",
        "input_schema": {
            "type": "object",
            "properties": {
                "domain": {"type": "string", "description": "Domain name to look up"}
            },
            "required": ["domain"]
        }
    },
    {
        "name": "dns_lookup",
        "description": "Look up DNS records for a domain",
        "input_schema": {
            "type": "object",
            "properties": {
                "domain": {"type": "string", "description": "Domain name"},
                "record_type": {"type": "string", "description": "Record type: A, AAAA, MX, TXT, CNAME, NS", "default": "A"}
            },
            "required": ["domain"]
        }
    },
    # ── Encoding / Hashing / Crypto ──
    {
        "name": "hash_text",
        "description": "Hash text using MD5, SHA1, SHA256, or SHA512",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to hash"},
                "algorithm": {"type": "string", "description": "Algorithm: md5, sha1, sha256, sha512", "default": "sha256"}
            },
            "required": ["text"]
        }
    },
    {
        "name": "encrypt_text",
        "description": "Encrypt text using Fernet symmetric encryption",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to encrypt"},
                "key": {"type": "string", "description": "Encryption key (if empty, generates a new one)"}
            },
            "required": ["text"]
        }
    },
    {
        "name": "decrypt_text",
        "description": "Decrypt text using Fernet symmetric encryption",
        "input_schema": {
            "type": "object",
            "properties": {
                "encrypted": {"type": "string", "description": "Encrypted text"},
                "key": {"type": "string", "description": "Encryption key"}
            },
            "required": ["encrypted", "key"]
        }
    },
    {
        "name": "base64_encode",
        "description": "Encode text or file to Base64",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to encode (or file path if is_file=true)"},
                "is_file": {"type": "boolean", "description": "If true, treat 'text' as file path", "default": False}
            },
            "required": ["text"]
        }
    },
    {
        "name": "base64_decode",
        "description": "Decode Base64 to text",
        "input_schema": {
            "type": "object",
            "properties": {
                "encoded": {"type": "string", "description": "Base64 encoded string"}
            },
            "required": ["encoded"]
        }
    },
    # ── Image Editing (Pillow) ──
    {
        "name": "image_resize",
        "description": "Resize an image to specified dimensions",
        "input_schema": {
            "type": "object",
            "properties": {
                "image_path": {"type": "string", "description": "Path to the image"},
                "width": {"type": "integer", "description": "New width"},
                "height": {"type": "integer", "description": "New height"},
                "save_as": {"type": "string", "description": "Output filename (default: overwrites)"}
            },
            "required": ["image_path", "width", "height"]
        }
    },
    {
        "name": "image_text_overlay",
        "description": "Add text overlay to an image (memes, watermarks, captions)",
        "input_schema": {
            "type": "object",
            "properties": {
                "image_path": {"type": "string", "description": "Path to the image"},
                "text": {"type": "string", "description": "Text to add"},
                "position": {"type": "string", "description": "Position: top, bottom, center", "default": "bottom"},
                "font_size": {"type": "integer", "description": "Font size (default: 36)", "default": 36},
                "color": {"type": "string", "description": "Text color (default: white)", "default": "white"},
                "save_as": {"type": "string", "description": "Output filename"}
            },
            "required": ["image_path", "text"]
        }
    },
    {
        "name": "image_convert",
        "description": "Convert image between formats (PNG, JPG, WEBP, BMP, GIF)",
        "input_schema": {
            "type": "object",
            "properties": {
                "image_path": {"type": "string", "description": "Path to the source image"},
                "format": {"type": "string", "description": "Target format: png, jpg, webp, bmp, gif"},
                "save_as": {"type": "string", "description": "Output filename"}
            },
            "required": ["image_path", "format"]
        }
    },
    # ── AI / LLM (multi-model) ──
    {
        "name": "ai_ask",
        "description": "Ask any AI model a question — routes to cheapest available (Groq free → Ollama local → Claude)",
        "input_schema": {
            "type": "object",
            "properties": {
                "prompt": {"type": "string", "description": "The prompt/question to ask"},
                "model": {"type": "string", "description": "Force model: 'haiku', 'sonnet', 'opus', 'groq', 'ollama', 'auto'", "default": "auto"},
                "system": {"type": "string", "description": "Optional system prompt"}
            },
            "required": ["prompt"]
        }
    },
    {
        "name": "ai_summarize",
        "description": "Summarize text, a file, or a URL using AI",
        "input_schema": {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "Text to summarize, OR a file path, OR a URL"},
                "style": {"type": "string", "description": "Summary style: brief, detailed, bullet_points, executive", "default": "brief"},
                "max_length": {"type": "integer", "description": "Max words in summary", "default": 200}
            },
            "required": ["content"]
        }
    },
    {
        "name": "ai_classify",
        "description": "Classify text into categories using AI",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to classify"},
                "categories": {"type": "array", "items": {"type": "string"}, "description": "List of possible categories"}
            },
            "required": ["text", "categories"]
        }
    },
    {
        "name": "ai_extract",
        "description": "Extract structured data from unstructured text using AI (names, dates, emails, addresses, etc.)",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to extract from"},
                "fields": {"type": "array", "items": {"type": "string"}, "description": "Fields to extract (e.g. ['name', 'email', 'phone', 'address'])"}
            },
            "required": ["text", "fields"]
        }
    },
    {
        "name": "ai_generate_code",
        "description": "Generate code in any language using AI, then optionally execute it",
        "input_schema": {
            "type": "object",
            "properties": {
                "description": {"type": "string", "description": "What the code should do"},
                "language": {"type": "string", "description": "Programming language (python, javascript, html, etc.)", "default": "python"},
                "execute": {"type": "boolean", "description": "Execute the generated code immediately (Python only)", "default": False}
            },
            "required": ["description"]
        }
    },
    # ── Workflow / Multi-Step Automation ──
    {
        "name": "workflow_create",
        "description": "Create a multi-step automated workflow (chain of tool calls)",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Workflow name"},
                "description": {"type": "string", "description": "What this workflow does"},
                "steps": {
                    "type": "array",
                    "description": "List of steps, each with a tool name and input",
                    "items": {
                        "type": "object",
                        "properties": {
                            "tool": {"type": "string", "description": "Tool name to call"},
                            "input": {"type": "object", "description": "Tool input parameters"},
                            "condition": {"type": "string", "description": "Optional: condition to check before running (Python expression)"}
                        },
                        "required": ["tool", "input"]
                    }
                }
            },
            "required": ["name", "steps"]
        }
    },
    {
        "name": "workflow_run",
        "description": "Execute a saved workflow by name",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Workflow name to run"},
                "variables": {"type": "object", "description": "Variables to pass to the workflow"}
            },
            "required": ["name"]
        }
    },
    {
        "name": "workflow_list",
        "description": "List all saved workflows",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    # ── Database (SQLite) ──
    {
        "name": "db_query",
        "description": "Execute a SQL query on the toolkit's SQLite database",
        "input_schema": {
            "type": "object",
            "properties": {
                "sql": {"type": "string", "description": "SQL query to execute"},
                "database": {"type": "string", "description": "Database name (default: toolkit.db)", "default": "toolkit.db"}
            },
            "required": ["sql"]
        }
    },
    {
        "name": "db_tables",
        "description": "List all tables in a database",
        "input_schema": {
            "type": "object",
            "properties": {
                "database": {"type": "string", "description": "Database name (default: toolkit.db)", "default": "toolkit.db"}
            }
        }
    },
    # ── Spreadsheet / Excel ──
    {
        "name": "spreadsheet_read",
        "description": "Read an Excel/Google Sheets file",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Path to .xlsx or .xls file"},
                "sheet": {"type": "string", "description": "Sheet name (default: first sheet)"},
                "max_rows": {"type": "integer", "description": "Max rows to return", "default": 100}
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "spreadsheet_write",
        "description": "Write data to an Excel spreadsheet",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Output .xlsx file path"},
                "headers": {"type": "array", "items": {"type": "string"}, "description": "Column headers"},
                "rows": {"type": "array", "items": {"type": "array"}, "description": "Data rows"},
                "sheet_name": {"type": "string", "description": "Sheet name", "default": "Sheet1"}
            },
            "required": ["file_path", "headers", "rows"]
        }
    },
    # ── Contact / CRM ──
    {
        "name": "contact_add",
        "description": "Add a contact to the CRM/contacts database",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Contact name"},
                "email": {"type": "string", "description": "Email address"},
                "phone": {"type": "string", "description": "Phone number"},
                "company": {"type": "string", "description": "Company name"},
                "notes": {"type": "string", "description": "Additional notes"},
                "tags": {"type": "array", "items": {"type": "string"}, "description": "Tags/labels"}
            },
            "required": ["name"]
        }
    },
    {
        "name": "contact_search",
        "description": "Search contacts by name, email, phone, company, or tags",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Search query"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "contact_list",
        "description": "List all contacts, optionally filtered by tag",
        "input_schema": {
            "type": "object",
            "properties": {
                "tag": {"type": "string", "description": "Optional tag to filter by"}
            }
        }
    },
    # ── Invoicing / Billing ──
    {
        "name": "invoice_create",
        "description": "Create a professional PDF invoice for a client",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_name": {"type": "string", "description": "Client/company name"},
                "client_email": {"type": "string", "description": "Client email"},
                "items": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "description": {"type": "string"},
                            "quantity": {"type": "number"},
                            "unit_price": {"type": "number"}
                        }
                    },
                    "description": "Line items"
                },
                "due_date": {"type": "string", "description": "Payment due date"},
                "notes": {"type": "string", "description": "Additional notes"}
            },
            "required": ["client_name", "items"]
        }
    },
    # ── Reporting / Analytics ──
    {
        "name": "report_generate",
        "description": "Generate a business report (revenue, clients, usage) as PDF or HTML",
        "input_schema": {
            "type": "object",
            "properties": {
                "report_type": {"type": "string", "description": "Type: revenue, clients, usage, custom"},
                "date_range": {"type": "string", "description": "Date range: 'today', 'this_week', 'this_month', 'last_30_days', or 'YYYY-MM-DD to YYYY-MM-DD'"},
                "format": {"type": "string", "description": "Output format: pdf, html, json", "default": "pdf"}
            },
            "required": ["report_type"]
        }
    },
    # ── Appointment / Calendar ──
    {
        "name": "appointment_book",
        "description": "Book an appointment for a client",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "customer_name": {"type": "string", "description": "Customer name"},
                "customer_phone": {"type": "string", "description": "Customer phone"},
                "service": {"type": "string", "description": "Service type"},
                "date": {"type": "string", "description": "Appointment date (YYYY-MM-DD)"},
                "time": {"type": "string", "description": "Appointment time (HH:MM)"},
                "staff": {"type": "string", "description": "Staff member (optional)"},
                "notes": {"type": "string", "description": "Additional notes"}
            },
            "required": ["client_id", "customer_name", "date", "time"]
        }
    },
    {
        "name": "appointment_list",
        "description": "List appointments for a client, optionally filtered by date",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "date": {"type": "string", "description": "Filter by date (YYYY-MM-DD)"}
            },
            "required": ["client_id"]
        }
    },
    {
        "name": "appointment_cancel",
        "description": "Cancel an appointment",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "appointment_id": {"type": "string", "description": "Appointment ID to cancel"}
            },
            "required": ["client_id", "appointment_id"]
        }
    },
    # ── Template System ──
    {
        "name": "template_list",
        "description": "List available templates (email, SMS, invoice, report)",
        "input_schema": {
            "type": "object",
            "properties": {
                "category": {"type": "string", "description": "Filter by category: email, sms, invoice, report"}
            }
        }
    },
    {
        "name": "template_render",
        "description": "Render a template with variables (e.g. fill in client name, date, amount)",
        "input_schema": {
            "type": "object",
            "properties": {
                "template_name": {"type": "string", "description": "Template name"},
                "variables": {"type": "object", "description": "Variables to fill in: {name: value, ...}"}
            },
            "required": ["template_name", "variables"]
        }
    },
    {
        "name": "template_create",
        "description": "Create a new reusable template",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Template name"},
                "category": {"type": "string", "description": "Category: email, sms, invoice, report, custom"},
                "content": {"type": "string", "description": "Template content with {{variable}} placeholders"}
            },
            "required": ["name", "category", "content"]
        }
    },
    # ── Monitoring / Uptime ──
    {
        "name": "monitor_url",
        "description": "Check if a URL/service is up and measure response time",
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "URL to check"},
                "expected_status": {"type": "integer", "description": "Expected HTTP status code (default: 200)", "default": 200}
            },
            "required": ["url"]
        }
    },
    {
        "name": "monitor_port",
        "description": "Check if a port is open on a host",
        "input_schema": {
            "type": "object",
            "properties": {
                "host": {"type": "string", "description": "Hostname or IP"},
                "port": {"type": "integer", "description": "Port number"}
            },
            "required": ["host", "port"]
        }
    },
    # ── Regex / Text Processing ──
    {
        "name": "regex_match",
        "description": "Find all regex matches in text",
        "input_schema": {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Text to search in"},
                "pattern": {"type": "string", "description": "Regex pattern"},
                "replace_with": {"type": "string", "description": "Optional: replace matches with this string"}
            },
            "required": ["text", "pattern"]
        }
    },
    {
        "name": "text_diff",
        "description": "Compare two texts and show the differences",
        "input_schema": {
            "type": "object",
            "properties": {
                "text1": {"type": "string", "description": "First text"},
                "text2": {"type": "string", "description": "Second text"}
            },
            "required": ["text1", "text2"]
        }
    },
    # ── Multi-Agent Orchestration ──
    {
        "name": "agent_spawn",
        "description": "Spawn a new AI agent to work on a task in parallel (can use different LLM providers)",
        "input_schema": {
            "type": "object",
            "properties": {
                "agent_id": {"type": "string", "description": "Unique identifier for this agent"},
                "task": {"type": "string", "description": "Task description for the agent"},
                "provider": {"type": "string", "description": "LLM provider: claude, groq, ollama, openai, pollinations", "default": "claude"},
                "model": {"type": "string", "description": "Specific model (e.g. haiku, sonnet, opus, llama3, gpt-4)"},
                "tools": {"type": "array", "items": {"type": "string"}, "description": "List of tool names this agent can use (default: all)"},
                "system_prompt": {"type": "string", "description": "Custom system prompt for this agent"},
                "max_turns": {"type": "integer", "description": "Max agent loop turns (default: 10)", "default": 10},
                "parallel": {"type": "boolean", "description": "Run in background (true) or wait for result (false)", "default": True}
            },
            "required": ["agent_id", "task"]
        }
    },
    {
        "name": "agent_status",
        "description": "Check the status of a running agent",
        "input_schema": {
            "type": "object",
            "properties": {
                "agent_id": {"type": "string", "description": "Agent ID to check"}
            },
            "required": ["agent_id"]
        }
    },
    {
        "name": "agent_result",
        "description": "Get the result/output from a completed agent",
        "input_schema": {
            "type": "object",
            "properties": {
                "agent_id": {"type": "string", "description": "Agent ID to get result from"},
                "wait": {"type": "boolean", "description": "Wait for agent to complete if still running", "default": False}
            },
            "required": ["agent_id"]
        }
    },
    {
        "name": "agent_stop",
        "description": "Stop a running agent",
        "input_schema": {
            "type": "object",
            "properties": {
                "agent_id": {"type": "string", "description": "Agent ID to stop"}
            },
            "required": ["agent_id"]
        }
    },
    {
        "name": "agent_list",
        "description": "List all active and completed agents",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "agent_message",
        "description": "Send a message to a running agent (inter-agent communication)",
        "input_schema": {
            "type": "object",
            "properties": {
                "from_agent": {"type": "string", "description": "Sender agent ID"},
                "to_agent": {"type": "string", "description": "Recipient agent ID"},
                "message": {"type": "string", "description": "Message content"}
            },
            "required": ["to_agent", "message"]
        }
    },
    {
        "name": "agent_team",
        "description": "Spawn a team of agents to work on related tasks in parallel, with a coordinator",
        "input_schema": {
            "type": "object",
            "properties": {
                "team_name": {"type": "string", "description": "Team identifier"},
                "goal": {"type": "string", "description": "The overall goal the team is working toward"},
                "agents": {
                    "type": "array",
                    "description": "List of agents to spawn",
                    "items": {
                        "type": "object",
                        "properties": {
                            "role": {"type": "string", "description": "Agent role (e.g. 'researcher', 'writer', 'reviewer')"},
                            "task": {"type": "string", "description": "Specific task for this agent"},
                            "provider": {"type": "string", "description": "LLM provider to use", "default": "claude"},
                            "tools": {"type": "array", "items": {"type": "string"}, "description": "Tools this agent can use"}
                        },
                        "required": ["role", "task"]
                    }
                }
            },
            "required": ["team_name", "goal", "agents"]
        }
    },
    # ── Pre-Built Business Bots (one-click setup) ──
    {
        "name": "bot_setup_receptionist",
        "description": "Set up an AI receptionist for a business — handles calls, books appointments, answers questions",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "business_name": {"type": "string", "description": "Business name"},
                "business_type": {"type": "string", "description": "Type: salon, dentist, restaurant, clinic, etc."},
                "phone": {"type": "string", "description": "Twilio phone number"},
                "services": {"type": "array", "items": {"type": "string"}, "description": "Services offered"},
                "business_hours": {"type": "string", "description": "Business hours"},
                "staff": {"type": "array", "items": {"type": "string"}, "description": "Staff names"},
                "greeting": {"type": "string", "description": "Custom greeting message"}
            },
            "required": ["client_id", "business_name", "business_type"]
        }
    },
    {
        "name": "bot_setup_email_assistant",
        "description": "Set up an AI email auto-responder for a business — monitors inbox, drafts replies",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "email": {"type": "string", "description": "Email address to monitor"},
                "email_password": {"type": "string", "description": "Email app password"},
                "business_name": {"type": "string", "description": "Business name"},
                "business_context": {"type": "string", "description": "What the business does"},
                "auto_send": {"type": "boolean", "description": "Auto-send replies or just draft", "default": False}
            },
            "required": ["client_id", "email", "email_password", "business_name"]
        }
    },
    {
        "name": "bot_setup_lead_gen",
        "description": "Set up an automated lead generation bot — finds businesses, enriches data, sends outreach",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "target_industry": {"type": "string", "description": "Industry to target (e.g. 'dentist', 'restaurant')"},
                "target_location": {"type": "string", "description": "Geographic area"},
                "outreach_template": {"type": "string", "description": "Email/message template for outreach"},
                "max_leads_per_day": {"type": "integer", "description": "Max leads to find per day", "default": 20}
            },
            "required": ["client_id", "target_industry", "target_location"]
        }
    },
    {
        "name": "bot_setup_social_media",
        "description": "Set up a social media management bot — schedules posts, monitors mentions, auto-replies",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "platforms": {"type": "array", "items": {"type": "string"}, "description": "Platforms: instagram, twitter, facebook, linkedin"},
                "business_name": {"type": "string", "description": "Business name"},
                "tone": {"type": "string", "description": "Brand voice/tone: professional, casual, fun, luxury"},
                "post_frequency": {"type": "string", "description": "How often to post: daily, 3x_week, weekly"}
            },
            "required": ["client_id", "platforms", "business_name"]
        }
    },
    {
        "name": "bot_setup_review_manager",
        "description": "Set up a review management bot — monitors reviews, generates responses, requests reviews from happy customers",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "business_name": {"type": "string", "description": "Business name"},
                "platforms": {"type": "array", "items": {"type": "string"}, "description": "Review platforms: google, yelp, facebook"},
                "auto_respond": {"type": "boolean", "description": "Auto-respond to reviews", "default": False}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_scheduler",
        "description": "Set up an appointment scheduling bot — online booking, reminders, calendar sync",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "business_name": {"type": "string", "description": "Business name"},
                "services": {"type": "array", "items": {"type": "string"}, "description": "Bookable services"},
                "slot_duration_minutes": {"type": "integer", "description": "Default appointment length", "default": 30},
                "business_hours": {"type": "string", "description": "Business hours"},
                "reminder_before_minutes": {"type": "integer", "description": "Send reminder N minutes before", "default": 60}
            },
            "required": ["client_id", "business_name", "services"]
        }
    },
    {
        "name": "bot_setup_chat_widget",
        "description": "Set up an AI chat widget for a website — embeddable, answers customer questions 24/7",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "business_name": {"type": "string", "description": "Business name"},
                "website_url": {"type": "string", "description": "Client's website URL"},
                "faq": {"type": "array", "items": {"type": "object"}, "description": "FAQ entries: [{question, answer}]"},
                "escalation_email": {"type": "string", "description": "Email to escalate to when bot can't help"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_invoice_collector",
        "description": "Set up an automated invoice/payment reminder bot",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Client ID"},
                "business_name": {"type": "string", "description": "Business name"},
                "reminder_days": {"type": "array", "items": {"type": "integer"}, "description": "Days before/after due to remind", "default": [7, 3, 1, -1, -7]},
                "payment_link": {"type": "string", "description": "Payment URL to include in reminders"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — Finance & Legal ──
    {
        "name": "bot_setup_accountant",
        "description": "AI Accountant — tracks income/expenses, categorizes transactions, generates financial reports, tax prep",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "business_type": {"type": "string", "description": "e.g. LLC, sole proprietor, S-Corp"},
                "currency": {"type": "string", "default": "USD"},
                "fiscal_year_start": {"type": "string", "description": "e.g. January or April", "default": "January"},
                "tax_state": {"type": "string", "description": "State for tax calculations"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_bookkeeper",
        "description": "AI Bookkeeper — daily transaction logging, receipt scanning (OCR), bank reconciliation, P&L statements",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "categories": {"type": "array", "items": {"type": "string"}, "description": "Expense categories (e.g. rent, utilities, supplies)"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_legal_assistant",
        "description": "AI Legal Assistant — contract review, terms of service generation, compliance checks, NDA drafting, legal research",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "industry": {"type": "string"},
                "jurisdiction": {"type": "string", "description": "State/country for legal context"},
                "specializations": {"type": "array", "items": {"type": "string"}, "description": "e.g. contracts, employment, IP, privacy"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_financial_advisor",
        "description": "AI Financial Advisor — cash flow forecasting, budget planning, investment analysis, financial health scoring",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "monthly_revenue": {"type": "number", "description": "Approximate monthly revenue"},
                "goals": {"type": "array", "items": {"type": "string"}, "description": "Financial goals"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_debt_collector",
        "description": "AI Debt Collector — tracks overdue invoices, sends escalating reminders, negotiates payment plans",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "reminder_schedule": {"type": "array", "items": {"type": "integer"}, "description": "Days after due date to send reminders"},
                "tone": {"type": "string", "description": "Communication tone: gentle, firm, formal", "default": "firm"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — Sales & Marketing ──
    {
        "name": "bot_setup_sales_rep",
        "description": "AI Sales Rep — cold outreach, follow-ups, pipeline management, objection handling, closes deals",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "product_service": {"type": "string", "description": "What you're selling"},
                "price_range": {"type": "string", "description": "Price range of offerings"},
                "target_audience": {"type": "string", "description": "Who you sell to"},
                "sales_script": {"type": "string", "description": "Custom sales script/talking points"},
                "follow_up_days": {"type": "array", "items": {"type": "integer"}, "description": "Days after initial contact to follow up", "default": [1, 3, 7, 14]}
            },
            "required": ["client_id", "business_name", "product_service"]
        }
    },
    {
        "name": "bot_setup_marketing_manager",
        "description": "AI Marketing Manager — campaign planning, content calendar, A/B testing, analytics, multi-channel strategy",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "target_audience": {"type": "string"},
                "channels": {"type": "array", "items": {"type": "string"}, "description": "Marketing channels: email, social, seo, ppc, content"},
                "monthly_budget": {"type": "number", "description": "Monthly marketing budget"},
                "brand_voice": {"type": "string", "description": "Brand voice/personality"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_seo_specialist",
        "description": "AI SEO Specialist — keyword research, on-page optimization, backlink analysis, content recommendations, rank tracking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "website_url": {"type": "string"},
                "target_keywords": {"type": "array", "items": {"type": "string"}},
                "competitors": {"type": "array", "items": {"type": "string"}, "description": "Competitor websites"}
            },
            "required": ["client_id", "business_name", "website_url"]
        }
    },
    {
        "name": "bot_setup_content_writer",
        "description": "AI Content Writer — blog posts, articles, newsletters, website copy, product descriptions, case studies",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "industry": {"type": "string"},
                "tone": {"type": "string", "description": "Writing tone: professional, casual, authoritative, friendly"},
                "content_types": {"type": "array", "items": {"type": "string"}, "description": "Types: blog, newsletter, social, website, product"},
                "posting_schedule": {"type": "string", "description": "How often to produce content"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_copywriter",
        "description": "AI Copywriter — ad copy, landing pages, email sequences, sales pages, CTAs, headlines",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "product_service": {"type": "string"},
                "unique_selling_points": {"type": "array", "items": {"type": "string"}},
                "target_audience": {"type": "string"},
                "brand_voice": {"type": "string"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_email_marketer",
        "description": "AI Email Marketer — drip campaigns, newsletters, segmentation, A/B subject lines, open rate optimization",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "email_platform": {"type": "string", "description": "Platform: mailchimp, sendgrid, smtp_direct"},
                "list_size": {"type": "integer", "description": "Approximate email list size"},
                "campaign_types": {"type": "array", "items": {"type": "string"}, "description": "Types: welcome, nurture, promotion, re-engagement, newsletter"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_pr_manager",
        "description": "AI PR Manager — press releases, media outreach, brand monitoring, crisis communication, reputation management",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "industry": {"type": "string"},
                "key_spokespeople": {"type": "array", "items": {"type": "string"}},
                "media_contacts": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_brand_manager",
        "description": "AI Brand Manager — brand guidelines enforcement, voice consistency, visual identity, brand health monitoring",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "brand_colors": {"type": "array", "items": {"type": "string"}},
                "brand_voice": {"type": "string"},
                "logo_path": {"type": "string"},
                "tagline": {"type": "string"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_competitor_analyst",
        "description": "AI Competitor Analyst — tracks competitor pricing, features, reviews, social media, job postings, news",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "competitors": {"type": "array", "items": {"type": "object"}, "description": "Competitor list: [{name, website, social}]"},
                "track": {"type": "array", "items": {"type": "string"}, "description": "What to track: pricing, features, reviews, social, hiring"}
            },
            "required": ["client_id", "business_name", "competitors"]
        }
    },
    {
        "name": "bot_setup_appointment_setter",
        "description": "AI Appointment Setter — cold calls/messages prospects, qualifies leads, books meetings for sales team",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "offering": {"type": "string", "description": "What you're offering"},
                "qualification_questions": {"type": "array", "items": {"type": "string"}},
                "calendar_link": {"type": "string", "description": "Booking link for meetings"},
                "daily_outreach_limit": {"type": "integer", "default": 50}
            },
            "required": ["client_id", "business_name", "offering"]
        }
    },
    # ── AI Employees — Customer Service ──
    {
        "name": "bot_setup_customer_support",
        "description": "AI Customer Support Agent — ticket handling, FAQ answers, issue resolution, escalation, satisfaction surveys",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "product_service": {"type": "string"},
                "common_issues": {"type": "array", "items": {"type": "string"}},
                "escalation_email": {"type": "string"},
                "sla_response_minutes": {"type": "integer", "description": "Target response time in minutes", "default": 15}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_customer_success",
        "description": "AI Customer Success Manager — onboarding, check-ins, churn prediction, upsell opportunities, NPS tracking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "product_service": {"type": "string"},
                "onboarding_steps": {"type": "array", "items": {"type": "string"}},
                "check_in_frequency_days": {"type": "integer", "default": 30}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — HR & Recruiting ──
    {
        "name": "bot_setup_hr_manager",
        "description": "AI HR Manager — employee handbook generation, policy compliance, time-off tracking, onboarding checklists, culture surveys",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "employee_count": {"type": "integer"},
                "policies": {"type": "array", "items": {"type": "string"}, "description": "Key policies to enforce"},
                "state": {"type": "string", "description": "State for employment law compliance"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_recruiter",
        "description": "AI Recruiter — job posting creation, resume screening, interview scheduling, candidate scoring, outreach",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "open_positions": {"type": "array", "items": {"type": "object"}, "description": "Positions: [{title, requirements, salary_range}]"},
                "hiring_platforms": {"type": "array", "items": {"type": "string"}, "description": "Where to post: indeed, linkedin, glassdoor"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_training_coach",
        "description": "AI Training Coach — employee onboarding, skill assessments, learning paths, quiz generation, progress tracking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "training_topics": {"type": "array", "items": {"type": "string"}},
                "assessment_frequency": {"type": "string", "description": "How often to assess: weekly, monthly, quarterly"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — Operations ──
    {
        "name": "bot_setup_project_manager",
        "description": "AI Project Manager — task tracking, deadline reminders, status reports, resource allocation, Gantt charts",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "team_members": {"type": "array", "items": {"type": "string"}},
                "project_methodology": {"type": "string", "description": "agile, waterfall, kanban", "default": "agile"},
                "reporting_frequency": {"type": "string", "default": "weekly"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_inventory_manager",
        "description": "AI Inventory Manager — stock tracking, reorder alerts, supplier management, demand forecasting, waste reduction",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "product_categories": {"type": "array", "items": {"type": "string"}},
                "reorder_threshold": {"type": "integer", "description": "Alert when stock falls below this", "default": 10},
                "suppliers": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_supply_chain",
        "description": "AI Supply Chain Manager — vendor evaluation, order tracking, logistics optimization, cost analysis",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "vendors": {"type": "array", "items": {"type": "object"}, "description": "Vendor list: [{name, product, lead_time}]"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_quality_assurance",
        "description": "AI QA Manager — checklists, inspection scheduling, defect tracking, compliance auditing, SOP enforcement",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "standards": {"type": "array", "items": {"type": "string"}, "description": "Quality standards to enforce"},
                "inspection_frequency": {"type": "string", "default": "weekly"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — IT & Tech ──
    {
        "name": "bot_setup_it_support",
        "description": "AI IT Support — troubleshooting, password resets, software setup, network diagnostics, security alerts",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "systems": {"type": "array", "items": {"type": "string"}, "description": "Systems managed: email, website, crm, pos, etc."},
                "escalation_contact": {"type": "string", "description": "Who to escalate critical issues to"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_security_analyst",
        "description": "AI Security Analyst — threat monitoring, vulnerability scanning, access audits, incident response, compliance reporting",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "domains": {"type": "array", "items": {"type": "string"}, "description": "Domains/IPs to monitor"},
                "compliance_frameworks": {"type": "array", "items": {"type": "string"}, "description": "e.g. HIPAA, PCI-DSS, SOC2, GDPR"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_devops",
        "description": "AI DevOps Engineer — server monitoring, deployment automation, log analysis, uptime tracking, auto-scaling",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "servers": {"type": "array", "items": {"type": "object"}, "description": "Servers: [{host, type, services}]"},
                "alert_channels": {"type": "array", "items": {"type": "string"}, "description": "Where to send alerts: telegram, discord, email, sms"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_data_analyst",
        "description": "AI Data Analyst — dashboards, trend analysis, KPI tracking, data visualization, predictive insights",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "data_sources": {"type": "array", "items": {"type": "string"}, "description": "Data sources: spreadsheets, database, api, csv"},
                "kpis": {"type": "array", "items": {"type": "string"}, "description": "Key metrics to track"},
                "report_frequency": {"type": "string", "default": "weekly"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — Creative ──
    {
        "name": "bot_setup_graphic_designer",
        "description": "AI Graphic Designer — social media graphics, logos, flyers, banners, brand assets, infographics",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "brand_colors": {"type": "array", "items": {"type": "string"}},
                "style": {"type": "string", "description": "Design style: modern, minimalist, bold, elegant, playful"},
                "output_formats": {"type": "array", "items": {"type": "string"}, "description": "Formats needed: instagram, facebook, linkedin, print"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_video_editor",
        "description": "AI Video Editor — clip trimming, captions, thumbnails, social media cuts, highlight reels",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "platforms": {"type": "array", "items": {"type": "string"}, "description": "Target platforms: youtube, tiktok, instagram, linkedin"},
                "style": {"type": "string", "description": "Editing style: fast-paced, cinematic, educational, casual"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_podcast_producer",
        "description": "AI Podcast Producer — episode planning, show notes, transcription, audiogram clips, guest research",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "podcast_name": {"type": "string"},
                "niche": {"type": "string"},
                "episode_frequency": {"type": "string", "default": "weekly"}
            },
            "required": ["client_id", "business_name", "podcast_name"]
        }
    },
    # ── AI Employees — Industry-Specific ──
    {
        "name": "bot_setup_real_estate_agent",
        "description": "AI Real Estate Agent — listing descriptions, market analysis, lead follow-up, showing scheduling, CMA reports",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "market_area": {"type": "string"},
                "specialization": {"type": "string", "description": "residential, commercial, luxury, investment"},
                "mls_access": {"type": "boolean", "default": False}
            },
            "required": ["client_id", "business_name", "market_area"]
        }
    },
    {
        "name": "bot_setup_insurance_agent",
        "description": "AI Insurance Agent — quote generation, policy comparison, claims assistance, renewal reminders, risk assessment",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "insurance_types": {"type": "array", "items": {"type": "string"}, "description": "Types: auto, home, life, health, business, liability"},
                "carriers": {"type": "array", "items": {"type": "string"}, "description": "Insurance carriers offered"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_medical_assistant",
        "description": "AI Medical Office Assistant — patient intake forms, appointment reminders, insurance verification, prescription refill requests",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "practice_type": {"type": "string", "description": "e.g. dental, general, dermatology, chiropractic"},
                "hipaa_compliant": {"type": "boolean", "default": True},
                "services": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name", "practice_type"]
        }
    },
    {
        "name": "bot_setup_restaurant_manager",
        "description": "AI Restaurant Manager — reservation handling, menu updates, review responses, inventory, staff scheduling, catering quotes",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "cuisine_type": {"type": "string"},
                "menu_items": {"type": "array", "items": {"type": "object"}, "description": "Menu: [{name, price, category}]"},
                "capacity": {"type": "integer"},
                "delivery_platforms": {"type": "array", "items": {"type": "string"}, "description": "e.g. doordash, ubereats, grubhub"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_fitness_coach",
        "description": "AI Fitness Coach — workout plans, nutrition tracking, progress monitoring, class scheduling, member engagement",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "gym_type": {"type": "string", "description": "e.g. crossfit, yoga, personal training, boxing"},
                "class_schedule": {"type": "array", "items": {"type": "object"}, "description": "Classes: [{name, day, time, instructor}]"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_salon_manager",
        "description": "AI Salon/Spa Manager — booking, stylist matching, service recommendations, loyalty programs, product suggestions",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "services": {"type": "array", "items": {"type": "object"}, "description": "Services: [{name, price, duration_min}]"},
                "stylists": {"type": "array", "items": {"type": "object"}, "description": "Staff: [{name, specialties}]"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_ecommerce_manager",
        "description": "AI E-Commerce Manager — product listings, price optimization, cart abandonment emails, inventory sync, order tracking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "platform": {"type": "string", "description": "e.g. shopify, woocommerce, etsy, amazon"},
                "product_count": {"type": "integer"},
                "shipping_options": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_property_manager",
        "description": "AI Property Manager — tenant communication, maintenance requests, rent collection, lease management, inspections",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "properties": {"type": "array", "items": {"type": "object"}, "description": "Properties: [{address, units, type}]"},
                "rent_due_day": {"type": "integer", "default": 1}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_tutor",
        "description": "AI Tutor — lesson planning, student assessment, homework help, progress reports, practice quiz generation",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "subjects": {"type": "array", "items": {"type": "string"}},
                "grade_levels": {"type": "array", "items": {"type": "string"}, "description": "e.g. K-5, middle school, high school, college"},
                "learning_style": {"type": "string", "description": "visual, auditory, hands-on, reading"}
            },
            "required": ["client_id", "business_name", "subjects"]
        }
    },
    {
        "name": "bot_setup_event_planner",
        "description": "AI Event Planner — venue research, vendor coordination, timeline creation, guest management, budget tracking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "event_types": {"type": "array", "items": {"type": "string"}, "description": "e.g. weddings, corporate, parties, conferences"},
                "service_area": {"type": "string"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_travel_agent",
        "description": "AI Travel Agent — trip planning, flight/hotel search, itinerary creation, budget optimization, travel alerts",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "specialization": {"type": "string", "description": "e.g. luxury, budget, adventure, business, family"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    # ── AI Employees — Executive / Strategic ──
    {
        "name": "bot_setup_executive_assistant",
        "description": "AI Executive Assistant — calendar management, meeting prep, email triage, travel booking, daily briefings",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "executive_name": {"type": "string"},
                "priorities": {"type": "array", "items": {"type": "string"}, "description": "Key focus areas"},
                "daily_briefing_time": {"type": "string", "description": "When to send daily briefing", "default": "08:00"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_business_consultant",
        "description": "AI Business Consultant — SWOT analysis, market research, growth strategies, process optimization, benchmarking",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "industry": {"type": "string"},
                "current_challenges": {"type": "array", "items": {"type": "string"}},
                "goals": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_compliance_officer",
        "description": "AI Compliance Officer — regulatory tracking, policy auditing, risk assessment, training requirements, filing deadlines",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "industry": {"type": "string"},
                "regulations": {"type": "array", "items": {"type": "string"}, "description": "e.g. HIPAA, GDPR, PCI-DSS, ADA, OSHA"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_research_analyst",
        "description": "AI Research Analyst — market research, industry reports, trend analysis, data collection, competitive intelligence",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "research_topics": {"type": "array", "items": {"type": "string"}},
                "sources": {"type": "array", "items": {"type": "string"}, "description": "Preferred sources: web, academic, news, social"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_proposal_writer",
        "description": "AI Proposal Writer — RFP responses, business proposals, pitch decks, grant applications, case studies",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "services": {"type": "array", "items": {"type": "string"}},
                "past_wins": {"type": "array", "items": {"type": "string"}, "description": "Notable past projects/clients"},
                "differentiators": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_setup_contract_manager",
        "description": "AI Contract Manager — contract drafting, renewal tracking, clause analysis, negotiation support, expiration alerts",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string"},
                "business_name": {"type": "string"},
                "contract_types": {"type": "array", "items": {"type": "string"}, "description": "Types: vendor, client, employment, NDA, lease"}
            },
            "required": ["client_id", "business_name"]
        }
    },
    {
        "name": "bot_list",
        "description": "List all configured bots/automations for a client or all clients",
        "input_schema": {
            "type": "object",
            "properties": {
                "client_id": {"type": "string", "description": "Optional: filter by client"}
            }
        }
    },
    # ── Payments / Crypto / Wallet ──
    {
        "name": "payment_create_invoice",
        "description": "Create a payment link/invoice (Stripe, PayPal, or crypto)",
        "input_schema": {
            "type": "object",
            "properties": {
                "amount": {"type": "number", "description": "Amount to charge"},
                "currency": {"type": "string", "description": "Currency: usd, eur, btc, eth, sol", "default": "usd"},
                "description": {"type": "string", "description": "What the payment is for"},
                "customer_email": {"type": "string", "description": "Customer email"},
                "provider": {"type": "string", "description": "Payment provider: stripe, paypal, crypto", "default": "stripe"}
            },
            "required": ["amount", "description"]
        }
    },
    {
        "name": "payment_check_status",
        "description": "Check the status of a payment/invoice",
        "input_schema": {
            "type": "object",
            "properties": {
                "payment_id": {"type": "string", "description": "Payment/invoice ID"}
            },
            "required": ["payment_id"]
        }
    },
    {
        "name": "crypto_wallet_balance",
        "description": "Check cryptocurrency wallet balance",
        "input_schema": {
            "type": "object",
            "properties": {
                "address": {"type": "string", "description": "Wallet address"},
                "network": {"type": "string", "description": "Network: ethereum, bitcoin, solana, polygon", "default": "ethereum"}
            },
            "required": ["address"]
        }
    },
    {
        "name": "crypto_send",
        "description": "Send cryptocurrency from toolkit wallet (requires private key in config)",
        "input_schema": {
            "type": "object",
            "properties": {
                "to_address": {"type": "string", "description": "Recipient wallet address"},
                "amount": {"type": "number", "description": "Amount to send"},
                "token": {"type": "string", "description": "Token: ETH, BTC, SOL, USDC, USDT"},
                "network": {"type": "string", "description": "Network: ethereum, bitcoin, solana", "default": "ethereum"}
            },
            "required": ["to_address", "amount", "token"]
        }
    },
    {
        "name": "crypto_price",
        "description": "Get real-time cryptocurrency prices",
        "input_schema": {
            "type": "object",
            "properties": {
                "tokens": {"type": "array", "items": {"type": "string"}, "description": "Token symbols: BTC, ETH, SOL, etc."}
            },
            "required": ["tokens"]
        }
    },
    # ── Math / Calculation ──
    {
        "name": "calculate",
        "description": "Evaluate a mathematical expression safely",
        "input_schema": {
            "type": "object",
            "properties": {
                "expression": {"type": "string", "description": "Math expression to evaluate (e.g. '2**10 + 5*3')"}
            },
            "required": ["expression"]
        }
    },
    # ── Date/Time ──
    {
        "name": "get_datetime",
        "description": "Get the current date, time, and timezone",
        "input_schema": {
            "type": "object",
            "properties": {
                "timezone": {"type": "string", "description": "Timezone (default: local)", "default": "local"}
            }
        }
    },
]


# ══════════════════════════════════════════════════════════════════════
# TOOL REGISTRY — combines everything
# ══════════════════════════════════════════════════════════════════════

# Map tool names to their module executors
_MODULE_TOOLS = {}
for t in WEB_TOOLS:
    _MODULE_TOOLS[t["name"]] = web_execute
for t in BROWSER_TOOLS:
    _MODULE_TOOLS[t["name"]] = browser_execute
for t in FILE_TOOLS:
    _MODULE_TOOLS[t["name"]] = file_execute


def get_all_tools():
    """Return all tool definitions for the Claude API tools parameter."""
    return WEB_TOOLS + BROWSER_TOOLS + FILE_TOOLS + BUILTIN_TOOLS


def get_tools_by_category():
    """Return tools grouped by category for UI display."""
    return {
        "Web Search": WEB_TOOLS,
        "Browser": BROWSER_TOOLS,
        "Files": FILE_TOOLS,
        "Email": [t for t in BUILTIN_TOOLS if t["name"].startswith("email_")],
        "Image": [t for t in BUILTIN_TOOLS if t["name"].startswith("image_")],
        "Code": [t for t in BUILTIN_TOOLS if t["name"] in ("code_execute", "shell_execute")],
        "Voice": [t for t in BUILTIN_TOOLS if t["name"].startswith("voice_")],
        "HTTP": [t for t in BUILTIN_TOOLS if t["name"] == "http_request"],
        "Scheduling": [t for t in BUILTIN_TOOLS if t["name"].startswith("schedule_")],
        "Memory": [t for t in BUILTIN_TOOLS if t["name"].startswith("memory_")],
        "Documents": [t for t in BUILTIN_TOOLS if t["name"].startswith("document_")],
        "Messaging": [t for t in BUILTIN_TOOLS if t["name"] in ("telegram_send", "discord_send", "sms_send")],
        "VPS": [t for t in BUILTIN_TOOLS if t["name"].startswith("vps_")],
        "Git": [t for t in BUILTIN_TOOLS if t["name"].startswith("git_")],
        "Skills": [t for t in BUILTIN_TOOLS if t["name"].startswith("skill_")],
        "Clients": [t for t in BUILTIN_TOOLS if t["name"].startswith("client_")],
        "Knowledge": [t for t in BUILTIN_TOOLS if t["name"].startswith("knowledge_")],
        "Account Automation": [t for t in BUILTIN_TOOLS if t["name"] in ("browser_signup", "browser_login", "browser_interactive")],
        "Data Intelligence": [t for t in BUILTIN_TOOLS if t["name"].startswith("scrape_")],
        "Lead Generation": [t for t in BUILTIN_TOOLS if t["name"].startswith("lead_")],
        "Notifications": [t for t in BUILTIN_TOOLS if t["name"] == "notify"],
        "Video / Camera": [t for t in BUILTIN_TOOLS if t["name"] in ("camera_capture", "video_record", "video_stream_start", "video_stream_stop", "screen_record")],
        "Audio": [t for t in BUILTIN_TOOLS if t["name"] in ("audio_record", "audio_play")],
        "Compression": [t for t in BUILTIN_TOOLS if t["name"].startswith("zip_")],
        "Translation": [t for t in BUILTIN_TOOLS if t["name"] == "translate"],
        "Weather": [t for t in BUILTIN_TOOLS if t["name"] == "weather_get"],
        "News": [t for t in BUILTIN_TOOLS if t["name"] == "news_get"],
        "Finance": [t for t in BUILTIN_TOOLS if t["name"] == "stock_price"],
        "Network": [t for t in BUILTIN_TOOLS if t["name"] in ("whois_lookup", "dns_lookup")],
        "Encoding / Crypto": [t for t in BUILTIN_TOOLS if t["name"] in ("hash_text", "encrypt_text", "decrypt_text", "base64_encode", "base64_decode")],
        "Image Editing": [t for t in BUILTIN_TOOLS if t["name"] in ("image_resize", "image_text_overlay", "image_convert")],
        "JSON / Data": [t for t in BUILTIN_TOOLS if t["name"] == "json_transform"],
        "QR Code": [t for t in BUILTIN_TOOLS if t["name"] == "qr_generate"],
        "Desktop": [t for t in BUILTIN_TOOLS if t["name"] in ("screenshot_desktop", "clipboard_read", "clipboard_write")],
        "PDF": [t for t in BUILTIN_TOOLS if t["name"] == "pdf_create"],
        "Processes": [t for t in BUILTIN_TOOLS if t["name"] in ("process_list", "process_kill")],
        "System": [t for t in BUILTIN_TOOLS if t["name"] in ("system_info", "wait")],
        "AI / LLM": [t for t in BUILTIN_TOOLS if t["name"].startswith("ai_")],
        "Workflows": [t for t in BUILTIN_TOOLS if t["name"].startswith("workflow_")],
        "Database": [t for t in BUILTIN_TOOLS if t["name"].startswith("db_")],
        "Spreadsheets": [t for t in BUILTIN_TOOLS if t["name"].startswith("spreadsheet_")],
        "Contacts / CRM": [t for t in BUILTIN_TOOLS if t["name"].startswith("contact_")],
        "Invoicing": [t for t in BUILTIN_TOOLS if t["name"].startswith("invoice_")],
        "Reports": [t for t in BUILTIN_TOOLS if t["name"].startswith("report_")],
        "Appointments": [t for t in BUILTIN_TOOLS if t["name"].startswith("appointment_")],
        "Templates": [t for t in BUILTIN_TOOLS if t["name"].startswith("template_")],
        "Monitoring": [t for t in BUILTIN_TOOLS if t["name"].startswith("monitor_")],
        "Text Processing": [t for t in BUILTIN_TOOLS if t["name"] in ("regex_match", "text_diff")],
        "Multi-Agent": [t for t in BUILTIN_TOOLS if t["name"].startswith("agent_")],
        "Pre-Built Bots": [t for t in BUILTIN_TOOLS if t["name"].startswith("bot_")],
        "Payments / Crypto": [t for t in BUILTIN_TOOLS if t["name"] in ("payment_create_invoice", "payment_check_status", "crypto_wallet_balance", "crypto_send", "crypto_price")],
        "Utilities": [t for t in BUILTIN_TOOLS if t["name"] in ("calculate", "get_datetime")],
    }


# ══════════════════════════════════════════════════════════════════════
# BUILT-IN TOOL EXECUTORS
# ══════════════════════════════════════════════════════════════════════

def _exec_email_send(inp, client_config=None):
    """Send an email via SMTP."""
    cfg = client_config or {}
    smtp_server = inp.get("smtp_server", cfg.get("smtp_server", "smtp.gmail.com"))
    smtp_port = inp.get("smtp_port", cfg.get("smtp_port", 587))
    from_email = cfg.get("email", "")
    password = cfg.get("email_password", "")

    if not from_email or not password:
        return {"error": "Email not configured. Set email and email_password in client config."}

    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = inp["to"]
    msg["Subject"] = inp["subject"]
    msg.attach(MIMEText(inp["body"], "plain"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(from_email, password)
            server.send_message(msg)
        return {"sent": True, "to": inp["to"], "subject": inp["subject"]}
    except Exception as e:
        return {"error": str(e)}


def _exec_email_read(inp, client_config=None):
    """Read emails from inbox via IMAP."""
    cfg = client_config or {}
    imap_server = inp.get("imap_server", cfg.get("imap_server", "imap.gmail.com"))
    email_addr = cfg.get("email", "")
    password = cfg.get("email_password", "")

    if not email_addr or not password:
        return {"error": "Email not configured. Set email and email_password in client config."}

    try:
        imap = imaplib.IMAP4_SSL(imap_server)
        imap.login(email_addr, password)
        imap.select(inp.get("folder", "INBOX"))

        search_criteria = "UNSEEN" if inp.get("unread_only", True) else "ALL"
        status, messages = imap.search(None, search_criteria)
        if status != "OK":
            return {"error": "Failed to search inbox"}

        msg_ids = messages[0].split()
        count = inp.get("count", 5)
        msg_ids = msg_ids[-count:]  # most recent

        emails = []
        for mid in reversed(msg_ids):
            status, data = imap.fetch(mid, "(RFC822)")
            if status != "OK":
                continue
            msg = email_lib.message_from_bytes(data[0][1])
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_payload(decode=True).decode("utf-8", errors="replace")
                        break
            else:
                body = msg.get_payload(decode=True).decode("utf-8", errors="replace")

            emails.append({
                "from": msg.get("From", ""),
                "to": msg.get("To", ""),
                "subject": msg.get("Subject", ""),
                "date": msg.get("Date", ""),
                "body": body[:3000]
            })

        imap.logout()
        return {"count": len(emails), "emails": emails}
    except Exception as e:
        return {"error": str(e)}


def _exec_image_generate(inp):
    """Generate an image using Pollinations.ai (free, no key)."""
    import requests
    from urllib.parse import quote

    prompt = inp["prompt"]
    width = inp.get("width", 1024)
    height = inp.get("height", 1024)

    url = f"https://image.pollinations.ai/prompt/{quote(prompt)}?width={width}&height={height}&nologo=true"

    try:
        resp = requests.get(url, headers={"User-Agent": "Janovum/1.0"}, timeout=60)
        if resp.status_code != 200:
            return {"error": f"Image generation failed: HTTP {resp.status_code}"}

        # Save to data/images/
        img_dir = PLATFORM_DIR / "data" / "images"
        img_dir.mkdir(parents=True, exist_ok=True)

        filename = inp.get("save_as", f"gen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        filepath = img_dir / filename

        with open(filepath, "wb") as f:
            f.write(resp.content)

        return {"saved_to": str(filepath), "size_bytes": len(resp.content), "prompt": prompt}
    except Exception as e:
        return {"error": str(e)}


def _exec_code_execute(inp):
    """Execute Python code in sandbox."""
    from .sandbox import get_sandbox
    sandbox = get_sandbox()
    result = sandbox.execute_python(inp["code"], timeout=inp.get("timeout", 30))
    return result.to_dict()


def _exec_shell_execute(inp):
    """Execute shell command in sandbox."""
    from .sandbox import get_sandbox
    sandbox = get_sandbox()
    result = sandbox.execute_shell(inp["command"], timeout=inp.get("timeout", 30))
    return result.to_dict()


def _exec_voice_tts(inp):
    """Convert text to speech using edge-tts."""
    try:
        import edge_tts
        import asyncio
    except ImportError:
        return {"error": "edge-tts not installed. Run: pip install edge-tts"}

    text = inp["text"]
    voice = inp.get("voice", "en-US-AriaNeural")

    audio_dir = PLATFORM_DIR / "data" / "audio"
    audio_dir.mkdir(parents=True, exist_ok=True)

    filename = inp.get("save_as", f"tts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp3")
    filepath = audio_dir / filename

    async def _generate():
        communicate = edge_tts.Communicate(text, voice)
        await communicate.save(str(filepath))

    asyncio.run(_generate())
    return {"saved_to": str(filepath), "voice": voice, "text_length": len(text)}


def _exec_voice_stt(inp):
    """Transcribe audio to text using faster-whisper."""
    try:
        from faster_whisper import WhisperModel
    except ImportError:
        return {"error": "faster-whisper not installed. Run: pip install faster-whisper"}

    audio_path = inp["audio_path"]
    if not os.path.exists(audio_path):
        return {"error": f"Audio file not found: {audio_path}"}

    model = WhisperModel("base", compute_type="int8")
    segments, info = model.transcribe(audio_path)

    text = " ".join(seg.text for seg in segments)
    return {"text": text.strip(), "language": info.language, "duration": info.duration}


def _exec_http_request(inp):
    """Make an HTTP request."""
    import requests

    method = inp.get("method", "GET").upper()
    url = inp["url"]
    headers = inp.get("headers", {})
    headers.setdefault("User-Agent", "Janovum/1.0")

    try:
        kwargs = {"headers": headers, "timeout": 30}
        if inp.get("json_body"):
            kwargs["json"] = inp["json_body"]
        elif inp.get("body"):
            kwargs["data"] = inp["body"]

        resp = requests.request(method, url, **kwargs)

        # Try to parse as JSON
        try:
            body = resp.json()
        except Exception:
            body = resp.text[:10000]

        return {
            "status_code": resp.status_code,
            "headers": dict(resp.headers),
            "body": body
        }
    except Exception as e:
        return {"error": str(e)}


def _exec_schedule_task(inp):
    """Schedule a task (persisted to disk)."""
    schedule_dir = PLATFORM_DIR / "data" / "schedules"
    schedule_dir.mkdir(parents=True, exist_ok=True)

    task = {
        "name": inp["name"],
        "action": inp["action"],
        "action_data": inp["action_data"],
        "run_at": inp.get("run_at"),
        "interval_minutes": inp.get("interval_minutes"),
        "created_at": datetime.now().isoformat(),
        "status": "scheduled"
    }

    filepath = schedule_dir / f"{inp['name']}.json"
    with open(filepath, "w") as f:
        json.dump(task, f, indent=2)

    return {"scheduled": True, "name": inp["name"], "saved_to": str(filepath)}


def _exec_schedule_list(inp):
    """List all scheduled tasks."""
    schedule_dir = PLATFORM_DIR / "data" / "schedules"
    if not schedule_dir.exists():
        return {"tasks": [], "count": 0}

    tasks = []
    for f in schedule_dir.glob("*.json"):
        with open(f) as fh:
            tasks.append(json.load(fh))

    return {"tasks": tasks, "count": len(tasks)}


def _exec_schedule_cancel(inp):
    """Cancel a scheduled task."""
    filepath = PLATFORM_DIR / "data" / "schedules" / f"{inp['name']}.json"
    if not filepath.exists():
        return {"error": f"Task not found: {inp['name']}"}
    filepath.unlink()
    return {"cancelled": True, "name": inp["name"]}


def _exec_memory_read(inp, client_id="default"):
    """Read agent memory."""
    from .memory import read_memory
    mem_type = inp.get("memory_type", "context")
    content = read_memory(client_id, mem_type)
    return {"memory_type": mem_type, "content": content}


def _exec_memory_write(inp, client_id="default"):
    """Write to agent memory."""
    from .memory import write_memory, append_memory
    mem_type = inp.get("memory_type", "notes")
    if inp.get("append", True):
        append_memory(client_id, mem_type, inp["content"])
    else:
        write_memory(client_id, mem_type, inp["content"])
    return {"saved": True, "memory_type": mem_type, "append": inp.get("append", True)}


def _exec_calculate(inp):
    """Safely evaluate a math expression."""
    expr = inp["expression"]
    # Only allow safe math operations
    allowed = set("0123456789+-*/().%** ,")
    import math
    safe_names = {k: v for k, v in math.__dict__.items() if not k.startswith("_")}
    safe_names["abs"] = abs
    safe_names["round"] = round
    safe_names["min"] = min
    safe_names["max"] = max
    safe_names["int"] = int
    safe_names["float"] = float

    try:
        result = eval(expr, {"__builtins__": {}}, safe_names)
        return {"expression": expr, "result": result}
    except Exception as e:
        return {"error": str(e), "expression": expr}


def _exec_get_datetime(inp):
    """Get current date and time."""
    now = datetime.now()
    return {
        "date": now.strftime("%Y-%m-%d"),
        "time": now.strftime("%H:%M:%S"),
        "day_of_week": now.strftime("%A"),
        "iso": now.isoformat(),
        "timestamp": int(now.timestamp())
    }


# ── Document Parsing Executors ──

def _exec_document_read_pdf(inp):
    """Extract text from a PDF."""
    filepath = inp["file_path"]
    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}

    try:
        import fitz  # PyMuPDF
    except ImportError:
        # Fallback: try pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                pages_range = inp.get("pages", "all")
                texts = []
                for i, page in enumerate(pdf.pages):
                    if pages_range != "all":
                        if "-" in pages_range:
                            start, end = map(int, pages_range.split("-"))
                            if i + 1 < start or i + 1 > end:
                                continue
                        elif int(pages_range) != i + 1:
                            continue
                    texts.append(f"--- Page {i+1} ---\n{page.extract_text() or ''}")
                return {"text": "\n".join(texts), "page_count": len(pdf.pages)}
        except ImportError:
            return {"error": "No PDF library installed. Run: pip install PyMuPDF or pip install pdfplumber"}

    doc = fitz.open(filepath)
    pages_range = inp.get("pages", "all")
    texts = []
    for i in range(len(doc)):
        if pages_range != "all":
            if "-" in str(pages_range):
                start, end = map(int, str(pages_range).split("-"))
                if i + 1 < start or i + 1 > end:
                    continue
            elif int(pages_range) != i + 1:
                continue
        texts.append(f"--- Page {i+1} ---\n{doc[i].get_text()}")
    doc.close()
    return {"text": "\n".join(texts), "page_count": len(doc)}


def _exec_document_read_docx(inp):
    """Extract text from a Word document."""
    filepath = inp["file_path"]
    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}
    try:
        from docx import Document
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}

    doc = Document(filepath)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return {"text": "\n".join(paragraphs), "paragraph_count": len(paragraphs)}


def _exec_document_read_csv(inp):
    """Read and parse a CSV file."""
    import csv
    filepath = inp["file_path"]
    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}

    max_rows = inp.get("max_rows", 100)
    rows = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(dict(row))

    return {"headers": headers, "rows": rows, "row_count": len(rows)}


def _exec_document_ocr(inp):
    """Extract text from an image using OCR."""
    image_path = inp["image_path"]
    if not os.path.exists(image_path):
        return {"error": f"File not found: {image_path}"}

    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        # Fallback: try OCR.space API (free, 500/day)
        try:
            import requests
            with open(image_path, "rb") as f:
                resp = requests.post(
                    "https://api.ocr.space/parse/image",
                    files={"file": f},
                    data={"apikey": "helloworld", "language": "eng"},
                    timeout=30
                )
            data = resp.json()
            if data.get("ParsedResults"):
                text = data["ParsedResults"][0].get("ParsedText", "")
                return {"text": text, "method": "ocr.space"}
            return {"error": "OCR failed", "details": data}
        except Exception as e2:
            return {"error": f"No OCR library. Install: pip install pytesseract pillow. Fallback error: {e2}"}

    img = Image.open(image_path)
    text = pytesseract.image_to_string(img)
    return {"text": text.strip(), "method": "tesseract"}


# ── Messaging Executors ──

def _exec_telegram_send(inp, client_config=None):
    """Send a Telegram message."""
    import requests
    cfg = client_config or {}
    bot_token = cfg.get("telegram_bot_token", "")
    if not bot_token:
        return {"error": "Telegram bot token not configured. Set telegram_bot_token in client config."}

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {
        "chat_id": inp["chat_id"],
        "text": inp["text"],
        "parse_mode": inp.get("parse_mode", "HTML")
    }
    try:
        resp = requests.post(url, json=payload, timeout=10)
        return resp.json()
    except Exception as e:
        return {"error": str(e)}


def _exec_discord_send(inp):
    """Send a Discord webhook message."""
    import requests
    payload = {"content": inp["content"]}
    if inp.get("username"):
        payload["username"] = inp["username"]
    try:
        resp = requests.post(inp["webhook_url"], json=payload, timeout=10)
        return {"sent": True, "status_code": resp.status_code}
    except Exception as e:
        return {"error": str(e)}


def _exec_sms_send(inp, client_config=None):
    """Send SMS via Twilio."""
    cfg = client_config or {}
    account_sid = cfg.get("twilio_account_sid", "")
    auth_token = cfg.get("twilio_auth_token", "")
    from_number = cfg.get("twilio_phone", "")

    if not all([account_sid, auth_token, from_number]):
        return {"error": "Twilio not configured. Set twilio_account_sid, twilio_auth_token, twilio_phone in config."}

    import requests
    url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
    try:
        resp = requests.post(url, data={
            "To": inp["to"],
            "From": from_number,
            "Body": inp["body"]
        }, auth=(account_sid, auth_token), timeout=15)
        return resp.json()
    except Exception as e:
        return {"error": str(e)}


# ── VPS Management Executors ──

def _exec_vps_execute(inp, client_config=None):
    """Execute command on VPS via SSH."""
    cfg = client_config or {}
    vps_host = cfg.get("vps_host", "")
    vps_user = cfg.get("vps_user", "root")
    vps_key = cfg.get("vps_ssh_key", "")

    if not vps_host:
        return {"error": "VPS not configured. Set vps_host in config."}

    ssh_cmd = ["ssh", "-o", "StrictHostKeyChecking=no", "-o", "ConnectTimeout=10"]
    if vps_key:
        ssh_cmd += ["-i", vps_key]
    ssh_cmd += [f"{vps_user}@{vps_host}", inp["command"]]

    try:
        result = subprocess.run(ssh_cmd, capture_output=True, text=True, timeout=inp.get("timeout", 30))
        return {"stdout": result.stdout[:5000], "stderr": result.stderr[:2000], "returncode": result.returncode}
    except subprocess.TimeoutExpired:
        return {"error": "SSH command timed out"}
    except Exception as e:
        return {"error": str(e)}


def _exec_vps_upload(inp, client_config=None):
    """Upload file to VPS via SCP."""
    cfg = client_config or {}
    vps_host = cfg.get("vps_host", "")
    vps_user = cfg.get("vps_user", "root")
    vps_key = cfg.get("vps_ssh_key", "")

    if not vps_host:
        return {"error": "VPS not configured."}
    if not os.path.exists(inp["local_path"]):
        return {"error": f"Local file not found: {inp['local_path']}"}

    scp_cmd = ["scp", "-o", "StrictHostKeyChecking=no"]
    if vps_key:
        scp_cmd += ["-i", vps_key]
    scp_cmd += [inp["local_path"], f"{vps_user}@{vps_host}:{inp['remote_path']}"]

    try:
        result = subprocess.run(scp_cmd, capture_output=True, text=True, timeout=60)
        return {"uploaded": result.returncode == 0, "stderr": result.stderr[:500]}
    except Exception as e:
        return {"error": str(e)}


def _exec_vps_download(inp, client_config=None):
    """Download file from VPS via SCP."""
    cfg = client_config or {}
    vps_host = cfg.get("vps_host", "")
    vps_user = cfg.get("vps_user", "root")
    vps_key = cfg.get("vps_ssh_key", "")

    if not vps_host:
        return {"error": "VPS not configured."}

    scp_cmd = ["scp", "-o", "StrictHostKeyChecking=no"]
    if vps_key:
        scp_cmd += ["-i", vps_key]
    scp_cmd += [f"{vps_user}@{vps_host}:{inp['remote_path']}", inp["local_path"]]

    try:
        result = subprocess.run(scp_cmd, capture_output=True, text=True, timeout=60)
        return {"downloaded": result.returncode == 0, "stderr": result.stderr[:500]}
    except Exception as e:
        return {"error": str(e)}


# ── Git Executor ──

def _exec_git_execute(inp):
    """Run a git command."""
    command = inp["command"]
    working_dir = inp.get("working_dir", str(PLATFORM_DIR))

    # Block dangerous commands
    dangerous = ["push --force", "reset --hard", "clean -f"]
    for d in dangerous:
        if d in command:
            return {"error": f"BLOCKED: Dangerous git command — {d}. Use with caution."}

    try:
        result = subprocess.run(
            f"git {command}", shell=True, capture_output=True, text=True,
            timeout=30, cwd=working_dir
        )
        return {"stdout": result.stdout[:5000], "stderr": result.stderr[:2000], "returncode": result.returncode}
    except Exception as e:
        return {"error": str(e)}


# ── Skills / Plugin Executors ──

SKILLS_DIR = PLATFORM_DIR / "modules" / "skills"
CUSTOM_SKILLS_DIR = PLATFORM_DIR / "data" / "skills"


def _exec_skill_list(inp):
    """List all available skills."""
    skills = []

    # Built-in skills (modules/skills/*.md)
    if SKILLS_DIR.exists():
        for f in SKILLS_DIR.glob("*.md"):
            skills.append({"name": f.stem, "type": "builtin", "path": str(f)})

    # Custom skills (data/skills/*.md)
    if CUSTOM_SKILLS_DIR.exists():
        for f in CUSTOM_SKILLS_DIR.glob("*.md"):
            skills.append({"name": f.stem, "type": "custom", "path": str(f)})

    # Python modules (modules/*.py)
    modules_dir = PLATFORM_DIR / "modules"
    if modules_dir.exists():
        for f in modules_dir.glob("*.py"):
            if f.name != "__init__.py":
                skills.append({"name": f.stem, "type": "module", "path": str(f)})

    return {"skills": skills, "count": len(skills)}


def _exec_skill_load(inp):
    """Load a skill file."""
    name = inp["skill_name"]

    # Search order: custom > builtin > module
    for directory in [CUSTOM_SKILLS_DIR, SKILLS_DIR]:
        md_path = directory / f"{name}.md"
        if md_path.exists():
            with open(md_path, "r", encoding="utf-8") as f:
                return {"name": name, "content": f.read(), "type": "skill"}

    py_path = PLATFORM_DIR / "modules" / f"{name}.py"
    if py_path.exists():
        with open(py_path, "r", encoding="utf-8") as f:
            return {"name": name, "content": f.read(), "type": "module"}

    return {"error": f"Skill not found: {name}"}


def _exec_skill_create(inp):
    """Create a new custom skill."""
    CUSTOM_SKILLS_DIR.mkdir(parents=True, exist_ok=True)

    content = f"""---
name: {inp['skill_name']}
description: {inp['description']}
---

{inp['instructions']}
"""
    filepath = CUSTOM_SKILLS_DIR / f"{inp['skill_name']}.md"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)

    return {"created": True, "name": inp["skill_name"], "path": str(filepath)}


# ── Knowledge / Learning Executors ──

KNOWLEDGE_DIR = PLATFORM_DIR / "data" / "knowledge"


def _exec_knowledge_save(inp):
    """Save knowledge to persistent store."""
    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    topic_dir = KNOWLEDGE_DIR / inp["topic"]
    topic_dir.mkdir(parents=True, exist_ok=True)

    # Sanitize title for filename
    safe_title = "".join(c if c.isalnum() or c in " -_" else "" for c in inp["title"]).strip().replace(" ", "_")
    filepath = topic_dir / f"{safe_title}.json"

    entry = {
        "topic": inp["topic"],
        "title": inp["title"],
        "content": inp["content"],
        "tags": inp.get("tags", []),
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(entry, f, indent=2)

    return {"saved": True, "topic": inp["topic"], "title": inp["title"], "path": str(filepath)}


def _exec_knowledge_search(inp):
    """Search the knowledge base."""
    if not KNOWLEDGE_DIR.exists():
        return {"results": [], "count": 0}

    query = inp["query"].lower()
    topic_filter = inp.get("topic")
    results = []

    search_dirs = [KNOWLEDGE_DIR / topic_filter] if topic_filter else [d for d in KNOWLEDGE_DIR.iterdir() if d.is_dir()]

    for topic_dir in search_dirs:
        if not topic_dir.exists():
            continue
        for f in topic_dir.glob("*.json"):
            try:
                with open(f) as fh:
                    entry = json.load(fh)
                # Search in title, content, tags
                searchable = f"{entry.get('title', '')} {entry.get('content', '')} {' '.join(entry.get('tags', []))}".lower()
                if query in searchable:
                    results.append(entry)
            except Exception:
                pass

    return {"results": results, "count": len(results), "query": inp["query"]}


def _exec_knowledge_list(inp):
    """List all knowledge entries."""
    if not KNOWLEDGE_DIR.exists():
        return {"entries": [], "count": 0, "topics": []}

    topic_filter = inp.get("topic")
    entries = []
    topics = set()

    search_dirs = [KNOWLEDGE_DIR / topic_filter] if topic_filter else [d for d in KNOWLEDGE_DIR.iterdir() if d.is_dir()]

    for topic_dir in search_dirs:
        if not topic_dir.exists() or not topic_dir.is_dir():
            continue
        topics.add(topic_dir.name)
        for f in topic_dir.glob("*.json"):
            try:
                with open(f) as fh:
                    entry = json.load(fh)
                entries.append({"topic": entry.get("topic"), "title": entry.get("title"),
                                "tags": entry.get("tags", []), "created_at": entry.get("created_at")})
            except Exception:
                pass

    return {"entries": entries, "count": len(entries), "topics": sorted(topics)}


# ── Account / Signup Automation Executors ──

def _exec_browser_signup(inp):
    """Automate account signup."""
    from modules.browser_agent import get_driver
    from selenium.webdriver.common.by import By

    driver = get_driver(headless=inp.get("headless", False))
    try:
        driver.get(inp["url"])
        time.sleep(2)

        filled = []
        for selector, value in inp["fields"].items():
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                el.clear()
                el.send_keys(value)
                filled.append(selector)
                time.sleep(0.3)
            except Exception as e:
                filled.append(f"{selector} (FAILED: {e})")

        if inp.get("submit_selector"):
            try:
                btn = driver.find_element(By.CSS_SELECTOR, inp["submit_selector"])
                btn.click()
                time.sleep(3)
            except Exception as e:
                return {"error": f"Submit failed: {e}", "fields_filled": filled}

        return {
            "url": driver.current_url,
            "title": driver.title,
            "fields_filled": filled,
            "submitted": bool(inp.get("submit_selector"))
        }
    finally:
        driver.quit()


def _exec_browser_login(inp):
    """Log into a website."""
    from modules.browser_agent import get_driver
    from selenium.webdriver.common.by import By

    driver = get_driver(headless=True)
    try:
        driver.get(inp["url"])
        time.sleep(2)

        # Find and fill username
        user_sel = inp.get("username_selector", "input[name='username'], input[name='email'], input[type='email']")
        el = driver.find_element(By.CSS_SELECTOR, user_sel)
        el.clear()
        el.send_keys(inp["username"])

        # Find and fill password
        pass_sel = inp.get("password_selector", "input[name='password'], input[type='password']")
        el = driver.find_element(By.CSS_SELECTOR, pass_sel)
        el.clear()
        el.send_keys(inp["password"])

        # Submit
        submit_sel = inp.get("submit_selector", "button[type='submit'], input[type='submit']")
        btn = driver.find_element(By.CSS_SELECTOR, submit_sel)
        btn.click()
        time.sleep(3)

        cookies = driver.get_cookies()
        return {
            "logged_in": True,
            "url": driver.current_url,
            "title": driver.title,
            "cookies": cookies
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        driver.quit()


def _exec_browser_interactive(inp):
    """Run multi-step browser session."""
    from modules.browser_agent import get_driver
    from selenium.webdriver.common.by import By

    driver = get_driver(headless=inp.get("headless", True))
    results = []
    try:
        for i, step in enumerate(inp["steps"]):
            action = step["action"]
            selector = step.get("selector", "")
            value = step.get("value", "")

            try:
                if action == "navigate":
                    driver.get(value)
                    time.sleep(2)
                    results.append({"step": i, "action": "navigate", "url": driver.current_url})
                elif action == "click":
                    el = driver.find_element(By.CSS_SELECTOR, selector)
                    el.click()
                    time.sleep(1)
                    results.append({"step": i, "action": "click", "selector": selector})
                elif action == "type":
                    el = driver.find_element(By.CSS_SELECTOR, selector)
                    el.clear()
                    el.send_keys(value)
                    results.append({"step": i, "action": "type", "selector": selector})
                elif action == "screenshot":
                    img_dir = PLATFORM_DIR / "data" / "images"
                    img_dir.mkdir(parents=True, exist_ok=True)
                    path = img_dir / f"step_{i}_{datetime.now().strftime('%H%M%S')}.png"
                    driver.save_screenshot(str(path))
                    results.append({"step": i, "action": "screenshot", "saved_to": str(path)})
                elif action == "wait":
                    time.sleep(float(value) if value else 2)
                    results.append({"step": i, "action": "wait", "seconds": value})
                elif action == "extract_text":
                    if selector:
                        el = driver.find_element(By.CSS_SELECTOR, selector)
                        text = el.text
                    else:
                        text = driver.find_element(By.TAG_NAME, "body").text
                    results.append({"step": i, "action": "extract_text", "text": text[:5000]})
                elif action == "extract_links":
                    links = driver.find_elements(By.TAG_NAME, "a")
                    link_data = [{"text": l.text, "href": l.get_attribute("href")} for l in links if l.get_attribute("href")]
                    results.append({"step": i, "action": "extract_links", "links": link_data[:50]})
                elif action == "scroll":
                    driver.execute_script(f"window.scrollBy(0, {value or 500})")
                    results.append({"step": i, "action": "scroll"})
                else:
                    results.append({"step": i, "action": action, "error": "Unknown action"})
            except Exception as e:
                results.append({"step": i, "action": action, "error": str(e)})

        return {"steps_completed": len(results), "results": results, "final_url": driver.current_url}
    finally:
        driver.quit()


# ── Data Intelligence Executors ──

def _exec_scrape_google(inp):
    """Scrape Google search results."""
    from modules.browser_agent import get_driver
    from selenium.webdriver.common.by import By

    driver = get_driver(headless=True)
    try:
        query = inp["query"]
        num = inp.get("num_results", 10)
        driver.get(f"https://www.google.com/search?q={query}&num={num}")
        time.sleep(2)

        results = []
        search_divs = driver.find_elements(By.CSS_SELECTOR, "div.g")
        for div in search_divs[:num]:
            try:
                title_el = div.find_element(By.CSS_SELECTOR, "h3")
                link_el = div.find_element(By.CSS_SELECTOR, "a")
                snippet_els = div.find_elements(By.CSS_SELECTOR, "div.VwiC3b, span.aCOpRe")
                snippet = snippet_els[0].text if snippet_els else ""
                results.append({
                    "title": title_el.text,
                    "url": link_el.get_attribute("href"),
                    "snippet": snippet
                })
            except Exception:
                pass

        return {"query": query, "results": results, "count": len(results)}
    finally:
        driver.quit()


def _exec_scrape_social(inp):
    """Scrape social media (public data only)."""
    from modules.browser_agent import get_driver
    from selenium.webdriver.common.by import By

    platform = inp["platform"].lower()
    target = inp["target"]

    urls = {
        "twitter": f"https://x.com/{target}",
        "reddit": f"https://old.reddit.com/user/{target}" if inp.get("action") == "profile" else f"https://old.reddit.com/search?q={target}",
        "linkedin": f"https://www.linkedin.com/in/{target}",
        "instagram": f"https://www.instagram.com/{target}/"
    }

    url = urls.get(platform, f"https://www.google.com/search?q={platform}+{target}")

    driver = get_driver(headless=True)
    try:
        driver.get(url)
        time.sleep(3)
        body_text = driver.find_element(By.TAG_NAME, "body").text[:8000]
        return {"platform": platform, "target": target, "url": driver.current_url, "content": body_text}
    finally:
        driver.quit()


def _exec_scrape_business_info(inp):
    """Look up business information."""
    from modules.web_search import search
    query = f"{inp['business_name']} {inp.get('location', '')} contact phone email".strip()
    results = search(query, max_results=5)
    return {"business": inp["business_name"], "search_results": results}


# ── Lead Generation Executors ──

def _exec_lead_find(inp):
    """Find business leads."""
    from modules.web_search import search
    industry = inp["industry"]
    location = inp["location"]
    max_results = inp.get("max_results", 20)

    queries = [
        f"{industry} near {location}",
        f"{industry} in {location} phone number email",
        f"best {industry} {location}",
    ]

    all_results = []
    for q in queries:
        results = search(q, max_results=max_results)
        all_results.extend(results)

    # Deduplicate by URL
    seen = set()
    unique = []
    for r in all_results:
        url = r.get("url", "")
        if url not in seen:
            seen.add(url)
            unique.append(r)

    return {"industry": industry, "location": location, "leads": unique[:max_results], "count": len(unique[:max_results])}


def _exec_lead_enrich(inp):
    """Enrich a lead with more data."""
    from modules.web_search import search, fetch_url

    name = inp["business_name"]
    location = inp.get("location", "")

    # Search for contact info
    results = search(f"{name} {location} phone email website", max_results=3)

    # If we have a website, scrape it
    website_text = ""
    if inp.get("website"):
        fetched = fetch_url(inp["website"])
        website_text = fetched.get("text", "")[:3000]

    return {
        "business_name": name,
        "search_results": results,
        "website_text": website_text
    }


# ── Notifications Executor ──

def _exec_notify(inp, client_config=None):
    """Send notification via best available channel."""
    cfg = client_config or {}
    message = inp["message"]
    channel = inp.get("channel")

    # Try channels in priority order
    if channel:
        channels = [channel]
    else:
        channels = ["telegram", "discord", "email", "sms"]

    for ch in channels:
        try:
            if ch == "telegram" and cfg.get("telegram_bot_token") and cfg.get("telegram_chat_id"):
                result = _exec_telegram_send({"chat_id": cfg["telegram_chat_id"], "text": message}, cfg)
                if not result.get("error"):
                    return {"sent_via": "telegram", "result": result}
            elif ch == "discord" and cfg.get("discord_webhook_url"):
                result = _exec_discord_send({"webhook_url": cfg["discord_webhook_url"], "content": message})
                if not result.get("error"):
                    return {"sent_via": "discord", "result": result}
            elif ch == "email" and cfg.get("email"):
                result = _exec_email_send({"to": cfg["email"], "subject": f"[Janovum] {inp.get('priority', 'normal').upper()}", "body": message}, cfg)
                if not result.get("error"):
                    return {"sent_via": "email", "result": result}
            elif ch == "sms" and cfg.get("twilio_phone"):
                # SMS only for high/urgent
                if inp.get("priority") in ("high", "urgent"):
                    result = _exec_sms_send({"to": cfg.get("owner_phone", ""), "body": message}, cfg)
                    if not result.get("error"):
                        return {"sent_via": "sms", "result": result}
        except Exception:
            continue

    return {"error": "No notification channel available. Configure telegram, discord, email, or sms in client config."}


# ── JSON Transform Executor ──

def _exec_json_transform(inp):
    """Transform JSON data."""
    data = inp["data"]
    expression = inp["expression"]
    try:
        result = eval(expression, {"__builtins__": {"len": len, "str": str, "int": int, "float": float,
                                                     "list": list, "dict": dict, "sorted": sorted,
                                                     "filter": filter, "map": map, "sum": sum,
                                                     "min": min, "max": max, "enumerate": enumerate,
                                                     "zip": zip, "range": range, "True": True, "False": False,
                                                     "None": None}},
                      {"data": data})
        return {"result": result}
    except Exception as e:
        return {"error": str(e)}


# ── QR Code Executor ──

def _exec_qr_generate(inp):
    """Generate QR code."""
    try:
        import qrcode
    except ImportError:
        # Fallback: use free API
        import requests
        from urllib.parse import quote
        url = f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&data={quote(inp['content'])}"
        resp = requests.get(url, timeout=10)
        img_dir = PLATFORM_DIR / "data" / "images"
        img_dir.mkdir(parents=True, exist_ok=True)
        filename = inp.get("save_as", f"qr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        filepath = img_dir / filename
        with open(filepath, "wb") as f:
            f.write(resp.content)
        return {"saved_to": str(filepath), "method": "api"}

    img_dir = PLATFORM_DIR / "data" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"qr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
    filepath = img_dir / filename

    qr = qrcode.make(inp["content"])
    qr.save(str(filepath))
    return {"saved_to": str(filepath), "method": "qrcode"}


# ── Screenshot Desktop Executor ──

def _exec_screenshot_desktop(inp):
    """Take a desktop screenshot."""
    try:
        from PIL import ImageGrab
    except ImportError:
        return {"error": "Pillow not installed. Run: pip install pillow"}

    img_dir = PLATFORM_DIR / "data" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"desktop_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
    filepath = img_dir / filename

    screenshot = ImageGrab.grab()
    screenshot.save(str(filepath))
    return {"saved_to": str(filepath), "size": screenshot.size}


# ── PDF Creation Executor ──

def _exec_pdf_create(inp):
    """Create a PDF from text/HTML."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas as pdf_canvas
    except ImportError:
        # Fallback: basic text file as PDF via fpdf
        try:
            from fpdf import FPDF
        except ImportError:
            return {"error": "No PDF library. Run: pip install fpdf2 or pip install reportlab"}

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        if inp.get("title"):
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, inp["title"], ln=True)
            pdf.set_font("Helvetica", size=12)
            pdf.ln(5)
        pdf.multi_cell(0, 7, inp["content"])

        doc_dir = PLATFORM_DIR / "data" / "documents"
        doc_dir.mkdir(parents=True, exist_ok=True)
        filename = inp.get("save_as", f"doc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        filepath = doc_dir / filename
        pdf.output(str(filepath))
        return {"saved_to": str(filepath)}

    doc_dir = PLATFORM_DIR / "data" / "documents"
    doc_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"doc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    filepath = doc_dir / filename

    c = pdf_canvas.Canvas(str(filepath), pagesize=letter)
    if inp.get("title"):
        c.setFont("Helvetica-Bold", 16)
        c.drawString(72, 750, inp["title"])
        y = 720
    else:
        y = 750

    c.setFont("Helvetica", 12)
    for line in inp["content"].split("\n"):
        if y < 72:
            c.showPage()
            y = 750
        c.drawString(72, y, line[:100])
        y -= 15

    c.save()
    return {"saved_to": str(filepath)}


# ── Clipboard Executors ──

def _exec_clipboard_read(inp):
    """Read clipboard."""
    try:
        result = subprocess.run(["powershell", "-command", "Get-Clipboard"], capture_output=True, text=True, timeout=5)
        return {"content": result.stdout.strip()}
    except Exception as e:
        return {"error": str(e)}


def _exec_clipboard_write(inp):
    """Write to clipboard."""
    try:
        process = subprocess.Popen(["powershell", "-command", "Set-Clipboard", "-Value", inp["text"]], timeout=5)
        process.wait()
        return {"copied": True, "text_length": len(inp["text"])}
    except Exception as e:
        return {"error": str(e)}


# ── Process Management Executors ──

def _exec_process_list(inp):
    """List running processes."""
    try:
        cmd = "tasklist /FO CSV /NH"
        if inp.get("filter"):
            cmd += f' /FI "IMAGENAME eq *{inp["filter"]}*"'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
        lines = result.stdout.strip().split("\n")
        processes = []
        for line in lines[:50]:
            parts = line.replace('"', '').split(",")
            if len(parts) >= 2:
                processes.append({"name": parts[0], "pid": parts[1]})
        return {"processes": processes, "count": len(processes)}
    except Exception as e:
        return {"error": str(e)}


def _exec_process_kill(inp):
    """Kill a process."""
    target = inp["target"]
    try:
        if target.isdigit():
            result = subprocess.run(f"taskkill /PID {target} /F", shell=True, capture_output=True, text=True, timeout=10)
        else:
            result = subprocess.run(f"taskkill /IM {target} /F", shell=True, capture_output=True, text=True, timeout=10)
        return {"killed": result.returncode == 0, "stdout": result.stdout, "stderr": result.stderr}
    except Exception as e:
        return {"error": str(e)}


# ── System Info Executor ──

def _exec_system_info(inp):
    """Get system information."""
    import platform
    detail = inp.get("detail", "all")

    info = {}
    if detail in ("all", "os"):
        info["os"] = {"system": platform.system(), "release": platform.release(),
                       "version": platform.version(), "machine": platform.machine()}
    if detail in ("all", "python"):
        info["python"] = {"version": platform.python_version(), "executable": sys.executable}
    if detail in ("all", "cpu"):
        info["cpu"] = {"processor": platform.processor(), "cores": os.cpu_count()}
    if detail in ("all", "ram"):
        try:
            import psutil
            mem = psutil.virtual_memory()
            info["ram"] = {"total_gb": round(mem.total / 1e9, 1), "available_gb": round(mem.available / 1e9, 1),
                           "percent_used": mem.percent}
        except ImportError:
            info["ram"] = {"note": "Install psutil for RAM info: pip install psutil"}
    if detail in ("all", "disk"):
        try:
            import shutil
            total, used, free = shutil.disk_usage("/")
            info["disk"] = {"total_gb": round(total / 1e9, 1), "used_gb": round(used / 1e9, 1),
                            "free_gb": round(free / 1e9, 1)}
        except Exception:
            pass
    if detail in ("all", "packages"):
        result = subprocess.run([sys.executable, "-m", "pip", "list", "--format=json"],
                                capture_output=True, text=True, timeout=15)
        try:
            info["packages"] = json.loads(result.stdout)[:50]
        except Exception:
            info["packages"] = []

    return info


# ── Wait Executor ──

def _exec_wait(inp):
    """Wait for specified seconds."""
    seconds = min(float(inp["seconds"]), 300)  # Max 5 minutes
    time.sleep(seconds)
    return {"waited": seconds}


# ── Video / Camera Executors ──

_active_streams = {}  # stream_id -> {"running": bool, "thread": Thread}


def _exec_camera_capture(inp):
    """Capture photo from webcam."""
    try:
        import cv2
    except ImportError:
        return {"error": "OpenCV not installed. Run: pip install opencv-python"}

    cam_idx = inp.get("camera_index", 0)
    cap = cv2.VideoCapture(cam_idx)
    if not cap.isOpened():
        return {"error": f"Cannot open camera {cam_idx}"}

    ret, frame = cap.read()
    cap.release()

    if not ret:
        return {"error": "Failed to capture frame"}

    img_dir = PLATFORM_DIR / "data" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"cam_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
    filepath = img_dir / filename
    cv2.imwrite(str(filepath), frame)
    return {"saved_to": str(filepath), "resolution": f"{frame.shape[1]}x{frame.shape[0]}"}


def _exec_video_record(inp):
    """Record video from webcam."""
    try:
        import cv2
    except ImportError:
        return {"error": "OpenCV not installed. Run: pip install opencv-python"}

    cam_idx = inp.get("camera_index", 0)
    duration = min(inp["duration"], 300)  # Max 5 minutes

    cap = cv2.VideoCapture(cam_idx)
    if not cap.isOpened():
        return {"error": f"Cannot open camera {cam_idx}"}

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    vid_dir = PLATFORM_DIR / "data" / "videos"
    vid_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"vid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4")
    filepath = vid_dir / filename

    fps = 20
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    out = cv2.VideoWriter(str(filepath), fourcc, fps, (w, h))

    start = time.time()
    frames = 0
    while time.time() - start < duration:
        ret, frame = cap.read()
        if not ret:
            break
        out.write(frame)
        frames += 1

    cap.release()
    out.release()
    return {"saved_to": str(filepath), "duration": duration, "frames": frames, "resolution": f"{w}x{h}"}


def _exec_video_stream_start(inp):
    """Start live video stream — saves frames to disk."""
    import threading
    try:
        import cv2
    except ImportError:
        return {"error": "OpenCV not installed. Run: pip install opencv-python"}

    stream_id = inp["stream_id"]
    if stream_id in _active_streams and _active_streams[stream_id].get("running"):
        return {"error": f"Stream {stream_id} already running"}

    fps = inp.get("fps", 2)
    cam_idx = inp.get("camera_index", 0)
    stream_dir = PLATFORM_DIR / "data" / "streams" / stream_id
    stream_dir.mkdir(parents=True, exist_ok=True)

    def _stream_loop():
        cap = cv2.VideoCapture(cam_idx)
        frame_count = 0
        while _active_streams.get(stream_id, {}).get("running", False):
            ret, frame = cap.read()
            if ret:
                cv2.imwrite(str(stream_dir / "latest.jpg"), frame)
                frame_count += 1
                # Save metadata
                meta = {"stream_id": stream_id, "frame": frame_count,
                        "timestamp": datetime.now().isoformat(), "status": "streaming"}
                with open(stream_dir / "meta.json", "w") as f:
                    json.dump(meta, f)
            time.sleep(1.0 / fps)
        cap.release()
        meta = {"stream_id": stream_id, "frame": frame_count, "status": "stopped",
                "timestamp": datetime.now().isoformat()}
        with open(stream_dir / "meta.json", "w") as f:
            json.dump(meta, f)

    _active_streams[stream_id] = {"running": True}
    t = threading.Thread(target=_stream_loop, daemon=True)
    t.start()
    _active_streams[stream_id]["thread"] = t

    return {"started": True, "stream_id": stream_id, "stream_dir": str(stream_dir),
            "view_latest": str(stream_dir / "latest.jpg")}


def _exec_video_stream_stop(inp):
    """Stop a live video stream."""
    stream_id = inp["stream_id"]
    if stream_id in _active_streams:
        _active_streams[stream_id]["running"] = False
        return {"stopped": True, "stream_id": stream_id}
    return {"error": f"Stream {stream_id} not found"}


def _exec_screen_record(inp):
    """Record the desktop screen."""
    try:
        import cv2
        import numpy as np
        from PIL import ImageGrab
    except ImportError:
        return {"error": "Requires: pip install opencv-python pillow numpy"}

    duration = min(inp["duration"], 300)
    fps = inp.get("fps", 10)

    vid_dir = PLATFORM_DIR / "data" / "videos"
    vid_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"screen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4")
    filepath = vid_dir / filename

    # Get screen size
    screen = ImageGrab.grab()
    w, h = screen.size

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(str(filepath), fourcc, fps, (w, h))

    start = time.time()
    frames = 0
    while time.time() - start < duration:
        img = ImageGrab.grab()
        frame = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
        out.write(frame)
        frames += 1
        time.sleep(1.0 / fps)

    out.release()
    return {"saved_to": str(filepath), "duration": duration, "frames": frames, "resolution": f"{w}x{h}"}


# ── Audio Recording Executors ──

def _exec_audio_record(inp):
    """Record audio from microphone."""
    try:
        import sounddevice as sd
        import scipy.io.wavfile as wav
    except ImportError:
        # Fallback: use ffmpeg
        audio_dir = PLATFORM_DIR / "data" / "audio"
        audio_dir.mkdir(parents=True, exist_ok=True)
        filename = inp.get("save_as", f"rec_{datetime.now().strftime('%Y%m%d_%H%M%S')}.wav")
        filepath = audio_dir / filename
        duration = min(inp["duration"], 300)
        try:
            result = subprocess.run(
                ["ffmpeg", "-f", "dshow", "-i", "audio=Microphone", "-t", str(duration), str(filepath)],
                capture_output=True, text=True, timeout=duration + 10
            )
            if result.returncode == 0:
                return {"saved_to": str(filepath), "method": "ffmpeg"}
            return {"error": f"ffmpeg failed: {result.stderr[:500]}"}
        except FileNotFoundError:
            return {"error": "Neither sounddevice nor ffmpeg available. Run: pip install sounddevice scipy"}

    duration = min(inp["duration"], 300)
    sample_rate = 44100

    recording = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='int16')
    sd.wait()

    audio_dir = PLATFORM_DIR / "data" / "audio"
    audio_dir.mkdir(parents=True, exist_ok=True)
    filename = inp.get("save_as", f"rec_{datetime.now().strftime('%Y%m%d_%H%M%S')}.wav")
    filepath = audio_dir / filename
    wav.write(str(filepath), sample_rate, recording)
    return {"saved_to": str(filepath), "duration": duration, "sample_rate": sample_rate}


def _exec_audio_play(inp):
    """Play an audio file."""
    filepath = inp["file_path"]
    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}
    try:
        # Windows: use start command
        subprocess.Popen(["start", "", filepath], shell=True)
        return {"playing": True, "file": filepath}
    except Exception as e:
        return {"error": str(e)}


# ── Compression Executors ──

def _exec_zip_create(inp):
    """Create ZIP archive."""
    import zipfile
    save_as = inp["save_as"]
    if not save_as.endswith(".zip"):
        save_as += ".zip"

    with zipfile.ZipFile(save_as, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in inp["paths"]:
            if os.path.isdir(path):
                for root, dirs, files in os.walk(path):
                    for f in files:
                        full = os.path.join(root, f)
                        arcname = os.path.relpath(full, os.path.dirname(path))
                        zf.write(full, arcname)
            elif os.path.isfile(path):
                zf.write(path, os.path.basename(path))

    return {"created": True, "path": save_as, "size_bytes": os.path.getsize(save_as)}


def _exec_zip_extract(inp):
    """Extract ZIP archive."""
    import zipfile
    zip_path = inp["zip_path"]
    extract_to = inp["extract_to"]

    if not os.path.exists(zip_path):
        return {"error": f"ZIP not found: {zip_path}"}

    os.makedirs(extract_to, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_to)
        files = zf.namelist()

    return {"extracted": True, "extract_to": extract_to, "files": files[:50], "count": len(files)}


# ── Translation Executor ──

def _exec_translate(inp):
    """Translate text."""
    import requests
    text = inp["text"]
    from_lang = inp.get("from_lang", "en")
    to_lang = inp["to_lang"]

    # Try MyMemory API (free, 5000 chars/day)
    try:
        url = f"https://api.mymemory.translated.net/get?q={text[:500]}&langpair={from_lang}|{to_lang}"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        translated = data.get("responseData", {}).get("translatedText", "")
        return {"original": text, "translated": translated, "from": from_lang, "to": to_lang}
    except Exception as e:
        return {"error": str(e)}


# ── Weather Executor ──

def _exec_weather_get(inp):
    """Get weather data."""
    import requests
    location = inp["location"]

    # First geocode the location
    try:
        geo_url = f"https://geocoding-api.open-meteo.com/v1/search?name={location}&count=1"
        geo = requests.get(geo_url, timeout=10).json()
        if not geo.get("results"):
            return {"error": f"Location not found: {location}"}

        lat = geo["results"][0]["latitude"]
        lon = geo["results"][0]["longitude"]
        name = geo["results"][0].get("name", location)

        # Get weather
        wx_url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current_weather=true&temperature_unit=fahrenheit"
        wx = requests.get(wx_url, timeout=10).json()
        current = wx.get("current_weather", {})

        return {
            "location": name,
            "temperature_f": current.get("temperature"),
            "windspeed_mph": current.get("windspeed"),
            "wind_direction": current.get("winddirection"),
            "weather_code": current.get("weathercode"),
            "time": current.get("time")
        }
    except Exception as e:
        return {"error": str(e)}


# ── News Executor ──

def _exec_news_get(inp):
    """Get news headlines from Google News RSS."""
    import requests
    import xml.etree.ElementTree as ET
    from urllib.parse import quote

    query = inp.get("query", "")
    count = inp.get("count", 10)

    url = f"https://news.google.com/rss/search?q={quote(query)}&hl=en-US&gl=US&ceid=US:en" if query else "https://news.google.com/rss?hl=en-US&gl=US&ceid=US:en"

    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Janovum/1.0"})
        root = ET.fromstring(resp.content)
        items = root.findall(".//item")

        articles = []
        for item in items[:count]:
            articles.append({
                "title": item.findtext("title", ""),
                "link": item.findtext("link", ""),
                "pub_date": item.findtext("pubDate", ""),
                "source": item.findtext("source", "")
            })

        return {"articles": articles, "count": len(articles), "query": query}
    except Exception as e:
        return {"error": str(e)}


# ── Finance Executor ──

def _exec_stock_price(inp):
    """Get stock/crypto price."""
    import requests
    symbol = inp["symbol"].upper()

    # Try CoinGecko for crypto
    crypto_map = {"BTC": "bitcoin", "ETH": "ethereum", "DOGE": "dogecoin", "SOL": "solana",
                  "ADA": "cardano", "XRP": "ripple", "DOT": "polkadot", "AVAX": "avalanche-2"}

    if symbol.lower() in crypto_map.values() or symbol in crypto_map:
        coin_id = crypto_map.get(symbol, symbol.lower())
        try:
            url = f"https://api.coingecko.com/api/v3/simple/price?ids={coin_id}&vs_currencies=usd&include_24hr_change=true"
            resp = requests.get(url, timeout=10).json()
            if coin_id in resp:
                return {"symbol": symbol, "price_usd": resp[coin_id].get("usd"),
                        "change_24h": resp[coin_id].get("usd_24h_change"), "type": "crypto"}
        except Exception:
            pass

    # Try Yahoo Finance scraping
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=1d"
        resp = requests.get(url, headers={"User-Agent": "Janovum/1.0"}, timeout=10).json()
        meta = resp.get("chart", {}).get("result", [{}])[0].get("meta", {})
        return {
            "symbol": symbol,
            "price": meta.get("regularMarketPrice"),
            "previous_close": meta.get("previousClose"),
            "currency": meta.get("currency"),
            "exchange": meta.get("exchangeName"),
            "type": "stock"
        }
    except Exception as e:
        return {"error": f"Could not fetch price for {symbol}: {str(e)}"}


# ── Network Executors ──

def _exec_whois_lookup(inp):
    """WHOIS lookup."""
    domain = inp["domain"]
    try:
        result = subprocess.run(["whois", domain], capture_output=True, text=True, timeout=15)
        return {"domain": domain, "data": result.stdout[:5000]}
    except FileNotFoundError:
        # Fallback: use web API
        import requests
        try:
            resp = requests.get(f"https://api.api-ninjas.com/v1/whois?domain={domain}",
                                headers={"User-Agent": "Janovum/1.0"}, timeout=10)
            return {"domain": domain, "data": resp.text[:5000]}
        except Exception as e:
            return {"error": str(e)}


def _exec_dns_lookup(inp):
    """DNS lookup."""
    import socket
    domain = inp["domain"]
    record_type = inp.get("record_type", "A")

    try:
        if record_type == "A":
            result = socket.getaddrinfo(domain, None, socket.AF_INET)
            ips = list(set(r[4][0] for r in result))
            return {"domain": domain, "type": "A", "records": ips}
        else:
            # Use nslookup for other types
            result = subprocess.run(["nslookup", f"-type={record_type}", domain],
                                    capture_output=True, text=True, timeout=10)
            return {"domain": domain, "type": record_type, "output": result.stdout[:3000]}
    except Exception as e:
        return {"error": str(e)}


# ── Encoding / Hashing Executors ──

def _exec_hash_text(inp):
    """Hash text."""
    import hashlib
    text = inp["text"].encode("utf-8")
    algo = inp.get("algorithm", "sha256")

    algos = {"md5": hashlib.md5, "sha1": hashlib.sha1, "sha256": hashlib.sha256, "sha512": hashlib.sha512}
    if algo not in algos:
        return {"error": f"Unknown algorithm: {algo}. Use: md5, sha1, sha256, sha512"}

    return {"hash": algos[algo](text).hexdigest(), "algorithm": algo}


def _exec_encrypt_text(inp):
    """Encrypt text with Fernet."""
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        return {"error": "cryptography not installed. Run: pip install cryptography"}

    if inp.get("key"):
        key = inp["key"].encode()
    else:
        key = Fernet.generate_key()

    f = Fernet(key)
    encrypted = f.encrypt(inp["text"].encode())
    return {"encrypted": encrypted.decode(), "key": key.decode()}


def _exec_decrypt_text(inp):
    """Decrypt text with Fernet."""
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        return {"error": "cryptography not installed. Run: pip install cryptography"}

    f = Fernet(inp["key"].encode())
    decrypted = f.decrypt(inp["encrypted"].encode())
    return {"decrypted": decrypted.decode()}


def _exec_base64_encode(inp):
    """Base64 encode."""
    import base64
    if inp.get("is_file") and os.path.exists(inp["text"]):
        with open(inp["text"], "rb") as f:
            data = f.read()
    else:
        data = inp["text"].encode()
    return {"encoded": base64.b64encode(data).decode()}


def _exec_base64_decode(inp):
    """Base64 decode."""
    import base64
    try:
        decoded = base64.b64decode(inp["encoded"])
        return {"decoded": decoded.decode("utf-8", errors="replace")}
    except Exception as e:
        return {"error": str(e)}


# ── Image Editing Executors ──

def _exec_image_resize(inp):
    """Resize an image."""
    try:
        from PIL import Image
    except ImportError:
        return {"error": "Pillow not installed. Run: pip install pillow"}

    if not os.path.exists(inp["image_path"]):
        return {"error": f"Image not found: {inp['image_path']}"}

    img = Image.open(inp["image_path"])
    img = img.resize((inp["width"], inp["height"]), Image.LANCZOS)

    save_path = inp.get("save_as", inp["image_path"])
    img.save(save_path)
    return {"saved_to": save_path, "size": f"{inp['width']}x{inp['height']}"}


def _exec_image_text_overlay(inp):
    """Add text overlay to image."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return {"error": "Pillow not installed. Run: pip install pillow"}

    if not os.path.exists(inp["image_path"]):
        return {"error": f"Image not found: {inp['image_path']}"}

    img = Image.open(inp["image_path"]).convert("RGBA")
    draw = ImageDraw.Draw(img)

    font_size = inp.get("font_size", 36)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except Exception:
        font = ImageFont.load_default()

    text = inp["text"]
    position = inp.get("position", "bottom")
    color = inp.get("color", "white")

    # Calculate position
    bbox = draw.textbbox((0, 0), text, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    iw, ih = img.size

    if position == "top":
        xy = ((iw - tw) // 2, 10)
    elif position == "center":
        xy = ((iw - tw) // 2, (ih - th) // 2)
    else:  # bottom
        xy = ((iw - tw) // 2, ih - th - 10)

    # Draw shadow then text
    draw.text((xy[0] + 2, xy[1] + 2), text, fill="black", font=font)
    draw.text(xy, text, fill=color, font=font)

    img_dir = PLATFORM_DIR / "data" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    save_path = inp.get("save_as", str(img_dir / f"overlay_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"))
    img.save(save_path)
    return {"saved_to": save_path}


def _exec_image_convert(inp):
    """Convert image format."""
    try:
        from PIL import Image
    except ImportError:
        return {"error": "Pillow not installed. Run: pip install pillow"}

    if not os.path.exists(inp["image_path"]):
        return {"error": f"Image not found: {inp['image_path']}"}

    img = Image.open(inp["image_path"])
    fmt = inp["format"].upper()
    if fmt == "JPG":
        fmt = "JPEG"

    img_dir = PLATFORM_DIR / "data" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    base = os.path.splitext(os.path.basename(inp["image_path"]))[0]
    save_path = inp.get("save_as", str(img_dir / f"{base}.{inp['format'].lower()}"))

    if fmt == "JPEG" and img.mode == "RGBA":
        img = img.convert("RGB")

    img.save(save_path, format=fmt)
    return {"saved_to": save_path, "format": fmt}


# ── AI / LLM Executors ──

def _exec_ai_ask(inp):
    """Ask an AI model."""
    from .engine import quick_ask, MODELS
    model = inp.get("model", "auto")
    force = None
    if model in MODELS:
        force = MODELS[model]
    return {"response": quick_ask(inp["prompt"], system_prompt=inp.get("system"), force_model=force)}


def _exec_ai_summarize(inp):
    """Summarize content."""
    from .engine import quick_ask
    content = inp["content"]
    style = inp.get("style", "brief")
    max_len = inp.get("max_length", 200)

    # Check if it's a URL
    if content.startswith("http"):
        from modules.web_search import fetch_url
        fetched = fetch_url(content)
        content = fetched.get("text", content)[:8000]
    # Check if it's a file
    elif os.path.exists(content):
        with open(content, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()[:8000]

    prompt = f"Summarize the following in {style} style, max {max_len} words:\n\n{content}"
    return {"summary": quick_ask(prompt)}


def _exec_ai_classify(inp):
    """Classify text."""
    from .engine import quick_ask
    cats = ", ".join(inp["categories"])
    prompt = f"Classify this text into ONE of these categories: {cats}\n\nText: {inp['text']}\n\nRespond with ONLY the category name."
    result = quick_ask(prompt)
    return {"text": inp["text"][:100], "category": result.strip(), "categories": inp["categories"]}


def _exec_ai_extract(inp):
    """Extract structured data from text."""
    from .engine import quick_ask
    fields = ", ".join(inp["fields"])
    prompt = f"Extract these fields from the text: {fields}\n\nText: {inp['text']}\n\nRespond in JSON format with the field names as keys. If a field is not found, use null."
    result = quick_ask(prompt)
    try:
        extracted = json.loads(result)
    except Exception:
        extracted = {"raw_response": result}
    return {"extracted": extracted, "fields": inp["fields"]}


def _exec_ai_generate_code(inp):
    """Generate and optionally execute code."""
    from .engine import quick_ask
    lang = inp.get("language", "python")
    prompt = f"Write {lang} code that does the following: {inp['description']}\n\nRespond with ONLY the code, no explanation."
    code = quick_ask(prompt)

    # Clean markdown code blocks
    if "```" in code:
        lines = code.split("\n")
        code_lines = []
        in_block = False
        for line in lines:
            if line.startswith("```"):
                in_block = not in_block
                continue
            if in_block:
                code_lines.append(line)
        code = "\n".join(code_lines)

    result = {"code": code, "language": lang}

    if inp.get("execute") and lang == "python":
        exec_result = _exec_code_execute({"code": code, "timeout": 30})
        result["execution"] = exec_result

    return result


# ── Workflow Executors ──

WORKFLOW_DIR = PLATFORM_DIR / "data" / "workflows"


def _exec_workflow_create(inp):
    """Create a workflow."""
    WORKFLOW_DIR.mkdir(parents=True, exist_ok=True)
    workflow = {
        "name": inp["name"],
        "description": inp.get("description", ""),
        "steps": inp["steps"],
        "created_at": datetime.now().isoformat()
    }
    filepath = WORKFLOW_DIR / f"{inp['name']}.json"
    with open(filepath, "w") as f:
        json.dump(workflow, f, indent=2)
    return {"created": True, "name": inp["name"], "steps": len(inp["steps"])}


def _exec_workflow_run(inp, client_id="default", client_config=None):
    """Execute a saved workflow."""
    filepath = WORKFLOW_DIR / f"{inp['name']}.json"
    if not filepath.exists():
        return {"error": f"Workflow not found: {inp['name']}"}

    with open(filepath) as f:
        workflow = json.load(f)

    variables = inp.get("variables", {})
    results = []

    for i, step in enumerate(workflow["steps"]):
        # Check condition
        if step.get("condition"):
            try:
                if not eval(step["condition"], {"__builtins__": {}}, {"variables": variables, "results": results}):
                    results.append({"step": i, "skipped": True, "reason": "condition not met"})
                    continue
            except Exception:
                pass

        # Replace variables in input
        step_input = json.loads(json.dumps(step["input"]))
        for key, val in step_input.items():
            if isinstance(val, str):
                for var_name, var_val in variables.items():
                    step_input[key] = step_input[key].replace(f"{{{{{var_name}}}}}", str(var_val))

        # Execute
        tool_result = execute_tool(step["tool"], step_input, client_id, client_config)
        try:
            parsed = json.loads(tool_result)
        except Exception:
            parsed = {"raw": tool_result}
        results.append({"step": i, "tool": step["tool"], "result": parsed})

    return {"workflow": inp["name"], "steps_run": len(results), "results": results}


def _exec_workflow_list(inp):
    """List all workflows."""
    if not WORKFLOW_DIR.exists():
        return {"workflows": [], "count": 0}
    workflows = []
    for f in WORKFLOW_DIR.glob("*.json"):
        with open(f) as fh:
            wf = json.load(fh)
        workflows.append({"name": wf["name"], "description": wf.get("description"), "steps": len(wf.get("steps", []))})
    return {"workflows": workflows, "count": len(workflows)}


# ── Database Executors ──

def _exec_db_query(inp):
    """Execute SQL query."""
    import sqlite3
    db_name = inp.get("database", "toolkit.db")
    db_path = PLATFORM_DIR / "data" / db_name

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.execute(inp["sql"])
        if inp["sql"].strip().upper().startswith("SELECT"):
            rows = [dict(row) for row in cursor.fetchall()[:500]]
            return {"rows": rows, "count": len(rows)}
        else:
            conn.commit()
            return {"affected_rows": cursor.rowcount, "success": True}
    except Exception as e:
        return {"error": str(e)}
    finally:
        conn.close()


def _exec_db_tables(inp):
    """List database tables."""
    import sqlite3
    db_name = inp.get("database", "toolkit.db")
    db_path = PLATFORM_DIR / "data" / db_name
    if not db_path.exists():
        return {"tables": [], "note": "Database does not exist yet"}
    conn = sqlite3.connect(str(db_path))
    cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [row[0] for row in cursor.fetchall()]
    conn.close()
    return {"tables": tables, "count": len(tables)}


# ── Spreadsheet Executors ──

def _exec_spreadsheet_read(inp):
    """Read Excel file."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not os.path.exists(inp["file_path"]):
        return {"error": f"File not found: {inp['file_path']}"}

    wb = openpyxl.load_workbook(inp["file_path"], read_only=True)
    sheet_name = inp.get("sheet") or wb.sheetnames[0]
    ws = wb[sheet_name]
    max_rows = inp.get("max_rows", 100)

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows + 1:  # +1 for header
            break
        rows.append([str(c) if c is not None else "" for c in row])

    wb.close()
    headers = rows[0] if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return {"headers": headers, "rows": data, "row_count": len(data), "sheet": sheet_name}


def _exec_spreadsheet_write(inp):
    """Write Excel file."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = inp.get("sheet_name", "Sheet1")
    ws.append(inp["headers"])
    for row in inp["rows"]:
        ws.append(row)
    wb.save(inp["file_path"])
    return {"saved_to": inp["file_path"], "rows": len(inp["rows"]), "columns": len(inp["headers"])}


# ── Contact / CRM Executors ──

CONTACTS_DB = PLATFORM_DIR / "data" / "contacts.json"


def _load_contacts():
    if CONTACTS_DB.exists():
        with open(CONTACTS_DB) as f:
            return json.load(f)
    return []


def _save_contacts(contacts):
    CONTACTS_DB.parent.mkdir(parents=True, exist_ok=True)
    with open(CONTACTS_DB, "w") as f:
        json.dump(contacts, f, indent=2)


def _exec_contact_add(inp):
    contacts = _load_contacts()
    contact = {
        "id": f"c_{int(time.time())}",
        "name": inp["name"],
        "email": inp.get("email", ""),
        "phone": inp.get("phone", ""),
        "company": inp.get("company", ""),
        "notes": inp.get("notes", ""),
        "tags": inp.get("tags", []),
        "created_at": datetime.now().isoformat()
    }
    contacts.append(contact)
    _save_contacts(contacts)
    return {"added": True, "contact": contact}


def _exec_contact_search(inp):
    contacts = _load_contacts()
    query = inp["query"].lower()
    results = [c for c in contacts if query in json.dumps(c).lower()]
    return {"results": results, "count": len(results)}


def _exec_contact_list(inp):
    contacts = _load_contacts()
    tag = inp.get("tag")
    if tag:
        contacts = [c for c in contacts if tag in c.get("tags", [])]
    return {"contacts": contacts, "count": len(contacts)}


# ── Invoice Executor ──

def _exec_invoice_create(inp):
    """Create a PDF invoice."""
    items = inp["items"]
    total = sum(i.get("quantity", 1) * i.get("unit_price", 0) for i in items)

    invoice_num = f"INV-{datetime.now().strftime('%Y%m%d%H%M')}"
    content = f"""INVOICE {invoice_num}

To: {inp['client_name']}
{f"Email: {inp['client_email']}" if inp.get('client_email') else ""}
Date: {datetime.now().strftime('%Y-%m-%d')}
{f"Due: {inp['due_date']}" if inp.get('due_date') else ""}

Items:
{'=' * 60}
"""
    for item in items:
        qty = item.get("quantity", 1)
        price = item.get("unit_price", 0)
        line_total = qty * price
        content += f"{item.get('description', 'Item'):<35} {qty:>5} x ${price:>8.2f} = ${line_total:>10.2f}\n"

    content += f"{'=' * 60}\n"
    content += f"{'TOTAL':>52} ${total:>10.2f}\n"
    if inp.get("notes"):
        content += f"\nNotes: {inp['notes']}"

    # Generate PDF
    result = _exec_pdf_create({"content": content, "title": f"Invoice {invoice_num}",
                                "save_as": f"invoice_{invoice_num}.pdf"})

    return {"invoice_number": invoice_num, "total": total, "pdf": result.get("saved_to", ""),
            "client": inp["client_name"]}


# ── Report Executor ──

def _exec_report_generate(inp):
    """Generate a business report."""
    report_type = inp["report_type"]
    date_range = inp.get("date_range", "this_month")

    # Gather data based on type
    data = {"report_type": report_type, "date_range": date_range, "generated_at": datetime.now().isoformat()}

    if report_type == "clients":
        data["clients"] = _exec_client_list({})
    elif report_type == "revenue":
        # Read from cost tracker or invoices
        data["note"] = "Revenue report — connect to payment/invoice data for actual figures"

    content = f"JANOVUM BUSINESS REPORT\nType: {report_type}\nPeriod: {date_range}\nGenerated: {data['generated_at']}\n\n"
    content += json.dumps(data, indent=2, default=str)

    if inp.get("format") == "json":
        return data

    result = _exec_pdf_create({"content": content, "title": f"Report - {report_type}",
                                "save_as": f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.pdf"})
    data["pdf"] = result.get("saved_to", "")
    return data


# ── Appointment Executors ──

def _exec_appointment_book(inp):
    """Book an appointment."""
    appt_dir = PLATFORM_DIR / "data" / "clients" / inp["client_id"] / "appointments"
    appt_dir.mkdir(parents=True, exist_ok=True)

    appt = {
        "id": f"appt_{int(time.time())}",
        "customer_name": inp["customer_name"],
        "customer_phone": inp.get("customer_phone", ""),
        "service": inp.get("service", ""),
        "date": inp["date"],
        "time": inp["time"],
        "staff": inp.get("staff", ""),
        "notes": inp.get("notes", ""),
        "status": "confirmed",
        "created_at": datetime.now().isoformat()
    }

    filepath = appt_dir / f"{appt['id']}.json"
    with open(filepath, "w") as f:
        json.dump(appt, f, indent=2)

    return {"booked": True, "appointment": appt}


def _exec_appointment_list(inp):
    """List appointments."""
    appt_dir = PLATFORM_DIR / "data" / "clients" / inp["client_id"] / "appointments"
    if not appt_dir.exists():
        return {"appointments": [], "count": 0}

    appointments = []
    for f in appt_dir.glob("*.json"):
        with open(f) as fh:
            appt = json.load(fh)
        if inp.get("date") and appt.get("date") != inp["date"]:
            continue
        appointments.append(appt)

    appointments.sort(key=lambda a: f"{a.get('date', '')} {a.get('time', '')}")
    return {"appointments": appointments, "count": len(appointments)}


def _exec_appointment_cancel(inp):
    """Cancel an appointment."""
    appt_dir = PLATFORM_DIR / "data" / "clients" / inp["client_id"] / "appointments"
    filepath = appt_dir / f"{inp['appointment_id']}.json"
    if not filepath.exists():
        return {"error": "Appointment not found"}

    with open(filepath) as f:
        appt = json.load(f)
    appt["status"] = "cancelled"
    with open(filepath, "w") as f:
        json.dump(appt, f, indent=2)
    return {"cancelled": True, "appointment": appt}


# ── Template Executors ──

TEMPLATE_DIR = PLATFORM_DIR / "data" / "templates"


def _exec_template_list(inp):
    """List templates."""
    if not TEMPLATE_DIR.exists():
        return {"templates": [], "count": 0}
    templates = []
    category = inp.get("category")
    for f in TEMPLATE_DIR.glob("*.json"):
        with open(f) as fh:
            t = json.load(fh)
        if category and t.get("category") != category:
            continue
        templates.append({"name": t["name"], "category": t.get("category", "custom")})
    return {"templates": templates, "count": len(templates)}


def _exec_template_render(inp):
    """Render a template."""
    filepath = TEMPLATE_DIR / f"{inp['template_name']}.json"
    if not filepath.exists():
        return {"error": f"Template not found: {inp['template_name']}"}
    with open(filepath) as f:
        template = json.load(f)
    content = template["content"]
    for key, val in inp["variables"].items():
        content = content.replace(f"{{{{{key}}}}}", str(val))
    return {"rendered": content, "template": inp["template_name"]}


def _exec_template_create(inp):
    """Create a template."""
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    template = {
        "name": inp["name"],
        "category": inp.get("category", "custom"),
        "content": inp["content"],
        "created_at": datetime.now().isoformat()
    }
    filepath = TEMPLATE_DIR / f"{inp['name']}.json"
    with open(filepath, "w") as f:
        json.dump(template, f, indent=2)
    return {"created": True, "name": inp["name"]}


# ── Monitoring Executors ──

def _exec_monitor_url(inp):
    """Check if URL is up."""
    import requests
    url = inp["url"]
    expected = inp.get("expected_status", 200)
    try:
        start = time.time()
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Janovum-Monitor/1.0"})
        response_time_ms = round((time.time() - start) * 1000)
        return {
            "url": url, "status_code": resp.status_code,
            "is_up": resp.status_code == expected,
            "response_time_ms": response_time_ms
        }
    except Exception as e:
        return {"url": url, "is_up": False, "error": str(e)}


def _exec_monitor_port(inp):
    """Check if port is open."""
    import socket
    host = inp["host"]
    port = inp["port"]
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(5)
        result = sock.connect_ex((host, port))
        sock.close()
        return {"host": host, "port": port, "is_open": result == 0}
    except Exception as e:
        return {"host": host, "port": port, "is_open": False, "error": str(e)}


# ── Text Processing Executors ──

def _exec_regex_match(inp):
    """Find regex matches."""
    import re
    pattern = inp["pattern"]
    text = inp["text"]
    try:
        matches = re.findall(pattern, text)
        result = {"pattern": pattern, "matches": matches, "count": len(matches)}
        if inp.get("replace_with") is not None:
            result["replaced"] = re.sub(pattern, inp["replace_with"], text)
        return result
    except Exception as e:
        return {"error": str(e)}


def _exec_text_diff(inp):
    """Compare two texts."""
    import difflib
    lines1 = inp["text1"].splitlines()
    lines2 = inp["text2"].splitlines()
    diff = list(difflib.unified_diff(lines1, lines2, lineterm=""))
    return {"diff": "\n".join(diff), "lines_changed": len([l for l in diff if l.startswith("+") or l.startswith("-")])}


# ── Multi-Agent Orchestration Executors ──

_agents = {}  # agent_id -> {status, result, thread, ...}


def _exec_agent_spawn(inp, client_id="default", client_config=None):
    """Spawn a new AI agent."""
    import threading

    agent_id = inp["agent_id"]
    if agent_id in _agents and _agents[agent_id].get("status") == "running":
        return {"error": f"Agent {agent_id} already running"}

    _agents[agent_id] = {
        "status": "starting",
        "task": inp["task"],
        "provider": inp.get("provider", "claude"),
        "started_at": datetime.now().isoformat(),
        "result": None,
        "messages": []
    }

    def _agent_work():
        try:
            _agents[agent_id]["status"] = "running"

            # Build system prompt
            system = inp.get("system_prompt", f"You are Agent '{agent_id}', a Janovum AI assistant. Complete your assigned task.")

            # Get tools (all or filtered)
            if inp.get("tools"):
                all_tools = get_all_tools()
                filtered = [t for t in all_tools if t["name"] in inp["tools"]]
            else:
                filtered = get_all_tools()

            # Use the engine
            from .engine import agent_loop
            messages = [{"role": "user", "content": inp["task"]}]

            def tool_exec(name, tool_inp):
                return execute_tool(name, tool_inp, client_id, client_config)

            result = agent_loop(messages, system, filtered, tool_exec, max_turns=inp.get("max_turns", 10))
            _agents[agent_id]["result"] = result
            _agents[agent_id]["status"] = "completed"
        except Exception as e:
            _agents[agent_id]["result"] = f"Agent error: {str(e)}"
            _agents[agent_id]["status"] = "failed"

    if inp.get("parallel", True):
        t = threading.Thread(target=_agent_work, daemon=True)
        t.start()
        _agents[agent_id]["thread"] = t
        return {"spawned": True, "agent_id": agent_id, "status": "running", "parallel": True}
    else:
        _agent_work()
        return {"agent_id": agent_id, "status": _agents[agent_id]["status"],
                "result": _agents[agent_id]["result"]}


def _exec_agent_status(inp):
    """Check agent status."""
    agent_id = inp["agent_id"]
    if agent_id not in _agents:
        return {"error": f"Agent {agent_id} not found"}
    a = _agents[agent_id]
    return {"agent_id": agent_id, "status": a["status"], "task": a["task"],
            "started_at": a["started_at"], "provider": a.get("provider")}


def _exec_agent_result(inp):
    """Get agent result."""
    agent_id = inp["agent_id"]
    if agent_id not in _agents:
        return {"error": f"Agent {agent_id} not found"}

    a = _agents[agent_id]
    if inp.get("wait") and a["status"] == "running":
        thread = a.get("thread")
        if thread:
            thread.join(timeout=120)

    return {"agent_id": agent_id, "status": a["status"], "result": a.get("result")}


def _exec_agent_stop(inp):
    """Stop an agent."""
    agent_id = inp["agent_id"]
    if agent_id not in _agents:
        return {"error": f"Agent {agent_id} not found"}
    _agents[agent_id]["status"] = "stopped"
    return {"stopped": True, "agent_id": agent_id}


def _exec_agent_list(inp):
    """List all agents."""
    agents = []
    for aid, a in _agents.items():
        agents.append({"agent_id": aid, "status": a["status"], "task": a["task"][:100],
                       "started_at": a["started_at"], "provider": a.get("provider")})
    return {"agents": agents, "count": len(agents)}


def _exec_agent_message(inp):
    """Send message between agents."""
    to_id = inp["to_agent"]
    if to_id not in _agents:
        return {"error": f"Agent {to_id} not found"}
    _agents[to_id].setdefault("messages", []).append({
        "from": inp.get("from_agent", "director"),
        "message": inp["message"],
        "timestamp": datetime.now().isoformat()
    })
    return {"sent": True, "to": to_id}


def _exec_agent_team(inp, client_id="default", client_config=None):
    """Spawn a team of agents."""
    import threading

    team_name = inp["team_name"]
    results = []

    for i, agent_def in enumerate(inp["agents"]):
        agent_id = f"{team_name}_{agent_def['role']}_{i}"
        spawn_inp = {
            "agent_id": agent_id,
            "task": f"TEAM GOAL: {inp['goal']}\n\nYOUR ROLE: {agent_def['role']}\nYOUR TASK: {agent_def['task']}",
            "provider": agent_def.get("provider", "claude"),
            "tools": agent_def.get("tools"),
            "parallel": True
        }
        result = _exec_agent_spawn(spawn_inp, client_id, client_config)
        results.append(result)

    return {"team": team_name, "agents_spawned": len(results), "agents": results}


# ── Pre-Built Bot Setup Executors ──

BOT_DIR = PLATFORM_DIR / "data" / "bots"


def _save_bot_config(client_id, bot_type, config):
    """Save a bot configuration."""
    bot_dir = BOT_DIR / client_id
    bot_dir.mkdir(parents=True, exist_ok=True)
    config["bot_type"] = bot_type
    config["created_at"] = datetime.now().isoformat()
    config["status"] = "configured"
    filepath = bot_dir / f"{bot_type}.json"
    with open(filepath, "w") as f:
        json.dump(config, f, indent=2)
    return {"configured": True, "bot_type": bot_type, "client_id": client_id, "path": str(filepath)}


def _exec_bot_setup_receptionist(inp):
    """Set up AI receptionist."""
    config = {
        "business_name": inp["business_name"],
        "business_type": inp["business_type"],
        "phone": inp.get("phone", ""),
        "services": inp.get("services", []),
        "business_hours": inp.get("business_hours", "Mon-Fri 9am-5pm"),
        "staff": inp.get("staff", []),
        "greeting": inp.get("greeting", f"Thank you for calling {inp['business_name']}! How can I help you today?"),
        "capabilities": ["answer_calls", "book_appointments", "answer_questions", "take_messages", "transfer_calls"]
    }
    return _save_bot_config(inp["client_id"], "receptionist", config)


def _exec_bot_setup_email_assistant(inp):
    """Set up email auto-responder."""
    config = {
        "email": inp["email"],
        "email_password": inp["email_password"],
        "business_name": inp["business_name"],
        "business_context": inp.get("business_context", ""),
        "auto_send": inp.get("auto_send", False),
        "imap_server": "imap.gmail.com",
        "smtp_server": "smtp.gmail.com",
        "check_interval": 60
    }
    return _save_bot_config(inp["client_id"], "email_assistant", config)


def _exec_bot_setup_lead_gen(inp):
    """Set up lead generation bot."""
    config = {
        "target_industry": inp["target_industry"],
        "target_location": inp["target_location"],
        "outreach_template": inp.get("outreach_template", ""),
        "max_leads_per_day": inp.get("max_leads_per_day", 20),
        "sources": ["google_maps", "yelp", "web_search"],
        "enrich": True
    }
    return _save_bot_config(inp["client_id"], "lead_gen", config)


def _exec_bot_setup_social_media(inp):
    """Set up social media bot."""
    config = {
        "platforms": inp["platforms"],
        "business_name": inp["business_name"],
        "tone": inp.get("tone", "professional"),
        "post_frequency": inp.get("post_frequency", "3x_week"),
        "capabilities": ["schedule_posts", "auto_reply_comments", "monitor_mentions", "generate_content"]
    }
    return _save_bot_config(inp["client_id"], "social_media", config)


def _exec_bot_setup_review_manager(inp):
    """Set up review management bot."""
    config = {
        "business_name": inp["business_name"],
        "platforms": inp.get("platforms", ["google", "yelp"]),
        "auto_respond": inp.get("auto_respond", False),
        "capabilities": ["monitor_reviews", "generate_responses", "request_reviews", "sentiment_analysis"]
    }
    return _save_bot_config(inp["client_id"], "review_manager", config)


def _exec_bot_setup_scheduler(inp):
    """Set up scheduling bot."""
    config = {
        "business_name": inp["business_name"],
        "services": inp["services"],
        "slot_duration_minutes": inp.get("slot_duration_minutes", 30),
        "business_hours": inp.get("business_hours", "Mon-Fri 9am-5pm"),
        "reminder_before_minutes": inp.get("reminder_before_minutes", 60),
        "capabilities": ["online_booking", "sms_reminders", "email_reminders", "calendar_sync", "waitlist"]
    }
    return _save_bot_config(inp["client_id"], "scheduler", config)


def _exec_bot_setup_chat_widget(inp):
    """Set up website chat widget."""
    config = {
        "business_name": inp["business_name"],
        "website_url": inp.get("website_url", ""),
        "faq": inp.get("faq", []),
        "escalation_email": inp.get("escalation_email", ""),
        "capabilities": ["answer_questions", "collect_leads", "book_appointments", "live_chat_escalation"]
    }
    return _save_bot_config(inp["client_id"], "chat_widget", config)


def _exec_bot_setup_invoice_collector(inp):
    """Set up invoice/payment reminder bot."""
    config = {
        "business_name": inp["business_name"],
        "reminder_days": inp.get("reminder_days", [7, 3, 1, -1, -7]),
        "payment_link": inp.get("payment_link", ""),
        "capabilities": ["send_invoices", "payment_reminders", "overdue_notices", "receipt_generation"]
    }
    return _save_bot_config(inp["client_id"], "invoice_collector", config)


def _exec_bot_setup_generic(inp, bot_type, capabilities):
    """Generic bot setup — works for any AI employee type."""
    config = dict(inp)
    config.pop("client_id", None)
    config["capabilities"] = capabilities
    return _save_bot_config(inp["client_id"], bot_type, config)


def _exec_bot_setup_accountant(inp):
    return _exec_bot_setup_generic(inp, "accountant", [
        "income_tracking", "expense_categorization", "financial_reports", "tax_prep",
        "profit_loss_statements", "balance_sheets", "quarterly_reports", "receipt_scanning"])

def _exec_bot_setup_bookkeeper(inp):
    return _exec_bot_setup_generic(inp, "bookkeeper", [
        "transaction_logging", "receipt_ocr", "bank_reconciliation", "p_and_l",
        "cash_flow_tracking", "vendor_payments", "payroll_records"])

def _exec_bot_setup_legal_assistant(inp):
    return _exec_bot_setup_generic(inp, "legal_assistant", [
        "contract_review", "tos_generation", "nda_drafting", "compliance_checks",
        "legal_research", "cease_desist", "trademark_search", "privacy_policy"])

def _exec_bot_setup_financial_advisor(inp):
    return _exec_bot_setup_generic(inp, "financial_advisor", [
        "cash_flow_forecasting", "budget_planning", "investment_analysis",
        "financial_health_score", "scenario_modeling", "debt_reduction_plan"])

def _exec_bot_setup_debt_collector(inp):
    return _exec_bot_setup_generic(inp, "debt_collector", [
        "overdue_tracking", "escalating_reminders", "payment_plan_negotiation",
        "collection_reporting", "aging_reports"])

def _exec_bot_setup_sales_rep(inp):
    return _exec_bot_setup_generic(inp, "sales_rep", [
        "cold_outreach", "follow_ups", "pipeline_management", "objection_handling",
        "proposal_sending", "deal_closing", "crm_updates", "commission_tracking"])

def _exec_bot_setup_marketing_manager(inp):
    return _exec_bot_setup_generic(inp, "marketing_manager", [
        "campaign_planning", "content_calendar", "analytics_reporting", "ab_testing",
        "multi_channel_strategy", "roi_tracking", "audience_segmentation"])

def _exec_bot_setup_seo_specialist(inp):
    return _exec_bot_setup_generic(inp, "seo_specialist", [
        "keyword_research", "on_page_optimization", "backlink_analysis", "rank_tracking",
        "content_recommendations", "technical_seo_audit", "competitor_seo_analysis"])

def _exec_bot_setup_content_writer(inp):
    return _exec_bot_setup_generic(inp, "content_writer", [
        "blog_posts", "articles", "newsletters", "website_copy", "product_descriptions",
        "case_studies", "whitepapers", "social_media_captions"])

def _exec_bot_setup_copywriter(inp):
    return _exec_bot_setup_generic(inp, "copywriter", [
        "ad_copy", "landing_pages", "email_sequences", "sales_pages", "headlines",
        "cta_optimization", "a_b_variants", "brand_messaging"])

def _exec_bot_setup_email_marketer(inp):
    return _exec_bot_setup_generic(inp, "email_marketer", [
        "drip_campaigns", "newsletter_creation", "list_segmentation", "ab_subject_lines",
        "open_rate_optimization", "unsubscribe_analysis", "re_engagement_campaigns"])

def _exec_bot_setup_pr_manager(inp):
    return _exec_bot_setup_generic(inp, "pr_manager", [
        "press_releases", "media_outreach", "brand_monitoring", "crisis_communication",
        "media_list_building", "press_kit_creation", "interview_prep"])

def _exec_bot_setup_brand_manager(inp):
    return _exec_bot_setup_generic(inp, "brand_manager", [
        "brand_guidelines", "voice_consistency", "visual_identity", "brand_health_monitoring",
        "asset_management", "brand_audit", "style_guide_enforcement"])

def _exec_bot_setup_competitor_analyst(inp):
    return _exec_bot_setup_generic(inp, "competitor_analyst", [
        "pricing_tracking", "feature_comparison", "review_monitoring", "social_media_tracking",
        "job_posting_analysis", "news_alerts", "market_share_estimation"])

def _exec_bot_setup_appointment_setter(inp):
    return _exec_bot_setup_generic(inp, "appointment_setter", [
        "cold_outreach", "lead_qualification", "meeting_scheduling", "follow_up_sequences",
        "calendar_management", "no_show_handling"])

def _exec_bot_setup_customer_support(inp):
    return _exec_bot_setup_generic(inp, "customer_support", [
        "ticket_handling", "faq_answers", "issue_resolution", "escalation",
        "satisfaction_surveys", "response_templates", "sla_tracking", "multi_channel"])

def _exec_bot_setup_customer_success(inp):
    return _exec_bot_setup_generic(inp, "customer_success", [
        "onboarding", "check_ins", "churn_prediction", "upsell_identification",
        "nps_tracking", "health_scoring", "renewal_management"])

def _exec_bot_setup_hr_manager(inp):
    return _exec_bot_setup_generic(inp, "hr_manager", [
        "handbook_generation", "policy_compliance", "time_off_tracking", "onboarding_checklists",
        "culture_surveys", "performance_reviews", "benefits_management"])

def _exec_bot_setup_recruiter(inp):
    return _exec_bot_setup_generic(inp, "recruiter", [
        "job_posting", "resume_screening", "interview_scheduling", "candidate_scoring",
        "talent_sourcing", "offer_letter_generation", "pipeline_tracking"])

def _exec_bot_setup_training_coach(inp):
    return _exec_bot_setup_generic(inp, "training_coach", [
        "lesson_planning", "skill_assessment", "quiz_generation", "progress_tracking",
        "learning_paths", "certification_tracking", "knowledge_base"])

def _exec_bot_setup_project_manager(inp):
    return _exec_bot_setup_generic(inp, "project_manager", [
        "task_tracking", "deadline_reminders", "status_reports", "resource_allocation",
        "milestone_tracking", "risk_assessment", "team_coordination"])

def _exec_bot_setup_inventory_manager(inp):
    return _exec_bot_setup_generic(inp, "inventory_manager", [
        "stock_tracking", "reorder_alerts", "supplier_management", "demand_forecasting",
        "waste_reduction", "barcode_scanning", "multi_location"])

def _exec_bot_setup_supply_chain(inp):
    return _exec_bot_setup_generic(inp, "supply_chain", [
        "vendor_evaluation", "order_tracking", "logistics_optimization", "cost_analysis",
        "lead_time_tracking", "quality_scoring"])

def _exec_bot_setup_quality_assurance(inp):
    return _exec_bot_setup_generic(inp, "quality_assurance", [
        "checklists", "inspection_scheduling", "defect_tracking", "compliance_auditing",
        "sop_enforcement", "corrective_actions"])

def _exec_bot_setup_it_support(inp):
    return _exec_bot_setup_generic(inp, "it_support", [
        "troubleshooting", "password_resets", "software_setup", "network_diagnostics",
        "security_alerts", "backup_monitoring", "ticket_system"])

def _exec_bot_setup_security_analyst(inp):
    return _exec_bot_setup_generic(inp, "security_analyst", [
        "threat_monitoring", "vulnerability_scanning", "access_audits", "incident_response",
        "compliance_reporting", "phishing_detection", "security_training"])

def _exec_bot_setup_devops(inp):
    return _exec_bot_setup_generic(inp, "devops", [
        "server_monitoring", "deployment_automation", "log_analysis", "uptime_tracking",
        "auto_scaling", "ssl_management", "backup_verification"])

def _exec_bot_setup_data_analyst(inp):
    return _exec_bot_setup_generic(inp, "data_analyst", [
        "dashboards", "trend_analysis", "kpi_tracking", "data_visualization",
        "predictive_insights", "report_generation", "data_cleaning"])

def _exec_bot_setup_graphic_designer(inp):
    return _exec_bot_setup_generic(inp, "graphic_designer", [
        "social_media_graphics", "logo_concepts", "flyers", "banners", "infographics",
        "brand_assets", "presentation_design", "thumbnail_creation"])

def _exec_bot_setup_video_editor(inp):
    return _exec_bot_setup_generic(inp, "video_editor", [
        "clip_trimming", "caption_generation", "thumbnail_creation", "social_cuts",
        "highlight_reels", "intro_outro", "color_grading"])

def _exec_bot_setup_podcast_producer(inp):
    return _exec_bot_setup_generic(inp, "podcast_producer", [
        "episode_planning", "show_notes", "transcription", "audiogram_clips",
        "guest_research", "distribution", "analytics"])

def _exec_bot_setup_real_estate_agent(inp):
    return _exec_bot_setup_generic(inp, "real_estate_agent", [
        "listing_descriptions", "market_analysis", "lead_follow_up", "showing_scheduling",
        "cma_reports", "neighborhood_research", "mortgage_calculators"])

def _exec_bot_setup_insurance_agent(inp):
    return _exec_bot_setup_generic(inp, "insurance_agent", [
        "quote_generation", "policy_comparison", "claims_assistance", "renewal_reminders",
        "risk_assessment", "coverage_recommendations"])

def _exec_bot_setup_medical_assistant(inp):
    return _exec_bot_setup_generic(inp, "medical_assistant", [
        "patient_intake", "appointment_reminders", "insurance_verification",
        "prescription_refills", "lab_results_notification", "hipaa_compliance"])

def _exec_bot_setup_restaurant_manager(inp):
    return _exec_bot_setup_generic(inp, "restaurant_manager", [
        "reservations", "menu_updates", "review_responses", "inventory_tracking",
        "staff_scheduling", "catering_quotes", "delivery_management"])

def _exec_bot_setup_fitness_coach(inp):
    return _exec_bot_setup_generic(inp, "fitness_coach", [
        "workout_plans", "nutrition_tracking", "progress_monitoring", "class_scheduling",
        "member_engagement", "goal_setting", "challenge_creation"])

def _exec_bot_setup_salon_manager(inp):
    return _exec_bot_setup_generic(inp, "salon_manager", [
        "booking", "stylist_matching", "service_recommendations", "loyalty_programs",
        "product_suggestions", "waitlist_management", "review_collection"])

def _exec_bot_setup_ecommerce_manager(inp):
    return _exec_bot_setup_generic(inp, "ecommerce_manager", [
        "product_listings", "price_optimization", "cart_abandonment", "inventory_sync",
        "order_tracking", "review_management", "return_processing"])

def _exec_bot_setup_property_manager(inp):
    return _exec_bot_setup_generic(inp, "property_manager", [
        "tenant_communication", "maintenance_requests", "rent_collection", "lease_management",
        "inspections", "vendor_coordination", "vacancy_marketing"])

def _exec_bot_setup_tutor(inp):
    return _exec_bot_setup_generic(inp, "tutor", [
        "lesson_planning", "student_assessment", "homework_help", "progress_reports",
        "quiz_generation", "study_guides", "parent_updates"])

def _exec_bot_setup_event_planner(inp):
    return _exec_bot_setup_generic(inp, "event_planner", [
        "venue_research", "vendor_coordination", "timeline_creation", "guest_management",
        "budget_tracking", "rsvp_management", "day_of_coordination"])

def _exec_bot_setup_travel_agent(inp):
    return _exec_bot_setup_generic(inp, "travel_agent", [
        "trip_planning", "flight_search", "hotel_search", "itinerary_creation",
        "budget_optimization", "travel_alerts", "visa_requirements"])

def _exec_bot_setup_executive_assistant(inp):
    return _exec_bot_setup_generic(inp, "executive_assistant", [
        "calendar_management", "meeting_prep", "email_triage", "travel_booking",
        "daily_briefings", "document_preparation", "priority_management"])

def _exec_bot_setup_business_consultant(inp):
    return _exec_bot_setup_generic(inp, "business_consultant", [
        "swot_analysis", "market_research", "growth_strategies", "process_optimization",
        "benchmarking", "kpi_definition", "strategic_planning"])

def _exec_bot_setup_compliance_officer(inp):
    return _exec_bot_setup_generic(inp, "compliance_officer", [
        "regulatory_tracking", "policy_auditing", "risk_assessment", "training_requirements",
        "filing_deadlines", "incident_reporting", "documentation"])

def _exec_bot_setup_research_analyst(inp):
    return _exec_bot_setup_generic(inp, "research_analyst", [
        "market_research", "industry_reports", "trend_analysis", "data_collection",
        "competitive_intelligence", "survey_design", "white_papers"])

def _exec_bot_setup_proposal_writer(inp):
    return _exec_bot_setup_generic(inp, "proposal_writer", [
        "rfp_responses", "business_proposals", "pitch_decks", "grant_applications",
        "case_studies", "executive_summaries", "pricing_proposals"])

def _exec_bot_setup_contract_manager(inp):
    return _exec_bot_setup_generic(inp, "contract_manager", [
        "contract_drafting", "renewal_tracking", "clause_analysis", "negotiation_support",
        "expiration_alerts", "version_control", "compliance_verification"])


def _exec_bot_list(inp):
    """List all configured bots."""
    if not BOT_DIR.exists():
        return {"bots": [], "count": 0}

    bots = []
    client_filter = inp.get("client_id")

    for client_dir in BOT_DIR.iterdir():
        if not client_dir.is_dir():
            continue
        if client_filter and client_dir.name != client_filter:
            continue
        for f in client_dir.glob("*.json"):
            with open(f) as fh:
                config = json.load(fh)
            bots.append({
                "client_id": client_dir.name,
                "bot_type": config.get("bot_type"),
                "status": config.get("status"),
                "business_name": config.get("business_name", ""),
                "created_at": config.get("created_at")
            })

    return {"bots": bots, "count": len(bots)}


# ── Payment / Crypto Executors ──

PAYMENTS_DIR = PLATFORM_DIR / "data" / "payments"


def _exec_payment_create_invoice(inp, client_config=None):
    """Create a payment invoice."""
    PAYMENTS_DIR.mkdir(parents=True, exist_ok=True)
    cfg = client_config or {}
    provider = inp.get("provider", "stripe")
    payment_id = f"pay_{int(time.time())}"

    payment = {
        "id": payment_id,
        "amount": inp["amount"],
        "currency": inp.get("currency", "usd"),
        "description": inp["description"],
        "customer_email": inp.get("customer_email", ""),
        "provider": provider,
        "status": "pending",
        "created_at": datetime.now().isoformat()
    }

    # If Stripe configured, create actual payment link
    if provider == "stripe" and cfg.get("stripe_secret_key"):
        try:
            import requests
            resp = requests.post("https://api.stripe.com/v1/payment_links",
                headers={"Authorization": f"Bearer {cfg['stripe_secret_key']}"},
                data={"line_items[0][price_data][currency]": inp.get("currency", "usd"),
                      "line_items[0][price_data][product_data][name]": inp["description"],
                      "line_items[0][price_data][unit_amount]": int(inp["amount"] * 100),
                      "line_items[0][quantity]": 1}, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                payment["payment_url"] = data.get("url", "")
                payment["stripe_id"] = data.get("id", "")
        except Exception as e:
            payment["note"] = f"Stripe API error: {e}. Invoice saved locally."

    filepath = PAYMENTS_DIR / f"{payment_id}.json"
    with open(filepath, "w") as f:
        json.dump(payment, f, indent=2)

    return payment


def _exec_payment_check_status(inp):
    """Check payment status."""
    filepath = PAYMENTS_DIR / f"{inp['payment_id']}.json"
    if not filepath.exists():
        return {"error": f"Payment not found: {inp['payment_id']}"}
    with open(filepath) as f:
        return json.load(f)


def _exec_crypto_wallet_balance(inp):
    """Check crypto wallet balance."""
    import requests
    address = inp["address"]
    network = inp.get("network", "ethereum")

    try:
        if network == "ethereum":
            # Etherscan-like free API
            url = f"https://api.etherscan.io/api?module=account&action=balance&address={address}&tag=latest"
            resp = requests.get(url, timeout=10).json()
            balance_wei = int(resp.get("result", 0))
            return {"address": address, "balance_eth": balance_wei / 1e18, "network": "ethereum"}
        elif network == "solana":
            url = f"https://api.mainnet-beta.solana.com"
            payload = {"jsonrpc": "2.0", "id": 1, "method": "getBalance", "params": [address]}
            resp = requests.post(url, json=payload, timeout=10).json()
            balance = resp.get("result", {}).get("value", 0)
            return {"address": address, "balance_sol": balance / 1e9, "network": "solana"}
        else:
            return {"address": address, "network": network, "note": "Use blockchain explorer for this network"}
    except Exception as e:
        return {"error": str(e)}


def _exec_crypto_send(inp, client_config=None):
    """Send crypto — requires private key in config."""
    cfg = client_config or {}
    private_key = cfg.get("crypto_private_key", "")
    if not private_key:
        return {"error": "Crypto private key not configured. Set crypto_private_key in client config."}

    # This is a placeholder — actual implementation depends on network
    return {
        "note": "Crypto send requires web3.py (Ethereum) or solana-py. Install the appropriate library.",
        "to": inp["to_address"],
        "amount": inp["amount"],
        "token": inp["token"],
        "network": inp.get("network", "ethereum"),
        "status": "requires_library"
    }


def _exec_crypto_price(inp):
    """Get crypto prices."""
    import requests
    tokens = inp["tokens"]

    # Map common symbols to CoinGecko IDs
    id_map = {
        "BTC": "bitcoin", "ETH": "ethereum", "SOL": "solana", "DOGE": "dogecoin",
        "ADA": "cardano", "XRP": "ripple", "DOT": "polkadot", "AVAX": "avalanche-2",
        "MATIC": "matic-network", "LINK": "chainlink", "UNI": "uniswap", "ATOM": "cosmos",
        "LTC": "litecoin", "BNB": "binancecoin", "USDC": "usd-coin", "USDT": "tether"
    }

    ids = [id_map.get(t.upper(), t.lower()) for t in tokens]
    ids_str = ",".join(ids)

    try:
        url = f"https://api.coingecko.com/api/v3/simple/price?ids={ids_str}&vs_currencies=usd&include_24hr_change=true"
        resp = requests.get(url, timeout=10).json()

        prices = {}
        for token, coin_id in zip(tokens, ids):
            if coin_id in resp:
                prices[token.upper()] = {
                    "price_usd": resp[coin_id].get("usd"),
                    "change_24h": resp[coin_id].get("usd_24h_change")
                }

        return {"prices": prices, "count": len(prices)}
    except Exception as e:
        return {"error": str(e)}


# ── Client Management Executors ──

def _exec_client_list(inp):
    """List all clients."""
    clients_dir = PLATFORM_DIR / "data" / "clients"
    if not clients_dir.exists():
        return {"clients": [], "count": 0}

    clients = []
    for f in clients_dir.glob("*.json"):
        if f.name.startswith("_"):
            continue
        try:
            with open(f) as fh:
                cfg = json.load(fh)
            clients.append({
                "client_id": f.stem,
                "business_name": cfg.get("business_name", f.stem),
                "phone": cfg.get("phone", ""),
                "email": cfg.get("email", ""),
                "port": cfg.get("port", ""),
            })
        except Exception:
            pass

    return {"clients": clients, "count": len(clients)}


def _exec_client_add(inp):
    """Add a new client."""
    clients_dir = PLATFORM_DIR / "data" / "clients"
    clients_dir.mkdir(parents=True, exist_ok=True)

    client_config = {
        "client_id": inp["client_id"],
        "business_name": inp["business_name"],
        "phone": inp.get("phone", ""),
        "email": inp.get("email", ""),
        "services": inp.get("services", []),
        "business_hours": inp.get("business_hours", ""),
        "created_at": datetime.now().isoformat()
    }

    filepath = clients_dir / f"{inp['client_id']}.json"
    with open(filepath, "w") as f:
        json.dump(client_config, f, indent=2)

    return {"created": True, "client_id": inp["client_id"], "path": str(filepath)}


def _exec_client_start(inp):
    """Start a client's service."""
    try:
        from .client_manager import start_client
        result = start_client(inp["client_id"])
        return {"started": True, "client_id": inp["client_id"], "details": str(result)}
    except Exception as e:
        return {"error": f"Failed to start client: {str(e)}"}


def _exec_client_stop(inp):
    """Stop a client's service."""
    try:
        from .client_manager import stop_client
        result = stop_client(inp["client_id"])
        return {"stopped": True, "client_id": inp["client_id"], "details": str(result)}
    except Exception as e:
        return {"error": f"Failed to stop client: {str(e)}"}


# ══════════════════════════════════════════════════════════════════════
# UNIFIED EXECUTOR — the ONE function agent_loop calls
# ══════════════════════════════════════════════════════════════════════

def execute_tool(tool_name, tool_input, client_id="default", client_config=None):
    """
    Execute any tool by name. This is the single entry point for the engine.

    Usage in agent_loop:
        from core.tools import get_all_tools, execute_tool
        result = agent_loop(messages, prompt, get_all_tools(),
                           lambda name, inp: execute_tool(name, inp, client_id="abc"))
    """
    try:
        # Check module tools first (web, browser, files)
        if tool_name in _MODULE_TOOLS:
            executor = _MODULE_TOOLS[tool_name]
            # file_manager needs client_id
            if tool_name.startswith("file_"):
                return executor(tool_name, tool_input, client_id=client_id)
            return executor(tool_name, tool_input)

        # Built-in tools
        if tool_name == "email_send":
            return json.dumps(_exec_email_send(tool_input, client_config))
        elif tool_name == "email_read":
            return json.dumps(_exec_email_read(tool_input, client_config))
        elif tool_name == "image_generate":
            return json.dumps(_exec_image_generate(tool_input))
        elif tool_name == "code_execute":
            return json.dumps(_exec_code_execute(tool_input))
        elif tool_name == "shell_execute":
            return json.dumps(_exec_shell_execute(tool_input))
        elif tool_name == "voice_tts":
            return json.dumps(_exec_voice_tts(tool_input))
        elif tool_name == "voice_stt":
            return json.dumps(_exec_voice_stt(tool_input))
        elif tool_name == "http_request":
            return json.dumps(_exec_http_request(tool_input))
        elif tool_name == "schedule_task":
            return json.dumps(_exec_schedule_task(tool_input))
        elif tool_name == "schedule_list":
            return json.dumps(_exec_schedule_list(tool_input))
        elif tool_name == "schedule_cancel":
            return json.dumps(_exec_schedule_cancel(tool_input))
        elif tool_name == "memory_read":
            return json.dumps(_exec_memory_read(tool_input, client_id))
        elif tool_name == "memory_write":
            return json.dumps(_exec_memory_write(tool_input, client_id))
        elif tool_name == "calculate":
            return json.dumps(_exec_calculate(tool_input))
        elif tool_name == "get_datetime":
            return json.dumps(_exec_get_datetime(tool_input))
        # Document parsing
        elif tool_name == "document_read_pdf":
            return json.dumps(_exec_document_read_pdf(tool_input))
        elif tool_name == "document_read_docx":
            return json.dumps(_exec_document_read_docx(tool_input))
        elif tool_name == "document_read_csv":
            return json.dumps(_exec_document_read_csv(tool_input))
        elif tool_name == "document_ocr":
            return json.dumps(_exec_document_ocr(tool_input))
        # Messaging
        elif tool_name == "telegram_send":
            return json.dumps(_exec_telegram_send(tool_input, client_config))
        elif tool_name == "discord_send":
            return json.dumps(_exec_discord_send(tool_input))
        elif tool_name == "sms_send":
            return json.dumps(_exec_sms_send(tool_input, client_config))
        # VPS management
        elif tool_name == "vps_execute":
            return json.dumps(_exec_vps_execute(tool_input, client_config))
        elif tool_name == "vps_upload":
            return json.dumps(_exec_vps_upload(tool_input, client_config))
        elif tool_name == "vps_download":
            return json.dumps(_exec_vps_download(tool_input, client_config))
        # Git
        elif tool_name == "git_execute":
            return json.dumps(_exec_git_execute(tool_input))
        # Skills / Plugins
        elif tool_name == "skill_list":
            return json.dumps(_exec_skill_list(tool_input))
        elif tool_name == "skill_load":
            return json.dumps(_exec_skill_load(tool_input))
        elif tool_name == "skill_create":
            return json.dumps(_exec_skill_create(tool_input))
        # Knowledge
        elif tool_name == "knowledge_save":
            return json.dumps(_exec_knowledge_save(tool_input))
        elif tool_name == "knowledge_search":
            return json.dumps(_exec_knowledge_search(tool_input))
        elif tool_name == "knowledge_list":
            return json.dumps(_exec_knowledge_list(tool_input))
        # Account automation
        elif tool_name == "browser_signup":
            return json.dumps(_exec_browser_signup(tool_input))
        elif tool_name == "browser_login":
            return json.dumps(_exec_browser_login(tool_input))
        elif tool_name == "browser_interactive":
            return json.dumps(_exec_browser_interactive(tool_input))
        # Data intelligence
        elif tool_name == "scrape_google":
            return json.dumps(_exec_scrape_google(tool_input))
        elif tool_name == "scrape_social":
            return json.dumps(_exec_scrape_social(tool_input))
        elif tool_name == "scrape_business_info":
            return json.dumps(_exec_scrape_business_info(tool_input))
        # Lead generation
        elif tool_name == "lead_find":
            return json.dumps(_exec_lead_find(tool_input))
        elif tool_name == "lead_enrich":
            return json.dumps(_exec_lead_enrich(tool_input))
        # Notifications
        elif tool_name == "notify":
            return json.dumps(_exec_notify(tool_input, client_config))
        # JSON
        elif tool_name == "json_transform":
            return json.dumps(_exec_json_transform(tool_input))
        # QR
        elif tool_name == "qr_generate":
            return json.dumps(_exec_qr_generate(tool_input))
        # Desktop screenshot
        elif tool_name == "screenshot_desktop":
            return json.dumps(_exec_screenshot_desktop(tool_input))
        # PDF creation
        elif tool_name == "pdf_create":
            return json.dumps(_exec_pdf_create(tool_input))
        # Clipboard
        elif tool_name == "clipboard_read":
            return json.dumps(_exec_clipboard_read(tool_input))
        elif tool_name == "clipboard_write":
            return json.dumps(_exec_clipboard_write(tool_input))
        # Processes
        elif tool_name == "process_list":
            return json.dumps(_exec_process_list(tool_input))
        elif tool_name == "process_kill":
            return json.dumps(_exec_process_kill(tool_input))
        # System info
        elif tool_name == "system_info":
            return json.dumps(_exec_system_info(tool_input))
        # Wait
        elif tool_name == "wait":
            return json.dumps(_exec_wait(tool_input))
        # Video / Camera
        elif tool_name == "camera_capture":
            return json.dumps(_exec_camera_capture(tool_input))
        elif tool_name == "video_record":
            return json.dumps(_exec_video_record(tool_input))
        elif tool_name == "video_stream_start":
            return json.dumps(_exec_video_stream_start(tool_input))
        elif tool_name == "video_stream_stop":
            return json.dumps(_exec_video_stream_stop(tool_input))
        elif tool_name == "screen_record":
            return json.dumps(_exec_screen_record(tool_input))
        # Audio
        elif tool_name == "audio_record":
            return json.dumps(_exec_audio_record(tool_input))
        elif tool_name == "audio_play":
            return json.dumps(_exec_audio_play(tool_input))
        # Compression
        elif tool_name == "zip_create":
            return json.dumps(_exec_zip_create(tool_input))
        elif tool_name == "zip_extract":
            return json.dumps(_exec_zip_extract(tool_input))
        # Translation
        elif tool_name == "translate":
            return json.dumps(_exec_translate(tool_input))
        # Weather
        elif tool_name == "weather_get":
            return json.dumps(_exec_weather_get(tool_input))
        # News
        elif tool_name == "news_get":
            return json.dumps(_exec_news_get(tool_input))
        # Finance
        elif tool_name == "stock_price":
            return json.dumps(_exec_stock_price(tool_input))
        # Whois
        elif tool_name == "whois_lookup":
            return json.dumps(_exec_whois_lookup(tool_input))
        # DNS
        elif tool_name == "dns_lookup":
            return json.dumps(_exec_dns_lookup(tool_input))
        # Hash
        elif tool_name == "hash_text":
            return json.dumps(_exec_hash_text(tool_input))
        # Encrypt/Decrypt
        elif tool_name == "encrypt_text":
            return json.dumps(_exec_encrypt_text(tool_input))
        elif tool_name == "decrypt_text":
            return json.dumps(_exec_decrypt_text(tool_input))
        # Base64
        elif tool_name == "base64_encode":
            return json.dumps(_exec_base64_encode(tool_input))
        elif tool_name == "base64_decode":
            return json.dumps(_exec_base64_decode(tool_input))
        # AI tools
        elif tool_name == "ai_ask":
            return json.dumps(_exec_ai_ask(tool_input))
        elif tool_name == "ai_summarize":
            return json.dumps(_exec_ai_summarize(tool_input))
        elif tool_name == "ai_classify":
            return json.dumps(_exec_ai_classify(tool_input))
        elif tool_name == "ai_extract":
            return json.dumps(_exec_ai_extract(tool_input))
        elif tool_name == "ai_generate_code":
            return json.dumps(_exec_ai_generate_code(tool_input))
        # Workflow
        elif tool_name == "workflow_create":
            return json.dumps(_exec_workflow_create(tool_input))
        elif tool_name == "workflow_run":
            return json.dumps(_exec_workflow_run(tool_input, client_id, client_config))
        elif tool_name == "workflow_list":
            return json.dumps(_exec_workflow_list(tool_input))
        # Database
        elif tool_name == "db_query":
            return json.dumps(_exec_db_query(tool_input))
        elif tool_name == "db_tables":
            return json.dumps(_exec_db_tables(tool_input))
        # Spreadsheet
        elif tool_name == "spreadsheet_read":
            return json.dumps(_exec_spreadsheet_read(tool_input))
        elif tool_name == "spreadsheet_write":
            return json.dumps(_exec_spreadsheet_write(tool_input))
        # Contacts
        elif tool_name == "contact_add":
            return json.dumps(_exec_contact_add(tool_input))
        elif tool_name == "contact_search":
            return json.dumps(_exec_contact_search(tool_input))
        elif tool_name == "contact_list":
            return json.dumps(_exec_contact_list(tool_input))
        # Invoice
        elif tool_name == "invoice_create":
            return json.dumps(_exec_invoice_create(tool_input))
        # Report
        elif tool_name == "report_generate":
            return json.dumps(_exec_report_generate(tool_input))
        # Appointments
        elif tool_name == "appointment_book":
            return json.dumps(_exec_appointment_book(tool_input))
        elif tool_name == "appointment_list":
            return json.dumps(_exec_appointment_list(tool_input))
        elif tool_name == "appointment_cancel":
            return json.dumps(_exec_appointment_cancel(tool_input))
        # Templates
        elif tool_name == "template_list":
            return json.dumps(_exec_template_list(tool_input))
        elif tool_name == "template_render":
            return json.dumps(_exec_template_render(tool_input))
        elif tool_name == "template_create":
            return json.dumps(_exec_template_create(tool_input))
        # Monitoring
        elif tool_name == "monitor_url":
            return json.dumps(_exec_monitor_url(tool_input))
        elif tool_name == "monitor_port":
            return json.dumps(_exec_monitor_port(tool_input))
        # Text processing
        elif tool_name == "regex_match":
            return json.dumps(_exec_regex_match(tool_input))
        elif tool_name == "text_diff":
            return json.dumps(_exec_text_diff(tool_input))
        # Image edit
        elif tool_name == "image_resize":
            return json.dumps(_exec_image_resize(tool_input))
        elif tool_name == "image_text_overlay":
            return json.dumps(_exec_image_text_overlay(tool_input))
        elif tool_name == "image_convert":
            return json.dumps(_exec_image_convert(tool_input))
        # Multi-agent
        elif tool_name == "agent_spawn":
            return json.dumps(_exec_agent_spawn(tool_input, client_id, client_config))
        elif tool_name == "agent_status":
            return json.dumps(_exec_agent_status(tool_input))
        elif tool_name == "agent_result":
            return json.dumps(_exec_agent_result(tool_input))
        elif tool_name == "agent_stop":
            return json.dumps(_exec_agent_stop(tool_input))
        elif tool_name == "agent_list":
            return json.dumps(_exec_agent_list(tool_input))
        elif tool_name == "agent_message":
            return json.dumps(_exec_agent_message(tool_input))
        elif tool_name == "agent_team":
            return json.dumps(_exec_agent_team(tool_input, client_id, client_config))
        # Payments / Crypto
        elif tool_name == "payment_create_invoice":
            return json.dumps(_exec_payment_create_invoice(tool_input, client_config))
        elif tool_name == "payment_check_status":
            return json.dumps(_exec_payment_check_status(tool_input))
        elif tool_name == "crypto_wallet_balance":
            return json.dumps(_exec_crypto_wallet_balance(tool_input))
        elif tool_name == "crypto_send":
            return json.dumps(_exec_crypto_send(tool_input, client_config))
        elif tool_name == "crypto_price":
            return json.dumps(_exec_crypto_price(tool_input))
        # Bot setup — dynamic routing for all bot_setup_* tools
        elif tool_name.startswith("bot_setup_"):
            func_name = f"_exec_{tool_name}"
            executor = globals().get(func_name)
            if executor:
                return json.dumps(executor(tool_input))
            return json.dumps({"error": f"No executor for {tool_name}"})
        elif tool_name == "bot_list":
            return json.dumps(_exec_bot_list(tool_input))
        # Client management
        elif tool_name == "client_list":
            return json.dumps(_exec_client_list(tool_input))
        elif tool_name == "client_add":
            return json.dumps(_exec_client_add(tool_input))
        elif tool_name == "client_start":
            return json.dumps(_exec_client_start(tool_input))
        elif tool_name == "client_stop":
            return json.dumps(_exec_client_stop(tool_input))
        else:
            return json.dumps({"error": f"Unknown tool: {tool_name}"})

    except Exception as e:
        return json.dumps({"error": f"Tool execution failed: {tool_name} — {str(e)}"})


# ══════════════════════════════════════════════════════════════════════
# CONVENIENCE — run a full agent task with all tools
# ══════════════════════════════════════════════════════════════════════

def run_agent_task(task_prompt, system_prompt=None, client_id="default", client_config=None, max_turns=15):
    """
    Run a complete agent task with all tools available.
    This is the highest-level function — give it a task, get back a result.

    Example:
        result = run_agent_task("Search the web for AI news and email me a summary")
    """
    from .engine import agent_loop, build_system_prompt

    if not system_prompt:
        system_prompt = build_system_prompt("general")

    messages = [{"role": "user", "content": task_prompt}]
    tools = get_all_tools()

    def tool_executor(name, inp):
        return execute_tool(name, inp, client_id=client_id, client_config=client_config)

    return agent_loop(messages, system_prompt, tools, tool_executor, max_turns=max_turns)


# ══════════════════════════════════════════════════════════════════════
# INFO — for dashboard display
# ══════════════════════════════════════════════════════════════════════

def get_tool_count():
    """Return total number of available tools."""
    return len(get_all_tools())


def get_tool_names():
    """Return list of all tool names."""
    return [t["name"] for t in get_all_tools()]


def get_tool_summary():
    """Return a summary for dashboard display."""
    categories = get_tools_by_category()
    return {
        "total_tools": get_tool_count(),
        "categories": {name: len(tools) for name, tools in categories.items()},
        "tool_names": get_tool_names()
    }


if __name__ == "__main__":
    # Print tool inventory
    summary = get_tool_summary()
    print(f"\nJanovum Orchestrator — {summary['total_tools']} tools available\n")
    for cat, count in summary["categories"].items():
        print(f"  {cat}: {count} tools")
    print(f"\nAll tools: {', '.join(summary['tool_names'])}")
