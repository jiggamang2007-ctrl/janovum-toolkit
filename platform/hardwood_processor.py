"""
International Hardwoods — Takeoff Processor
Reads architectural PDFs, extracts all wood areas, generates Excel takeoff spreadsheet.

Usage:
  from hardwood_processor import process_permit_set
  result = process_permit_set(zip_path, project_name, address, claude_api_key)
"""

import os
import io
import json
import zipfile
import re
import requests
import tempfile
from datetime import datetime

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

CLAUDE_API_URL = "https://api.anthropic.com/v1/messages"

# PDF sheets we care about (by filename keywords)
RELEVANT_SHEETS = {
    "ffp_floor1": ["A1-2.01", "FIRST FLOOR - FFP", "FFP.01"],
    "ffp_floor2": ["A1-2.02", "SECOND FLOOR - FFP", "FFP.02"],
    "rcp_floor1": ["A1-3.02", "FIRST FLOOR - RCP", "RCP.02"],
    "rcp_floor2": ["A1-3.03", "SECOND FLOOR - RCP", "RCP.03"],
    "floor_schedule": ["A5-2.03", "FLOOR SCHEDULE"],
    "wall_schedule1": ["A5-2.01", "WALL SCHEDULE - FIRST"],
    "wall_schedule2": ["A5-2.02", "WALL SCHEDULE - SECOND"],
    "panel_schedule": ["A6-1.01", "D&W SCHEDULES"],
    "panel_map1": ["A6-1.02", "D&W MAP PLAN - FIRST"],
    "panel_map2": ["A6-1.03", "D&W MAP PLAN - SECOND"],
}


def _extract_pdf_text(pdf_bytes):
    """Extract all text from a PDF given as bytes."""
    if not HAS_PDF:
        return ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            parts = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                parts.append(text)
                # Also try tables
                for tbl in page.extract_tables():
                    for row in tbl:
                        row_text = " | ".join(str(c) for c in row if c)
                        if row_text.strip():
                            parts.append(row_text)
            return "\n".join(parts)
    except Exception as e:
        return f"[PDF read error: {e}]"


def _find_pdf_in_zip(zf, keywords):
    """Find a PDF in a zip file matching any of the keywords."""
    for name in zf.namelist():
        if not name.lower().endswith(".pdf"):
            continue
        basename = os.path.basename(name).upper()
        for kw in keywords:
            if kw.upper() in basename:
                return name
    return None


def extract_texts_from_zip(zip_path):
    """Extract text from all relevant PDFs in a zip file."""
    texts = {}
    with zipfile.ZipFile(zip_path, "r") as zf:
        for key, keywords in RELEVANT_SHEETS.items():
            match = _find_pdf_in_zip(zf, keywords)
            if match:
                pdf_bytes = zf.read(match)
                texts[key] = {
                    "filename": os.path.basename(match),
                    "text": _extract_pdf_text(pdf_bytes),
                }
    return texts


def extract_texts_from_folder(folder_path):
    """Extract text from PDFs in a flat folder."""
    texts = {}
    all_pdfs = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
    for key, keywords in RELEVANT_SHEETS.items():
        for fname in all_pdfs:
            if any(kw.upper() in fname.upper() for kw in keywords):
                path = os.path.join(folder_path, fname)
                with open(path, "rb") as f:
                    texts[key] = {
                        "filename": fname,
                        "text": _extract_pdf_text(f.read()),
                    }
                break
    return texts


def _call_claude(prompt, system, api_key):
    """Call Claude API."""
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    body = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 4096,
        "system": system,
        "messages": [{"role": "user", "content": prompt}],
    }
    resp = requests.post(CLAUDE_API_URL, headers=headers, json=body, timeout=120)
    data = resp.json()
    if resp.status_code != 200:
        raise Exception(data.get("error", {}).get("message", f"API error {resp.status_code}"))
    return data["content"][0]["text"]


ANALYSIS_SYSTEM = """You are an expert construction estimator specializing in hardwood and wood finish takeoffs.
You analyze architectural PDF text (extracted via OCR/text parser) and identify every area that requires wood treatment.

Wood areas include:
1. WOOD FLOORS: Rooms labeled IF-04 (interior wood), EF-02A or EF-02B (exterior wood decks)
2. WOOD CEILINGS: Rooms in Reflected Ceiling Plans (RCP) with wood slat soffits or wood panel ceilings
3. WOOD WALLS: Rooms with wood paneling, wood slats, or paginated wooden wall finishes
4. SITE WOOD: Fences, gates, or exterior slatted wood features

For each room/area you find, extract:
- Room name (e.g. "Living Room", "Master Bedroom")
- Level (Level 01 = First Floor, Level 02 = Second Floor)
- Area in square feet (labeled "A:" or "sq ft" in the text)
- Wood type/finish if mentioned

Return ONLY valid JSON in exactly this structure:
{
  "ceiling_exterior": [{"room": "Exterior", "level": "01", "sf": 4349, "material": "American Oak panels 2\" slats"}],
  "ceiling_interior": [{"room": "Living Room", "level": "01", "sf": 1048, "material": "American Oak panels 2\" slats"}],
  "floor_exterior": [{"room": "Lower Deck", "level": "01", "sf": 781, "material": "EF-02A Planks 8\" x 3/4\""}],
  "floor_interior": [{"room": "Master Bedroom", "level": "02", "sf": 585, "material": "IF-04 Planks 8\" x 3/4\""}],
  "walls": [{"room": "Corridor", "level": "01", "sf": 924, "material": "Paginated Wood 3'x2\" 10'H Light Oak"}],
  "site": [{"room": "Fence & Gates 4'-0\" High", "level": "site", "sf": 408, "material": "1\"x1\" Slatted Wood"}],
  "notes": ["any important notes about the scope"]
}

Be thorough — include every room that has ANY wood treatment.
If you can't determine exact SF, use 0 and note it.
"""


def analyze_with_claude(texts, project_name, address, api_key):
    """Send extracted PDF text to Claude for analysis."""
    # Build a combined prompt with all available text
    sections = []
    for key, data in texts.items():
        if data.get("text"):
            # Truncate very long texts to avoid token limits
            text = data["text"][:8000]
            sections.append(f"=== {key.upper()} ({data['filename']}) ===\n{text}")

    combined = "\n\n".join(sections)

    prompt = f"""Project: {project_name}
Address: {address}

I have extracted text from the following architectural PDF sheets for this project.
Please analyze all of them and identify every wood scope area.

{combined}

Return the complete wood takeoff as JSON."""

    return _call_claude(prompt, ANALYSIS_SYSTEM, api_key)


def _parse_claude_response(response_text):
    """Extract JSON from Claude's response."""
    # Try to find JSON block
    match = re.search(r'\{[\s\S]*\}', response_text)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return None


def generate_excel(takeoff_data, project_name, address, output_path=None):
    """Generate the Excel takeoff spreadsheet matching the IH format."""
    if not HAS_XLSX:
        raise ImportError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MATERIAL"

    # Color scheme
    GOLD = "C9A84C"
    DARK = "1A1A1A"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"
    SECTION_BG = "2C2C2C"
    SUB_BG = "3D3D3D"

    def set_cell(row, col, value, bold=False, bg=None, font_color="000000", size=10, wrap=False, align="left", formula=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, size=size, color=font_color, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        return cell

    def border_cell(row, col, sides="all"):
        thin = Side(style="thin", color="CCCCCC")
        thick = Side(style="medium", color="888888")
        cell = ws.cell(row=row, column=col)
        b = Border(
            left=thin if "l" in sides or sides == "all" else None,
            right=thin if "r" in sides or sides == "all" else None,
            top=thin if "t" in sides or sides == "all" else None,
            bottom=thin if "b" in sides or sides == "all" else None,
        )
        cell.border = b

    # ── Column widths ──
    col_widths = [40, 12, 12, 12, 14, 12, 16, 12, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Title block ──
    ws.merge_cells(f"A{row}:J{row}")
    set_cell(row, 1, "WOOD SCOPE OF WORK", bold=True, bg=DARK, font_color=GOLD, size=14, align="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:J{row}")
    set_cell(row, 1, project_name, bold=True, bg=DARK, font_color="FFFFFF", size=12, align="center")
    row += 1

    ws.merge_cells(f"A{row}:J{row}")
    set_cell(row, 1, address, bg=DARK, font_color="AAAAAA", size=10, align="center")
    ws.row_dimensions[row].height = 18
    row += 1

    ws.merge_cells(f"A{row}:J{row}")
    set_cell(row, 1, f"Generated: {datetime.now().strftime('%B %d, %Y')}", bg=DARK, font_color="666666", size=9, align="center")
    row += 2

    # ── Header row ──
    headers = ["DESCRIPTION", "Takeoff", "Waste (10%)", "Total", "Unit", "Unit Cost", "Material Cost", "Unit Cost", "Labor Total", "Material & Labor Total"]
    sub_headers = ["", "(SF)", "", "(SF)", "", "($)", "($)", "($)", "($)", "($)"]
    for i, h in enumerate(headers, 1):
        set_cell(row, i, h, bold=True, bg=SECTION_BG, font_color=GOLD, size=9, align="center", wrap=True)
    ws.row_dimensions[row].height = 32
    row += 1
    for i, h in enumerate(sub_headers, 1):
        set_cell(row, i, h, bold=False, bg=SUB_BG, font_color="AAAAAA", size=8, align="center")
    ws.row_dimensions[row].height = 14
    header_row = row
    row += 1

    def section_header(label):
        nonlocal row
        ws.merge_cells(f"A{row}:J{row}")
        set_cell(row, 1, label, bold=True, bg="1A3A5C", font_color="FFFFFF", size=10)
        ws.row_dimensions[row].height = 20
        row += 1

    def sub_section(label):
        nonlocal row
        ws.merge_cells(f"A{row}:J{row}")
        set_cell(row, 1, f"  {label}", bold=True, bg="2A2A2A", font_color="DDDDDD", size=9)
        ws.row_dimensions[row].height = 16
        row += 1

    def note_row(text):
        nonlocal row
        ws.merge_cells(f"A{row}:J{row}")
        set_cell(row, 1, f"    ↳ {text}", bold=False, bg=LIGHT_GRAY, font_color="888888", size=8, wrap=True)
        ws.row_dimensions[row].height = 13
        row += 1

    first_data_row = None
    data_rows = []

    def data_row(description, sf, material=""):
        nonlocal row, first_data_row
        bg = WHITE if row % 2 == 0 else "F7F7F7"
        sf_val = sf if isinstance(sf, (int, float)) and sf > 0 else None

        set_cell(row, 1, f"    {description}", bg=bg, size=9)
        ws.cell(row=row, column=2).value = sf_val
        ws.cell(row=row, column=2).number_format = "#,##0"
        ws.cell(row=row, column=2).font = Font(bold=True, size=9, name="Calibri")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=bg)

        # Waste = 10% of takeoff
        waste_cell = ws.cell(row=row, column=3)
        waste_cell.value = f"=0.1*B{row}" if sf_val else None
        waste_cell.number_format = "#,##0"
        waste_cell.font = Font(size=9, name="Calibri")
        waste_cell.alignment = Alignment(horizontal="center", vertical="center")
        waste_cell.fill = PatternFill("solid", fgColor=bg)

        # Total = takeoff + waste
        total_cell = ws.cell(row=row, column=4)
        total_cell.value = f"=C{row}+B{row}" if sf_val else None
        total_cell.number_format = "#,##0"
        total_cell.font = Font(bold=True, size=9, name="Calibri")
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.fill = PatternFill("solid", fgColor=bg)

        # Unit
        set_cell(row, 5, "Square Foot", bg=bg, size=8, align="center")

        # Unit Cost (blank - client fills)
        for col in [6, 7, 8, 9, 10]:
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=bg)
            if col == 7:
                c.value = f"=D{row}*F{row}" if sf_val else None
                c.number_format = '"$"#,##0.00'
            elif col == 9:
                c.value = f"=H{row}*B{row}" if sf_val else None
                c.number_format = '"$"#,##0.00'
            elif col == 10:
                c.value = f"=I{row}+G{row}" if sf_val else None
                c.number_format = '"$"#,##0.00'
                c.font = Font(bold=True, size=9, name="Calibri", color="1A3A5C")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9, name="Calibri")

        data_rows.append(row)
        ws.row_dimensions[row].height = 16
        row += 1

    def empty_row():
        nonlocal row
        ws.row_dimensions[row].height = 8
        row += 1

    # ════════════════════════════════════════
    # SECTION 1: CEILINGS
    # ════════════════════════════════════════
    section_header("CEILINGS — WOOD SLATS / SOFFITS")
    note_row("Refer to sheets A1-3.04 and A1-3.05 for wood soffit details")
    note_row("Refer to sheets A1-3.02 and A1-3.03 for locations")

    # Exterior ceilings
    sub_section("Exterior — Wooden American Oak panels with 2\" slats")
    ceiling_ext = takeoff_data.get("ceiling_exterior", [])
    if ceiling_ext:
        for item in ceiling_ext:
            data_row(f"Level {item.get('level','01').zfill(2).replace('0','0')}:{item['room']}", item.get("sf", 0))
    else:
        data_row("Level 01:Exterior", 0)
    empty_row()

    # Interior ceilings
    sub_section("Interior — Wooden American Oak panels with 2\" slats")
    ceiling_int = takeoff_data.get("ceiling_interior", [])
    if ceiling_int:
        for item in ceiling_int:
            label = f"Level {str(item.get('level','01')).zfill(2)}:{item['room']}"
            data_row(label, item.get("sf", 0))
    else:
        data_row("Level 01:Living Room", 0)
    empty_row()

    # ════════════════════════════════════════
    # SECTION 2: FLOORING
    # ════════════════════════════════════════
    section_header("FLOORING — WOOD FLOORS")
    note_row("Refer to sheet A5-2.03 for locations and flooring assembly types")
    note_row("Refer to sheets A1-2.01 and A1-2.02 for finish floor plans")

    # Exterior wood floors
    sub_section("Exterior — EF-02A Planks 8\" x 3/4\" (Decks & Verandas)")
    floor_ext = takeoff_data.get("floor_exterior", [])
    if floor_ext:
        for item in floor_ext:
            label = f"Level {str(item.get('level','01')).zfill(2)}:{item['room']}"
            data_row(label, item.get("sf", 0))
    else:
        data_row("Level 01:Deck", 0)
    empty_row()

    # Interior wood floors
    sub_section("Interior — IF-04 Planks 8\" x 3/4\" Solid Hardwood")
    floor_int = takeoff_data.get("floor_interior", [])
    if floor_int:
        for item in floor_int:
            label = f"Level {str(item.get('level','02')).zfill(2)}:{item['room']}"
            data_row(label, item.get("sf", 0))
    else:
        data_row("Level 02:Master Bedroom", 0)
    empty_row()

    # ════════════════════════════════════════
    # SECTION 3: WALLS
    # ════════════════════════════════════════
    section_header("WALLS — WOOD PANELING")
    note_row("Refer to sheets A5-2.01 and A5-2.02 for wall assembly schedules")
    note_row("Refer to sheet A6-1.01 for the wood panels schedule")
    note_row("Refer to sheets A6-1.02 and A6-1.03 for wood panel locations")

    walls = takeoff_data.get("walls", [])
    # Group walls by material/finish
    wall_groups = {}
    for item in walls:
        mat = item.get("material", "Paginated Wooden")
        if mat not in wall_groups:
            wall_groups[mat] = []
        wall_groups[mat].append(item)

    if wall_groups:
        for mat, items in wall_groups.items():
            sub_section(mat)
            for item in items:
                label = f"Level {str(item.get('level','01')).zfill(2)}:{item['room']}"
                data_row(label, item.get("sf", 0))
    else:
        sub_section("Paginated Wooden 3' x 2\" Thick — Light Oak Finish")
        data_row("Level 01:Living Room", 0)
    empty_row()

    # ════════════════════════════════════════
    # SECTION 4: SITE
    # ════════════════════════════════════════
    section_header("SITE — FENCES & GATES")
    note_row("Refer to sheet A5-1.08 for fence and gate location and details")

    site_items = takeoff_data.get("site", [])
    if site_items:
        for item in site_items:
            data_row(item["room"], item.get("sf", 0))
    else:
        sub_section("Fence & Gates — 4'-0\" High")
        data_row("1\" x 1\" Slatted Wood", 0)
    empty_row()

    # ════════════════════════════════════════
    # TOTALS
    # ════════════════════════════════════════
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(row, 1, "TOTAL", bold=True, bg=DARK, font_color=GOLD, size=11)
    for col in [6, 7, 8, 9, 10]:
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=DARK)
    if data_rows:
        first_dr = data_rows[0]
        last_dr = data_rows[-1]
        for col, letter in [(2, "B"), (4, "D"), (7, "G"), (9, "I"), (10, "J")]:
            c = ws.cell(row=row, column=col)
            c.value = f"=SUM({letter}{first_dr}:{letter}{last_dr})"
            c.number_format = '"$"#,##0' if col > 4 else "#,##0"
            c.font = Font(bold=True, size=11, color=GOLD if col in [2,4] else "00C853", name="Calibri")
            c.fill = PatternFill("solid", fgColor=DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    # ════════════════════════════════════════
    # QUALIFICATIONS
    # ════════════════════════════════════════
    ws.merge_cells(f"A{row}:J{row}")
    set_cell(row, 1, "QUALIFICATIONS / EXCLUSIONS", bold=True, bg="1A3A5C", font_color="FFFFFF", size=10)
    ws.row_dimensions[row].height = 20
    row += 1

    quals = takeoff_data.get("notes", []) or [
        "Trellis structure by others",
        "Canopy structure by others",
        "Aluminum louvers excluded",
        "Substrate for exterior siding by others",
        "Site fence and gates — metal structure excluded",
        "Bison adjustable pedestal system or similar excluded",
        "Verify all measurements on site before ordering",
        "Finishes and material specifications to be confirmed with client",
    ]
    for q in quals:
        ws.merge_cells(f"A{row}:J{row}")
        set_cell(row, 1, f"  • {q}", bg=LIGHT_GRAY, size=9, font_color="444444")
        ws.row_dimensions[row].height = 14
        row += 1

    # Save
    if output_path is None:
        output_path = tempfile.mktemp(suffix=".xlsx")
    wb.save(output_path)
    return output_path


def process_permit_set(zip_path_or_folder, project_name, address, api_key, output_dir=None):
    """
    Full pipeline: ZIP/folder → Excel takeoff.
    Returns: {"excel_path": "...", "data": {...}, "error": None}
    """
    try:
        # 1. Extract text from PDFs
        if os.path.isdir(zip_path_or_folder):
            texts = extract_texts_from_folder(zip_path_or_folder)
        else:
            texts = extract_texts_from_zip(zip_path_or_folder)

        if not texts:
            return {"error": "No relevant architectural sheets found in the files. Expected: FFP, RCP, or Floor Schedule PDFs."}

        # 2. Ask Claude to analyze
        claude_response = analyze_with_claude(texts, project_name, address, api_key)
        takeoff_data = _parse_claude_response(claude_response)

        if not takeoff_data:
            # Claude couldn't parse — return basic structure for manual fill
            takeoff_data = {
                "ceiling_exterior": [], "ceiling_interior": [],
                "floor_exterior": [], "floor_interior": [],
                "walls": [], "site": [],
                "notes": ["Could not auto-extract from PDFs — please verify measurements manually."]
            }

        # 3. Generate Excel
        if output_dir:
            safe_name = re.sub(r'[^a-zA-Z0-9_\- ]', '', project_name).strip().replace(' ', '_')
            output_path = os.path.join(output_dir, f"{safe_name}_Takeoff.xlsx")
        else:
            output_path = None

        excel_path = generate_excel(takeoff_data, project_name, address, output_path)

        return {
            "excel_path": excel_path,
            "data": takeoff_data,
            "sheets_processed": list(texts.keys()),
            "error": None
        }

    except Exception as e:
        return {"error": str(e), "excel_path": None, "data": None}


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 4:
        print("Usage: python hardwood_processor.py <zip_or_folder> <project_name> <address>")
        sys.exit(1)

    # Load API key from env or toolkit config
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        try:
            cfg_path = os.path.join(os.path.dirname(__file__), "data", "config.json")
            with open(cfg_path) as f:
                api_key = json.load(f).get("api_key", "")
        except Exception:
            pass

    if not api_key:
        print("Error: No Claude API key found. Set ANTHROPIC_API_KEY env var.")
        sys.exit(1)

    result = process_permit_set(sys.argv[1], sys.argv[2], sys.argv[3], api_key, output_dir=".")
    if result["error"]:
        print(f"Error: {result['error']}")
    else:
        print(f"Success! Excel saved to: {result['excel_path']}")
        print(f"Sheets processed: {result['sheets_processed']}")
